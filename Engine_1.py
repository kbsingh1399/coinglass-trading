# C:\Users\SIGMA\Documents\Project - Coinglass Trading\Engine_1\Engine_1.py
# Production-Grade Coinglass + Binance Footprint Scraper Terminal
# Built from scratch - fully modular, clean, and robust.

from __future__ import annotations
import os
import sys
import os

# Enable instant unbuffered stdout/stderr flushing for real-time live terminal logs
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass
if hasattr(sys.stderr, 'reconfigure'):
    try:
        sys.stderr.reconfigure(line_buffering=True)
    except Exception:
        pass

# Force-enable Windows ANSI Virtual Terminal Processing on module load
if sys.platform == "win32":
    try:
        import ctypes
        from ctypes import wintypes
        kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
        STD_OUTPUT_HANDLE = wintypes.DWORD(-11)
        ENABLE_VIRTUAL_TERMINAL_PROCESSING = 0x0004
        ENABLE_PROCESSED_OUTPUT = 0x0001
        h_stdout = kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
        mode_val = wintypes.DWORD()
        if kernel32.GetConsoleMode(h_stdout, ctypes.byref(mode_val)):
            kernel32.SetConsoleMode(h_stdout, mode_val.value | ENABLE_VIRTUAL_TERMINAL_PROCESSING | ENABLE_PROCESSED_OUTPUT)
    except Exception:
        pass

import time
from datetime import datetime
import json
import asyncio
import signal
import collections
import dataclasses
import threading
import math
from typing import Dict, Any, List, Optional
from dotenv import load_dotenv

import aiohttp
import websockets
import socket
import re
from playwright.async_api import async_playwright, Page, BrowserContext

base_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(base_dir, "engine_components"))
load_dotenv(os.path.join(base_dir, ".env"))
load_dotenv(os.path.join(base_dir, "..", ".env"))

# Six Strategy Engine (ports run_all_6.py verified strategies)
from six_strategy_engine import LiveSixStrategyPredictor, STRATEGY_NAMES as SIX_STRAT_NAMES
from rich.console import Console
from rich.live import Live
from rich.table import Table
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import concurrent.futures

# Dedicated thread pools — prevents ML predictor tasks from starving the renderer
ML_POOL = concurrent.futures.ThreadPoolExecutor(max_workers=8, thread_name_prefix="ML")
RENDER_POOL = concurrent.futures.ThreadPoolExecutor(max_workers=2, thread_name_prefix="Renderer")

# Reconfigure sys.stdout and sys.stderr to use UTF-8 encoding to prevent UnicodeEncodeError on Windows
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import ctypes

def get_process_memory_usage() -> int:
    try:
        class PROCESS_MEMORY_COUNTERS_EX(ctypes.Structure):
            _fields_ = [
                ("cb", ctypes.c_ulong),
                ("PageFaultCount", ctypes.c_ulong),
                ("PeakWorkingSetSize", ctypes.c_size_t),
                ("WorkingSetSize", ctypes.c_size_t),
                ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                ("QuotaPagedPoolUsage", ctypes.c_size_t),
                ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                ("PagefileUsage", ctypes.c_size_t),
                ("PeakPagefileUsage", ctypes.c_size_t),
                ("PrivateUsage", ctypes.c_size_t),
            ]
        psapi = ctypes.windll.psapi
        kernel32 = ctypes.windll.kernel32
        kernel32.GetCurrentProcess.restype = ctypes.c_void_p
        psapi.GetProcessMemoryInfo.argtypes = [ctypes.c_void_p, ctypes.c_void_p, ctypes.c_ulong]
        
        pmc = PROCESS_MEMORY_COUNTERS_EX()
        pmc.cb = ctypes.sizeof(pmc)
        handle = kernel32.GetCurrentProcess()
        if psapi.GetProcessMemoryInfo(handle, ctypes.byref(pmc), pmc.cb):
            return pmc.PrivateUsage
    except Exception:
        pass
    return 0

base_dir = os.path.dirname(os.path.abspath(__file__))
EXECUTION_MODE = os.environ.get("EXECUTION_MODE", "LIVE")
ENGINE_RISK_PCT = float(os.environ.get("ENGINE_RISK_PCT", "0.004"))
ENGINE_RISK_USD = float(os.environ.get("ENGINE_RISK_USD", "20.0"))
BINANCE_LIVE = os.environ.get("BINANCE_LIVE", "0") == "1"

# Strategy identity constants (used by Engine1TradeTracker cooldown logic)
ACTIVE_STRATEGY = os.environ.get("ACTIVE_STRATEGY", "ml_alpha_squeezer")
STRATEGY_DISPLAY_NAME = ACTIVE_STRATEGY.replace("_", " ").title().replace(" ", "_")


def _parse_suffix_float(val: Any) -> float | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("$", "")
    # Replace all unicode minus variants with ASCII minus
    for ch in ("\u2212", "\u2012", "\u2013", "\u2014"):
        s = s.replace(ch, "-")
    if not s or s.lower() == "n/a" or s == "--" or s.lower() == "nan" or s == "∅":
        return None
    try:
        mul = 1.0
        if s.endswith("%"):
            s = s[:-1]
        if s.lower().endswith("k"):
            mul = 1000.0
            s = s[:-1]
        elif s.lower().endswith("m"):
            mul = 1000000.0
            s = s[:-1]
        elif s.lower().endswith("b"):
            mul = 1000000000.0
            s = s[:-1]
        elif s.lower().endswith("t"):
            mul = 1000000000000.0
            s = s[:-1]
        f = float(s)
        return f * mul if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None

def parse_float(val: Any) -> float:
    res = _parse_suffix_float(val)
    return res if res is not None else 0.0

def finite_float_or_none(val: Any) -> float | None:
    return _parse_suffix_float(val)

def get_historical_timestamps(symbol: str, start_time_ts: int, steps: int) -> List[int]:
    is_crypto = symbol not in ["XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]
    if is_crypto:
        return [int(start_time_ts - i * 900) for i in range(steps)]

    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    ny_tz = ZoneInfo("America/New_York")

    def is_active(dt):
        day = dt.weekday()
        hour = dt.hour
        if day == 4:  # Friday
            if hour >= 17: return False
        elif day == 5:  # Saturday
            return False
        elif day == 6:  # Sunday
            if hour < 18: return False
        if day in (0, 1, 2, 3):  # Mon-Thu
            if hour == 17: return False
        return True

    dt = datetime.fromtimestamp(start_time_ts, tz=ny_tz)
    dt = dt.replace(minute=(dt.minute // 15) * 15, second=0, microsecond=0)

    while not is_active(dt):
        dt -= timedelta(minutes=15)

    timestamps = []
    for _ in range(steps):
        timestamps.append(int(dt.timestamp()))
        dt -= timedelta(minutes=15)
        while not is_active(dt):
            dt -= timedelta(minutes=15)

    return timestamps

def calculate_commodity_gap(symbol: str, latest_time: int, current_time: int) -> int:
    is_crypto = symbol not in ["XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]
    if is_crypto:
        return max(0, int((current_time - latest_time) / 900))
    timestamps = get_historical_timestamps(symbol, current_time, 2000)
    for idx, ts in enumerate(timestamps):
        if ts <= latest_time:
            return idx
    return 1000

# --- GLOBAL CONFIGURATION ---
URL = "https://www.coinglass.com/tv/layout/s9"
TAB1_SYMBOLS = ["BTCUSDT", "ETHUSDT", "XRPUSDT", "SOLUSDT", "BNBUSDT", "DOGEUSDT", "ADAUSDT", "TRXUSDT", "LINKUSDT"]
TAB2_SYMBOLS = ["AVAXUSDT", "SUIUSDT", "NEARUSDT", "DOTUSDT", "LTCUSDT", "XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]
ALL_SYMBOLS = TAB1_SYMBOLS + TAB2_SYMBOLS
REFRESH_HZ = 2.0  # 2 Hz rendering = 0.5s interval
STALE_NS = 15_000_000_000  # 15 seconds staleness threshold

# Column-level staleness tracking for purple bold formatting (>60s unchanged)
_COLUMN_LAST_VALUES: Dict[str, Any] = {}
_COLUMN_LAST_CHANGED_TIME: Dict[str, float] = {}
_COLUMN_STALE_THRESHOLD = 60.0  # seconds

# Live Event Log Ring Buffer for zero-flicker terminal log panel
_LIVE_LOG_FEED: collections.deque = collections.deque(maxlen=6)

def log_live_event(msg: str, tag: str = "SYS") -> None:
    stamp = datetime.now().strftime("%H:%M:%S")
    _LIVE_LOG_FEED.append(f"[{stamp}] [{tag}] {msg}")

# --- STATE MANAGEMENT ---
_FLOAT_FIELDS = {
    'price', 'volume', 'rsi', 'fut_cvd', 'spot_cvd', 'liq_long', 'liq_short',
    'funding', 'ls_ratio', 'oi', 'fp_delta', 'fp_poc', 'coins_bid', 'coins_ask',
    'dollars_bid', 'dollars_ask', 'whale_idx', 'tk_buy_cnt', 'tk_sell_cnt', 'tk_delta',
    'ema_8', 'ema_21', 'ema_50', 'ema_200', 'ema_800', 'atr_14', 'atr_100', 'atr',
    'zc4', 'zc10', 'zc20', 'zb4', 'zb10', 'zb20', 'vr', 'zoi', 'zls', 'zfr',
    'p8', 'p21', 'p50'
}

@dataclasses.dataclass
class EngineConfig:
    tp_mult: float = 5.0
    trail_atr: float = 0.8
    risk_per_trade: float = 10.0
    max_daily_risk: float = 150.0
    max_drawdown_pct: float = 8.0
    fee_pct: float = 0.0008

config = EngineConfig()

@dataclasses.dataclass
class AssetSnapshot:
    symbol: str
    price: float = 0.0
    volume: float = 0.0
    rsi: float = 0.0
    fut_cvd: float = 0.0
    spot_cvd: float = 0.0
    liq_long: float = 0.0
    liq_short: float = 0.0
    funding: float = 0.0
    ls_ratio: float = 0.0
    oi: float = 0.0
    fp_delta: float = 0.0
    fp_poc: float = 0.0
    coins_bid: float = 0.0
    coins_ask: float = 0.0
    dollars_bid: float = 0.0
    dollars_ask: float = 0.0
    whale_idx: float = 0.0
    tk_buy_cnt: float = 0.0
    tk_sell_cnt: float = 0.0
    tk_delta: float = 0.0
    strategy_armed: str = ""
    ts_ns: int = 0
    seq: int = 0
    # Enriched ML Strategy Fields
    ema_8: float = 0.0
    ema_21: float = 0.0
    ema_50: float = 0.0
    ema_200: float = 0.0
    ema_800: float = 0.0
    atr: float = 0.0
    atr_14: float = 0.0
    atr_100: float = 0.0
    zc4: float = 0.0
    zc10: float = 0.0
    zc20: float = 0.0
    zb4: float = 0.0
    zb10: float = 0.0
    zb20: float = 0.0
    vr: float = 0.0
    zoi: float = 0.0
    zls: float = 0.0
    zfr: float = 0.0
    p8: float = 0.0
    p21: float = 0.0
    p50: float = 0.0

    def __post_init__(self):
        for f in _FLOAT_FIELDS:
            try:
                setattr(self, f, float(getattr(self, f)))
            except (ValueError, TypeError):
                setattr(self, f, 0.0)

class BinanceBrokerAdapter:
    def __init__(self, binance_broker, tracker):
        self.broker = binance_broker
        self.tracker = tracker
        self.dry_run = binance_broker.dry_run

    @property
    def account_size(self):
        return self.tracker.current_capital

    @account_size.setter
    def account_size(self, val):
        pass

    def connect(self) -> bool:
        return self.broker.connect()

    def execute_trade(self, symbol, direction, entry_price, sl, tp, strategy):
        # Determine risk capital
        import os
        env_risk_usd = float(os.environ.get("ENGINE_RISK_USD", str(ENGINE_RISK_USD)))
        if env_risk_usd > 0.0:
            risk_capital = env_risk_usd
        else:
            risk_capital = self.tracker.current_capital * ENGINE_RISK_PCT

        res = self.broker.execute_trade(
            binance_symbol=symbol,
            direction=direction,
            bin_entry=entry_price,
            bin_sl=sl,
            bin_tp=tp,
            strategy=strategy,
            risk_capital=risk_capital
        )
        if res:
            return {
                "symbol": res["symbol"],
                "order_id": res["order_id"],
                "order_id": res["order_id"],
                "deal_id": res["order_id"],
                "exec_entry": res["entry_price"],
                "exec_sl": res["sl_price"],
                "exec_tp": res["tp_price"],
                "lot": res["lot"],
                "is_pending": res.get("is_pending", False)
            }
        else:
            print(f"[Broker] [WARNING] Order execution returned None for {symbol} ({strategy}). Possible margin/API restriction.")
        return None

    def close_position(self, symbol, reason="ENGINE_EXIT") -> bool:
        return self.broker.close_position(symbol, reason)

    def modify_sltp(self, symbol, ticket, sl, tp) -> bool:
        return self.broker.modify_sltp(symbol, ticket, sl, tp)

    def is_order_pending(self, order_ticket) -> bool:
        return False

    def has_position(self, ticket) -> bool:
        if self.dry_run:
            return True
        symbol = None
        for t in self.tracker.active_trades.values():
            if t.get("order_id") == ticket or t.get("order_id") == ticket:
                symbol = t.get("symbol")
                break
        if not symbol or not self.broker.is_valid_symbol(symbol):
            return False

        try:
            res = self.broker._request("GET", "/fapi/v2/positionRisk", params={"symbol": symbol}, signed=True, max_retries=1)
            if res:
                for p in res:
                    if p["symbol"] == symbol:
                        return float(p.get("positionAmt", 0.0)) != 0.0
        except Exception:
            pass
        return False

    def list_engine_positions(self) -> list:
        if self.dry_run:
            return []
        try:
            res = self.broker._request("GET", "/fapi/v2/positionRisk", signed=True, max_retries=1)
            if res:
                class PositionObj:
                    def __init__(self, ticket):
                        self.ticket = ticket
                active_positions = []
                for p in res:
                    if float(p.get("positionAmt", 0.0)) != 0.0:
                        # Find corresponding order ID ticket from active trades
                        ticket = None
                        for t in self.tracker.active_trades.values():
                            if t.get("symbol") == p.get("symbol"):
                                ticket = t.get("order_id")
                                break
                        if ticket is not None:
                            active_positions.append(PositionObj(ticket))
                return active_positions
        except Exception as e:
            print(f"[Binance] Error querying live positionRisk: {e}")
        return []


class LiveTradeTracker:
    REENTRY_COOLDOWN_TP_SECS = 3600   # 1 hour
    REENTRY_COOLDOWN_SL_SECS = 1800   # 30 minutes

    def _cooldown_key(self, strategy: str, symbol: str) -> str:
        return f"{strategy}:{symbol}"

    def _cooldown_secs_after_close(self, strategy: str, reason: str) -> int:
        six_strat_names = {
            "S1_Liquidation", "S2_CVD_Momentum", "S3_Trend_Follow",
            "S4_Mean_Reversion", "S5_Vol_Breakout", "S6_OI_Coherence"
        }
        if strategy in six_strat_names:
            if reason == "TP": return self.REENTRY_COOLDOWN_TP_SECS
            return self.REENTRY_COOLDOWN_SL_SECS
        if strategy == "ML_Liquidation_Runner":
            if reason == "TP": return 0
            if reason in ("SL", "BE", "TRAIL"): return 15 * 60
        if strategy in (STRATEGY_DISPLAY_NAME, "AlphaSqueezer_V17", "AlphaSqueezer_V11"):
            if reason == "TP": return self.REENTRY_COOLDOWN_TP_SECS
            return self.REENTRY_COOLDOWN_SL_SECS
        if strategy == "ML_Trend_Pull":
            if reason == "TP": return self.REENTRY_COOLDOWN_TP_SECS
            return self.REENTRY_COOLDOWN_SL_SECS
        return 15 * 60

    def __init__(self, initial_capital: float = 10000.0, base_dir: str = "."):
        self.base_dir = base_dir
        self.tracker_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "Engine_1_trade_logs.json"
        )
        self.log_file = self.tracker_file
        self.lock = threading.RLock()
        
        # --- Binance Broker Initialization ---
        from concurrent.futures import ThreadPoolExecutor
        self.broker_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="BinanceBroker")

        from engine_components.binance_broker import BinanceBroker
        binance_live = os.environ.get("BINANCE_LIVE", "1") != "0"
        use_testnet = os.environ.get("BINANCE_USE_TESTNET", "true").lower() == "true"
        
        raw_binance_broker = BinanceBroker(
            dry_run=not binance_live,
            account_size=initial_capital,
            risk_pct=ENGINE_RISK_PCT,
            use_testnet=use_testnet
        )
        self.broker = BinanceBrokerAdapter(raw_binance_broker, self)
        
        if self.broker.connect():
            details = raw_binance_broker.get_account_details()
            if details and details.get("balance", 0.0) > 0.0:
                initial_capital = details["balance"]
        # -------------------------------------
        
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.daily_start_capital = initial_capital
        self.emergency_halt = False
        self.on_close_callbacks = []
        
        import zoneinfo
        from datetime import datetime
        broker_tz = zoneinfo.ZoneInfo("Europe/Athens")
        self.last_rollover_day = datetime.now(broker_tz).strftime("%Y-%m-%d")
        self.active_trades: Dict[str, dict] = {}
        self.closed_trades: List[dict] = []
        self.history: List[dict] = []
        self.last_entry_bar: Dict[str, str] = {}
        self.reentry_cooldown_until: Dict[str, float] = {}
        
        self._load_state = self.load_history
        self.load_history()

    def _translate_to_binance_price(self, trade: dict, price: float) -> float:
        return float(price)

    def _broker_submit_checked(self, trade_id, fn, *args) -> None:
        if not hasattr(self, "broker_executor") or self.broker_executor is None:
            print(f"[Binance][FATAL] broker_executor missing — cannot dispatch {fn.__name__}")
            return
        try:
            fut = self.broker_executor.submit(fn, *args)
            def _log_done(f):
                try:
                    ok = f.result()
                except Exception as exc:
                    ok = False
                    print(f"[Binance] Async broker call {fn.__name__} failed: {exc}")
                
                if not ok:
                    with self.lock:
                        tr = self.active_trades.get(trade_id)
                        if tr:
                            tr["broker_sync_error"] = f"{fn.__name__}_FAILED"
                            tr["needs_manual_attention"] = True
                            print(f"[Binance] SL modify failed for {trade_id}! Trade tagged for manual attention.")
            fut.add_done_callback(_log_done)
        except Exception as exc:
            print(f"[Binance] Failed to submit broker action: {exc}")

    def load_history(self):
        """Load only from this engine's own log — never from Engine 3 whose
        pnl_usd uses an incompatible unit-based sizing model."""
        with self.lock:
            if not os.path.exists(self.log_file):
                return
            try:
                with open(self.log_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # Support envelope format with __meta__
                meta = {}
                if isinstance(data, dict) and '__meta__' in data:
                    meta = data['__meta__']
                    data = data.get('trades', [])
                if not isinstance(data, list):
                    return
                self.last_entry_bar = meta.get('last_entry_bar', {})
                # FIX 2: Strip test/dry-run artifacts injected by boot scripts.
                # These have synthetic trade_ids and corrupt the P&L baseline.
                raw_history = [t for t in data if t.get('exit_price')]
                self.history = [
                    t for t in raw_history
                    if 'test' not in t.get('trade_id', '').lower()
                    and 'emergency_test' not in t.get('trade_id', '').lower()
                ]
                self.current_capital = self.initial_capital
                self.daily_start_capital = meta.get('daily_start_capital', self.current_capital)
                self.last_rollover_day = meta.get('last_rollover_day', self.last_rollover_day)

                # FIX: If the log is from a prior session (different day), reset daily DD baseline
                # to prevent stale cumulative losses from triggering an immediate emergency halt.
                import zoneinfo
                from datetime import datetime as _dt
                broker_tz = zoneinfo.ZoneInfo("Europe/Athens")
                today = _dt.now(broker_tz).strftime("%Y-%m-%d")
                if self.last_rollover_day != today:
                    self.daily_start_capital = self.current_capital
                    self.last_rollover_day = today
                    print(f"[RiskGovernor] New session detected (last rollover: {meta.get('last_rollover_day', 'unknown')}). "
                          f"Daily DD baseline reset to ${self.current_capital:.2f}")
                for t in data:
                    if not t.get('exit_price') and t.get('trade_id'):
                        self.active_trades[t['trade_id']] = t
            except Exception as e:
                print(f"[TradeTracker] [ERROR] Failed to load trade history: {e}")

    def save_history(self):
        with self.lock:
            try:
                if len(self.history) > 5000:
                    self.history = self.history[-5000:]
                all_trades = list(self.history) + list(self.active_trades.values())
                envelope = {
                    '__meta__': {
                        'last_entry_bar': dict(self.last_entry_bar),
                        'daily_start_capital': self.daily_start_capital,
                        'last_rollover_day': self.last_rollover_day
                    },
                    'trades': all_trades
                }
                tmp = self.log_file + ".tmp"
                with open(tmp, 'w', encoding='utf-8') as f:
                    json.dump(envelope, f, indent=4)
                os.replace(tmp, self.log_file)
            except Exception:
                pass

    def update_day(self) -> None:
        with self.lock:
            import zoneinfo
            from datetime import datetime
            broker_tz = zoneinfo.ZoneInfo("Europe/Athens")
            now_day = datetime.now(broker_tz).strftime("%Y-%m-%d")
            
            if self.last_rollover_day != now_day:
                self.daily_start_capital = self.current_capital
                self.last_rollover_day = now_day
                print(f"[RiskGovernor] Daily starting capital rolled over to ${self.daily_start_capital:.2f} (realized balance) at Athens server day {now_day}")

    def trigger_entry(self, symbol: str, strategy: str, direction: int, entry_price: float, sl: float, tp: float, atr: float, macro: int, vol_regime: float, risk_mult: float = 1.0, trail_act: float = 0.5, regime_val: int = 0) -> None:
        with self.lock:
            if getattr(self, 'emergency_halt', False):
                log_live_event(f"Entry blocked. Symbol={symbol} Strategy={strategy}. Emergency halt active.", "RiskGov")
                return

            # --- GLOBAL RISK GOVERNOR (10% Daily Governance Drawdown Limit) ---
            active_list = list(self.active_trades.values())
            unrealized_pnl = sum(t.get('live_pnl_usd', 0.0) for t in active_list)
            current_equity = self.current_capital + unrealized_pnl

            # 1. Daily Drawdown Check (Hard limit 10%, Guardrail 9.0%)
            daily_dd = (self.daily_start_capital - current_equity) / self.daily_start_capital * 100.0 if self.daily_start_capital > 0 else 0.0
            if daily_dd >= 9.0:
                log_live_event(f"Entry blocked. Symbol={symbol} Strategy={strategy}. Daily DD ({daily_dd:.2f}%) > 9%.", "RiskGov")
                return

            # 2. Total Drawdown Check (Hard limit 15%, Guardrail 14.0% of initial capital)
            total_dd = (self.initial_capital - current_equity) / self.initial_capital * 100.0
            if total_dd >= 14.0:
                log_live_event(f"Entry blocked. Symbol={symbol} Strategy={strategy}. Total DD ({total_dd:.2f}%) > 14%.", "RiskGov")
                return

            cool_key = self._cooldown_key(strategy, symbol)
            cooldown_until = self.reentry_cooldown_until.get(cool_key, 0.0)
            if time.time() < cooldown_until:
                log_live_event(f"Entry blocked by cooldown. {symbol} {strategy} Rem: {(cooldown_until - time.time()):.0f}s", "RiskGov")
                return

            strategy_trades = [t for t in self.active_trades.values() if t['strategy'] == strategy]
            if any(t['symbol'] == symbol for t in strategy_trades):
                return
            
            max_concurrent = 3 if regime_val == 1 else 5
            if len(strategy_trades) >= max_concurrent:
                return
                
            if direction not in (1, -1):
                return
            
            if direction == 1 and not (sl < entry_price < tp):
                return
            if direction == -1 and not (tp < entry_price < sl):
                return
                
            stop_dist = abs(entry_price - sl)

            # --- TIGHT-SL FLOOR: Reject entries where SL is tighter than minimum % of price ---
            # Per-symbol minimum stop distances (wider for low-priced/high-spread assets)
            MIN_STOP_PCT = {
                'BTCUSDT': 0.0008, 'ETHUSDT': 0.0008, 'BNBUSDT': 0.001,
                'SOLUSDT': 0.001, 'XRPUSDT': 0.001, 'LINKUSDT': 0.001,
                'AVAXUSDT': 0.001, 'LTCUSDT': 0.001, 'DOTUSDT': 0.001,
                'ADAUSDT': 0.0015, 'NEARUSDT': 0.0015, 'SUIUSDT': 0.0015,
                'DOGEUSDT': 0.002, 'TRXUSDT': 0.002,
                'XAUUSDT': 0.0005, 'XAGUSDT': 0.001,
                'CLUSDT': 0.0015, 'NATGASUSDT': 0.003,
            }
            min_stop_pct = MIN_STOP_PCT.get(symbol, 0.003)  # Default 0.3%
            min_stop_dist = entry_price * min_stop_pct
            if stop_dist < min_stop_dist:
                # Adaptive widening: floor SL to minimum safe distance instead of blocking
                stop_dist = min_stop_dist
                if direction == 1:
                    sl = entry_price - stop_dist
                else:
                    sl = entry_price + stop_dist
                log_live_event(f"{symbol} {strategy} SL widened to {stop_dist:.6f} ({min_stop_pct*100:.1f}% floor)", "RiskGov")

            env_risk_usd = float(os.environ.get("ENGINE_RISK_USD", str(ENGINE_RISK_USD)))
            if env_risk_usd > 0.0:
                risk_capital = env_risk_usd * risk_mult
            else:
                risk_capital = max(0.0, self.current_capital) * ENGINE_RISK_PCT * risk_mult
            
            if risk_capital <= 0.0 or stop_dist <= 0:
                return
                
            # --- FRICTION-AWARE SIZING: Deduct 0.12% round-trip Binance taker friction from risk budget ---
            TOTAL_FRICTION = 0.0012
            effective_stop_dist = stop_dist + (entry_price * TOTAL_FRICTION)
            units = risk_capital / effective_stop_dist if effective_stop_dist > 0 else 0.0

            # --- NOTIONAL CAP: Never open a position > $50,000 notional ---
            MAX_NOTIONAL = 50_000.0
            notional = units * entry_price
            if notional > MAX_NOTIONAL:
                units = MAX_NOTIONAL / entry_price
                log_live_event(f"{symbol} notional capped: ${notional:.0f} -> ${MAX_NOTIONAL:.0f}", "RiskGov")

            # 3. Overall Portfolio Open Stop Risk Check (Max 4% of current equity)
            open_stop_risk = 0.0
            for t in active_list:
                t_units = t.get('units', 0.0)
                t_dir = t.get('direction', 1)
                t_ep = t.get('entry_price', 0.0)
                t_sl = t.get('sl', 0.0)
                if t_dir == 1:
                    risk_pts = max(0.0, t_ep - t_sl)
                else:
                    risk_pts = max(0.0, t_sl - t_ep)
                open_stop_risk += t_units * risk_pts
            new_trade_risk = units * stop_dist
            total_portfolio_risk = open_stop_risk + new_trade_risk
            if total_portfolio_risk > current_equity * 0.04:
                now_wall = time.time()
                if now_wall - getattr(self, '_last_risk_cap_log_time', 0) > 60.0:
                    self._last_risk_cap_log_time = now_wall
                    log_live_event(f"{symbol} {strategy} blocked: risk (${total_portfolio_risk:.0f}) > 4% cap", "RiskGov")
                return

            trade_id = f"{strategy}_{symbol}_{'LONG' if direction == 1 else 'SHORT'}_{int(time.time_ns())}"
            log_live_event(f"ENTRY: {symbol} {'LONG' if direction == 1 else 'SHORT'} @ {entry_price:.4f} (Lot: {units:.2f}) [{strategy}]", "EXEC")
            self.active_trades[trade_id] = {
                "trade_id": trade_id,
                "symbol": symbol,
                "strategy": strategy,
                "direction": direction,
                "entry_price": entry_price,
                "entry_time": time.strftime("%Y-%m-%d %H:%M:%S"),
                "entry_timestamp": time.time(),
                "sl": sl,
                "tp": tp,
                "units": units,
                "live_pnl_pct": 0.0,
                "live_pnl_usd": 0.0,
                "atr": atr,
                "macro": macro,
                "vol_regime": vol_regime,
                "sl_dist": stop_dist,
                "intended_tp_dist": abs(tp - entry_price),
                "trail_act": trail_act,
                "trail_buf": 0.8
            }
            
            # --- Binance Execution Dispatch ---
            try:
                broker_res = self.broker.execute_trade(symbol, direction, entry_price, sl, tp, strategy)
            except Exception as e:
                print(f"[TradeTracker] execute_trade raised exception for {symbol} ({strategy}): {e} — aborting phantom trade.")
                self.active_trades.pop(trade_id, None)
                return
            if broker_res:
                self.active_trades[trade_id]["symbol"] = broker_res.get("symbol")
                self.active_trades[trade_id]["order_id"] = broker_res.get("order_id")
                self.active_trades[trade_id]["order_id"] = broker_res.get("order_id")
                self.active_trades[trade_id]["deal_id"] = broker_res.get("deal_id")
                self.active_trades[trade_id]["exec_entry"] = broker_res.get("exec_entry")
                self.active_trades[trade_id]["exec_sl"] = broker_res.get("exec_sl")
                self.active_trades[trade_id]["exec_tp"] = broker_res.get("exec_tp")
                self.active_trades[trade_id]["exec_lot"] = broker_res.get("lot")
                if broker_res.get("lot"):
                    self.active_trades[trade_id]["units"] = broker_res["lot"]
                self.active_trades[trade_id]["is_pending"] = broker_res.get("is_pending", False)
            else:
                print(f"[TradeTracker] Broker rejected {symbol} ({strategy}) - removing phantom trade.")
                self.active_trades.pop(trade_id, None)
                return
            # ------------------------------
            
            self.save_history()

    def update_live_pnl(self, symbol: str, current_price: float, store: Optional[Any] = None) -> None:
        with self.lock:
            # 1. Update individual trade PnL
            trades_for_symbol = [t for t in self.active_trades.values() if t['symbol'] == symbol]
            for trade in trades_for_symbol:
                if trade.get("is_pending"):
                    order_id = trade.get("order_id")
                    if order_id and not self.broker.dry_run:
                        if not self.broker.is_order_pending(order_id):
                            # Resolve real position ticket (order ticket != position ticket)
                            pos_ticket = None
                            if hasattr(self.broker, "resolve_position_from_order"):
                                pos_ticket = self.broker.resolve_position_from_order(
                                    order_id, trade.get("symbol")
                                )
                            if pos_ticket is None and self.broker.has_position(order_id):
                                pos_ticket = order_id  # fallback
                            if pos_ticket:
                                log_live_event(f"Limit order {order_id} for {symbol} filled -> pos={pos_ticket}. Activating trade.", "Binance")
                                trade["is_pending"] = False
                                trade["order_id"] = pos_ticket
                            else:
                                log_live_event(f"Limit order {order_id} for {symbol} cancelled/expired. Removing phantom.", "Binance")
                                del self.active_trades[trade["trade_id"]]
                                continue
                    elif self.broker.dry_run:
                        trade["is_pending"] = False
                    
                    if trade.get("is_pending"):
                        continue
                        
                direction = trade['direction']
                entry_price = trade['entry_price']
                pnl_pct = (current_price - entry_price) / entry_price * 100.0 if direction == 1 else (entry_price - current_price) / entry_price * 100.0
                pnl_usd = trade['units'] * (current_price - entry_price) * direction
                trade['live_pnl_pct'] = pnl_pct
                trade['live_pnl_usd'] = pnl_usd

            # 2. Run global equity drawdown breach checks (Hard Kill-Switch)
            active_list = list(self.active_trades.values())
            unrealized_pnl = sum(t.get('live_pnl_usd', 0.0) for t in active_list)
            current_equity = self.current_capital + unrealized_pnl

            daily_dd = (self.daily_start_capital - current_equity) / self.daily_start_capital * 100.0 if self.daily_start_capital > 0 else 0.0
            total_dd = (self.initial_capital - current_equity) / self.initial_capital * 100.0

            if daily_dd >= 10.0 or total_dd >= 15.0:
                if not getattr(self, 'emergency_halt', False):
                    self.emergency_halt = True
                    log_live_event(f"[CRITICAL] EMERGENCY HALT! Daily DD={daily_dd:.2f}%, Total DD={total_dd:.2f}%. Closing all.", "RiskGov")
                
                # Pre-dispatch parallel closes
                close_futures = {}
                if not self.broker.dry_run and hasattr(self, "broker_executor") and self.broker_executor:
                    for trade in list(self.active_trades.values()):
                        if trade.get("order_id"):
                            fut = self.broker_executor.submit(self.broker.close_position, trade["symbol"], "EMERGENCY_HALT")
                            close_futures[trade['trade_id']] = fut

                any_closed = False
                for trade in list(self.active_trades.values()):
                    tid = trade['trade_id']
                    
                    if tid in close_futures:
                        try:
                            ok = close_futures[tid].result(timeout=10.0)
                        except Exception:
                            ok = False
                        if not ok or self.broker.has_position(trade.get("order_id")):
                            trade["emergency_close_failed"] = True
                            continue # Keep trying on next loop

                    trade_sym = trade['symbol']
                    if trade_sym == symbol:
                        exit_price = current_price
                    elif store is not None:
                        snap_obj = store._data.get(trade_sym)
                        exit_price = snap_obj.price if (snap_obj and snap_obj.price > 0.0) else trade['entry_price']
                    else:
                        exit_price = trade['entry_price']

                    trade['exit_price'] = exit_price
                    trade['exit_time'] = time.strftime("%Y-%m-%d %H:%M:%S")
                    trade['exit_reason'] = "EMERGENCY_HALT"
                    
                    entry_price = trade['entry_price']
                    direction = trade['direction']
                    live_pnl_pct = (exit_price - entry_price) / entry_price * 100.0 if direction == 1 else (entry_price - exit_price) / entry_price * 100.0
                    live_pnl_usd = trade['units'] * (exit_price - entry_price) * direction

                    pnl_pct = live_pnl_pct - 0.06
                    pnl_usd = live_pnl_usd - (trade['units'] * entry_price * 0.0006)
                    
                    trade['pnl_pct'] = pnl_pct
                    trade['pnl_usd'] = pnl_usd
                    
                    self.history.append(trade)
                    self.current_capital += pnl_usd
                    
                    del self.active_trades[trade['trade_id']]
                    any_closed = True
                    
                    closed_strategy = trade.get('strategy', '')
                    for cb in self.on_close_callbacks:
                        try:
                            cb(closed_strategy, self.current_capital)
                        except Exception:
                            pass
                if any_closed:
                    self.save_history()

    def check_exits(self, symbol: str, current_price: float, current_atr_or_dict: Any = 0.0) -> None:
        with self.lock:
            # SL Heartbeat: periodically push current trailing SL to exchange
            now = time.time()
            if now - getattr(self, 'last_sl_heartbeat', now) > 60:
                self.last_sl_heartbeat = now
                for t in self.active_trades.values():
                    if t.get("order_id") and not self.broker.dry_run:
                        exec_sl = self._translate_to_binance_price(t, t["sl"])
                        exec_tp = self._translate_to_binance_price(t, t["tp"]) if t.get("tp") else None
                        self._broker_submit_checked(t["trade_id"], self.broker.modify_sltp, t["symbol"], t["order_id"], exec_sl, exec_tp)

            trades_for_symbol = [t for t in self.active_trades.values() if t['symbol'] == symbol]
            any_closed = False
            for trade in trades_for_symbol:
                if trade.get("is_pending"):
                    continue
                direction = trade['direction']
                sl = trade['sl']
                tp = trade['tp']
                entry_price = trade['entry_price']
                
                # Resolve current ATR based on strategy if a dict is passed
                if isinstance(current_atr_or_dict, dict):
                    strategy = trade.get('strategy', STRATEGY_DISPLAY_NAME)
                    current_atr = float(current_atr_or_dict.get(strategy, 0.0) or 0.0)
                else:
                    current_atr = float(current_atr_or_dict or 0.0)
                
                # --- OOS EXACT RATCHET TRAILING STOP (run_all_6.py parity) ---
                entry_atr = trade.get('atr', 0.0)
                tp_dist = trade.get('intended_tp_dist', abs(tp - entry_price))
                sl_dist_val = trade.get('sl_dist', abs(entry_price - sl))
                trail_dist = 0.8 * entry_atr if entry_atr > 0 else (0.8 * sl_dist_val if sl_dist_val > 0 else 0.0)

                if direction == 1:
                    profit_from_entry = current_price - entry_price
                    if profit_from_entry >= tp_dist:  # ONLY activate after reaching 5.0R target
                        best_price = max(trade.get('best_price', current_price), current_price)
                        trade['best_price'] = best_price
                        new_sl = best_price - trail_dist
                        if new_sl > sl:
                            trade['sl'] = new_sl
                            sl = new_sl
                            if trade.get("order_id") and not self.broker.dry_run:
                                if now - trade.get('last_sl_modify_time', 0.0) >= 1.0:
                                    trade['last_sl_modify_time'] = now
                                    exec_sl = self._translate_to_binance_price(trade, sl)
                                    exec_tp = self._translate_to_binance_price(trade, trade["tp"])
                                    self._broker_submit_checked(trade["trade_id"], self.broker.modify_sltp, trade["symbol"], trade["order_id"], exec_sl, exec_tp)
                else:
                    profit_from_entry = entry_price - current_price
                    if profit_from_entry >= tp_dist:  # ONLY activate after reaching 5.0R target
                        best_price = min(trade.get('best_price', current_price), current_price)
                        trade['best_price'] = best_price
                        new_sl = best_price + trail_dist
                        if new_sl < sl:
                            trade['sl'] = new_sl
                            sl = new_sl
                            if trade.get("order_id") and not self.broker.dry_run:
                                if now - trade.get('last_sl_modify_time', 0.0) >= 1.0:
                                    trade['last_sl_modify_time'] = now
                                    exec_sl = self._translate_to_binance_price(trade, sl)
                                    exec_tp = self._translate_to_binance_price(trade, trade["tp"])
                                    self._broker_submit_checked(trade["trade_id"], self.broker.modify_sltp, trade["symbol"], trade["order_id"], exec_sl, exec_tp)
                
                should_close = False
                reason = ""
                
                # --- MAX_BARS Timeout Exit (Parity with run_all_6.py _sim_trade) ---
                # 288 bars of 15m = 72 hours (259200 seconds)
                elapsed_time = time.time() - trade.get('entry_timestamp', time.time())
                if elapsed_time >= 259200:
                    should_close = True
                    reason = "TIMEOUT"
                
                if not should_close:
                    if direction == 1:
                        if current_price <= sl:
                            should_close = True
                            reason = "SL"
                    else:
                        if current_price >= sl:
                            should_close = True
                            reason = "SL"
                    # NOTE: No hard TP exit — OOS backtest relies solely on
                    # trailing stop ratchet after 5R profit. Hard TP would
                    # close at exactly 5R, leaving 6R-8R+ gains on the table.
                        
                if should_close:
                    if trade.get("closing_dispatched"):
                        continue
                        
                    exit_price = current_price
                    trade['exit_price'] = exit_price
                    trade['exit_time'] = time.strftime("%Y-%m-%d %H:%M:%S")
                    trade['exit_reason'] = reason
                    
                    entry_price = trade['entry_price']
                    pnl_pct = (exit_price - entry_price) / entry_price * 100.0 if direction == 1 else (entry_price - exit_price) / entry_price * 100.0
                    pnl_pct -= 0.06
                    
                    pnl_usd = (trade['units'] * (exit_price - entry_price) * direction) - (trade['units'] * entry_price * 0.0006)
                    
                    trade['pnl_pct'] = pnl_pct
                    trade['pnl_usd'] = pnl_usd
                    
                    if trade.get("order_id") and not self.broker.dry_run:
                        trade["closing_dispatched"] = True
                        
                        def make_close_cb(t_id, t_dict):
                            def _cb(f):
                                try:
                                    res = f.result()
                                    with self.lock:
                                        if res and t_id in self.active_trades:
                                            self.history.append(t_dict)
                                            self.current_capital += t_dict.get('pnl_usd', 0)
                                            del self.active_trades[t_id]
                                            self.save_history()
                                            
                                            try:
                                                details = self.broker.broker.get_account_details()
                                                if details and details.get("balance", 0.0) > 0.0:
                                                    self.current_capital = details["balance"]
                                            except Exception:
                                                pass
                                        elif not res and t_id in self.active_trades:
                                            log_live_event(f"Close rejected/failed for {t_id}. Re-arming local state.", "EXIT")
                                            self.active_trades[t_id]["closing_dispatched"] = False
                                except Exception as e:
                                    log_live_event(f"Exception during async close for {t_id}: {e}", "EXIT")
                                    with self.lock:
                                        if t_id in self.active_trades:
                                            self.active_trades[t_id]["closing_dispatched"] = False
                            return _cb
                            
                        if hasattr(self, "broker_executor") and self.broker_executor:
                            fut = self.broker_executor.submit(self.broker.close_position, trade["symbol"], reason)
                            fut.add_done_callback(make_close_cb(trade["trade_id"], trade.copy()))
                    else:
                        self.history.append(trade)
                        self.current_capital += pnl_usd
                        del self.active_trades[trade['trade_id']]
                        any_closed = True
                        
                    closed_strategy = trade.get("strategy", "")
                    cooldown_secs = self._cooldown_secs_after_close(closed_strategy, reason)
                    if cooldown_secs > 0:
                        self.reentry_cooldown_until[self._cooldown_key(closed_strategy, symbol)] = time.time() + cooldown_secs

                    for cb in self.on_close_callbacks:
                        try:
                            cb(closed_strategy, self.current_capital)
                        except Exception:
                            pass
                    
            if any_closed:
                self.save_history()

    def get_stats(self) -> dict:
        with self.lock:
            total = len(self.history)
            if total == 0:
                return {"total": 0, "winrate": 0.0, "total_pnl_usd": 0.0, "current_capital": self.current_capital}
            wins = sum(1 for t in self.history if t.get('pnl_usd', 0.0) > 0)
            total_pnl_usd = sum(t.get('pnl_usd', 0.0) for t in self.history)
            return {
                "total": total,
                "winrate": (wins / total) * 100.0,
                "total_pnl_usd": total_pnl_usd,
                "current_capital": self.current_capital
            }

    def reconcile_with_broker(self) -> None:
        """
        Keep active_trades in absolute sync with Binance positions.
        - Drop local trades whose Binance position is gone (broker SL/TP hit).
        - Promote filled pending orders to live tickets.
        Called periodically from the rollover watchdog (non-blocking path).
        """
        if getattr(self.broker, "dry_run", True):
            return
        with self.lock:
            try:
                broker_positions = {}
                if hasattr(self.broker, "list_engine_positions"):
                    for p in self.broker.list_engine_positions():
                        broker_positions[int(p.ticket)] = p

                stale_ids = []
                for tid, trade in list(self.active_trades.items()):
                    if trade.get("is_pending"):
                        order_id = trade.get("order_id")
                        if order_id and not self.broker.is_order_pending(order_id):
                            pos_ticket = None
                            if hasattr(self.broker, "resolve_position_from_order"):
                                pos_ticket = self.broker.resolve_position_from_order(
                                    order_id, trade.get("symbol")
                                )
                            if pos_ticket:
                                trade["is_pending"] = False
                                trade["order_id"] = pos_ticket
                            else:
                                stale_ids.append(tid)
                        continue

                    ticket = trade.get("order_id")
                    if not ticket:
                        continue
                    if ticket not in broker_positions and not self.broker.has_position(ticket):
                        # Broker already closed (SL/TP) — archive locally at last known price
                        trade["exit_price"] = trade.get("entry_price")
                        trade["exit_time"] = time.strftime("%Y-%m-%d %H:%M:%S")
                        trade["exit_reason"] = "BROKER_SYNC"
                        trade["pnl_pct"] = trade.get("live_pnl_pct", 0.0) - 0.06
                        trade["pnl_usd"] = trade.get("live_pnl_usd", 0.0)
                        self.history.append(trade)
                        self.current_capital += trade.get("pnl_usd", 0.0)
                        stale_ids.append(tid)
                        log_live_event(f"SYNC: Reconciled {trade.get('symbol')} position exit (PnL: ${trade.get('pnl_usd', 0):+.2f})", "Binance")

                for tid in stale_ids:
                    self.active_trades.pop(tid, None)
                if stale_ids:
                    self.save_history()
            except Exception as e:
                log_live_event(f"Reconcile error: {e}", "Binance")

INDICATOR_FRESHNESS_CONTRACTS: Dict[str, Dict[str, float]] = {
    "price": {"interval": 2.0, "tolerance": 3.0},          # Stale > 6.0s
    "fp_delta": {"interval": 2.0, "tolerance": 3.0},       # Stale > 6.0s
    "fp_poc": {"interval": 2.0, "tolerance": 3.0},         # Stale > 6.0s
    "rsi": {"interval": 60.0, "tolerance": 2.5},           # Stale > 150.0s
    "fut_cvd": {"interval": 30.0, "tolerance": 2.5},       # Stale > 75.0s
    "spot_cvd": {"interval": 30.0, "tolerance": 2.5},      # Stale > 75.0s
    "oi": {"interval": 60.0, "tolerance": 2.5},            # Stale > 150.0s
    "ls_ratio": {"interval": 60.0, "tolerance": 2.5},      # Stale > 150.0s
    "whale_idx": {"interval": 60.0, "tolerance": 2.5},     # Stale > 150.0s
    "dollars_bid": {"interval": 60.0, "tolerance": 2.5},   # Stale > 150.0s
    "dollars_ask": {"interval": 60.0, "tolerance": 2.5},   # Stale > 150.0s
    "liq_long": {"interval": 300.0, "tolerance": 2.0},     # Event-driven
    "liq_short": {"interval": 300.0, "tolerance": 2.0},    # Event-driven
    "funding": {"interval": 28800.0, "tolerance": 1.5},    # 8-Hour Exchange Settlement
}

Engine1TradeTracker = LiveTradeTracker


class SnapshotStore:
    def __init__(self, symbols: List[str], predictor=None, trade_tracker: Any = None):
        self._data: Dict[str, AssetSnapshot] = {s: AssetSnapshot(symbol=s) for s in symbols}
        self._locks = {s: asyncio.Lock() for s in symbols}
        self._seq = 0
        self.predictor = predictor
        self.trade_tracker = trade_tracker
        self._global_lock = threading.RLock()
        self._field_last_updated: Dict[str, Dict[str, float]] = {s: {} for s in symbols}

        # Proactively seed initial values from predictor candle histories so no symbol starts with zero
        if predictor and hasattr(predictor, "candles_history"):
            for s in symbols:
                hist = getattr(predictor, "candles_history", {}).get(s, [])
                if hist:
                    last_c = hist[-1]
                    cur = self._data[s]
                    self._data[s] = dataclasses.replace(
                        cur,
                        price=float(last_c.get("close", cur.price) or cur.price or 0.0),
                        rsi=float(last_c.get("rsi", cur.rsi) or cur.rsi or 0.0),
                        fut_cvd=float(last_c.get("fut_cvd", cur.fut_cvd) or cur.fut_cvd or 0.0),
                        spot_cvd=float(last_c.get("spot_cvd", cur.spot_cvd) or cur.spot_cvd or 0.0),
                        funding=float(last_c.get("funding", cur.funding) or cur.funding or 0.0),
                        oi=float(last_c.get("oi", cur.oi) or cur.oi or 0.0),
                        volume=float(last_c.get("volume", cur.volume) or cur.volume or 0.0),
                        fp_delta=float(last_c.get("fp_delta", cur.fp_delta) or cur.fp_delta or 0.0),
                        fp_poc=float(last_c.get("close", cur.fp_poc) or cur.fp_poc or 0.0),
                        ema_8=float(last_c.get("ema_8", cur.ema_8) or cur.ema_8 or 0.0),
                        ema_21=float(last_c.get("ema_21", cur.ema_21) or cur.ema_21 or 0.0),
                        ema_50=float(last_c.get("ema_50", cur.ema_50) or cur.ema_50 or 0.0),
                        ema_200=float(last_c.get("ema_200", cur.ema_200) or cur.ema_200 or 0.0),
                        ema_800=float(last_c.get("ema_800", cur.ema_800) or cur.ema_800 or 0.0),
                        atr_100=float(last_c.get("atr", cur.atr_100) or cur.atr_100 or 0.0),
                    )

        # Pipeline health metrics — updated by each subsystem
        self.pipeline_health: Dict[str, Any] = {
            "chrome_status": "INIT",
            "chrome_latency_ms": 0.0,
            "chrome_polls": 0,
            "chrome_last_poll_ns": 0,
            "binance_ws_status": "INIT",
            "binance_ws_url": "",
            "binance_ws_ticks": 0,
            "binance_broker_status": "INIT",
            "binance_broker_balance": 0.0,
            "binance_broker_positions": 0,
            "scraper_fps": 0.0,
            "scraper_last_parse_ns": time.time_ns(),
            "scraper_valid_ns": {s: time.time_ns() for s in ALL_SYMBOLS},
            "footprint_status": "INIT",
            "footprint_ticks": 0,
        }

    def is_field_stale(self, symbol: str, field_name: str) -> bool:
        contract = INDICATOR_FRESHNESS_CONTRACTS.get(field_name, {"interval": 60.0, "tolerance": 2.5})
        last_ts = self._field_last_updated.get(symbol, {}).get(field_name, 0.0)
        if last_ts == 0.0:
            return False
        threshold = contract["interval"] * contract["tolerance"]
        return (time.time() - last_ts) > threshold

    def is_scraper_dead(self, tab_id: str, timeout: float = 60.0) -> bool:
        last = self.pipeline_health.get(f"{tab_id}_last_heartbeat", time.time())
        return (time.time() - last) > timeout

    async def update(self, symbol: str, source: str = "binance", **patch: Any) -> None:
        if symbol not in self._data:
            return
        async with self._locks[symbol]:
            cur = self._data[symbol]
            clean_patch = {}
            _now_sec = time.time()
            for k, v in patch.items():
                if not hasattr(cur, k):
                    continue
                if k in ("price", "open", "high", "low", "close"):
                    fv = finite_float_or_none(v)
                    if fv is None or fv <= 0.0:
                        continue
                    if k == "price" and source == "coinglass" and cur.price > 0.0 and symbol not in ("XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"):
                        continue
                    clean_patch[k] = fv
                    self._field_last_updated[symbol][k] = _now_sec
                elif k in ("rsi", "oi", "ls_ratio", "dollars_bid", "dollars_ask"):
                    fv = finite_float_or_none(v)
                    if fv is None or fv <= 0.0:
                        continue
                    if k == "rsi" and (fv <= 0.0 or fv >= 100.0):
                        continue
                    clean_patch[k] = fv
                    self._field_last_updated[symbol][k] = _now_sec
                    cur_val = getattr(cur, k, 0.0)
                    if abs(fv - cur_val) > 1e-9:
                        self.pipeline_health.setdefault("last_change_ns", {})[symbol] = time.time_ns()
                elif k in (
                    "fut_cvd", "spot_cvd", "liq_long", "liq_short",
                    "funding", "whale_idx", "coins_bid", "coins_ask",
                    "tk_buy_cnt", "tk_sell_cnt", "fp_delta", "fp_poc"
                ):
                    fv = finite_float_or_none(v)
                    if fv is None:
                        continue
                    clean_patch[k] = fv
                    self._field_last_updated[symbol][k] = _now_sec
                    cur_val = getattr(cur, k, 0.0)
                    if abs(fv - cur_val) > 1e-9:
                        self.pipeline_health.setdefault("last_change_ns", {})[symbol] = time.time_ns()
                else:
                    clean_patch[k] = v
                    self._field_last_updated[symbol][k] = _now_sec

            now_ns = time.time_ns()
            self._seq += 1

            if not clean_patch:
                # Heartbeat: scraper is alive but price was filtered; bump ts_ns so UI stays green
                self._data[symbol] = dataclasses.replace(cur, ts_ns=now_ns)
                return

            new_snap = dataclasses.replace(cur, seq=self._seq, ts_ns=now_ns, **clean_patch)
            
            # Track if any actual indicators (not just price/volume) were updated
            indicator_keys = {
                "rsi", "fut_cvd", "spot_cvd", "liq_long", "liq_short", "funding", "ls_ratio", "oi",
                "ema_8", "ema_21", "ema_50", "ema_200", "ema_800", "atr_100", "atr_14", "atr",
                "volume", "coins_bid", "coins_ask", "dollars_bid", "dollars_ask"
            }
            
            if "scraper_valid_ns" not in self.pipeline_health:
                self.pipeline_health["scraper_valid_ns"] = {}

            if source == "coinglass" or any(k in clean_patch for k in indicator_keys):
                self.pipeline_health["scraper_valid_ns"][symbol] = now_ns

            if self.trade_tracker:
                self.trade_tracker.update_day()

            price_updated = "price" in clean_patch
            
            if self.trade_tracker and price_updated:
                # Use ATR from the unified predictor's cached signals
                atr_dict = {}
                if self.predictor and hasattr(self.predictor, '_cached_signals'):
                    cached = self.predictor._cached_signals.get(symbol, {})
                    atr_val = cached.get('atr_val', 0.0)
                    for strat_name in SIX_STRAT_NAMES.values():
                        atr_dict[strat_name] = atr_val
                self.trade_tracker.check_exits(symbol, new_snap.price, atr_dict)
                self.trade_tracker.update_live_pnl(symbol, new_snap.price, self)
            price_fresh = price_updated and new_snap.price > 0.0
            self._data[symbol] = new_snap

            if price_fresh and self.predictor:
                # --- Staleness Guardrail ---
                last_valid_ns = self.pipeline_health.get("scraper_valid_ns", {}).get(symbol, 0)
                
                # If valid indicators haven't updated in 5 minutes (300 seconds), block predictions
                if last_valid_ns > 0 and (now_ns - last_valid_ns) > 300 * 1_000_000_000:
                    new_snap = dataclasses.replace(new_snap, strategy_armed="STALE_DATA")
                    self._data[symbol] = new_snap
                    return # Skip ML predictions to prevent bad entries

                # Fire-and-forget ML predictions — deduplicated per symbol to prevent
                # async task flooding on every WS tick (was causing 2-8s lag bursts)
                if not getattr(self, '_ml_pending', None):
                    self._ml_pending = set()
                if symbol not in self._ml_pending:
                    self._ml_pending.add(symbol)
                    def _run_ml_predictors(sym: str, snap_obj, tracker):
                        try:
                            updated_snap = self.predictor.on_tick_update(sym, snap_obj, tracker)
                            if updated_snap is not None and getattr(updated_snap, 'strategy_armed', None):
                                with self._global_lock:
                                    existing = self._data.get(sym)
                                    if existing:
                                        self._data[sym] = dataclasses.replace(existing, strategy_armed=updated_snap.strategy_armed)
                        except Exception as e:
                            print(f"[ML Predictor] Exception for {sym}: {e}")
                        finally:
                            self._ml_pending.discard(sym)
                    loop = asyncio.get_running_loop()
                    asyncio.create_task(loop.run_in_executor(ML_POOL, _run_ml_predictors, symbol, new_snap, self.trade_tracker))

    def snapshot(self) -> Dict[str, AssetSnapshot]:
        # GIL-atomic shallow copy of dict references; safe for lock-free reads
        return dict(self._data)

# --- BINANCE FOOTPRINT ENGINE ---
TICK_SIZES = {
    "BTCUSDT": 10.0,       # ~0.016% - 12-31 rows
    "ETHUSDT": 0.25,       # ~0.015% - 12-30 rows (was 0.5, too blocky)
    "XRPUSDT": 0.0002,     # ~0.018% - 11-27 rows
    "BNBUSDT": 0.05,       # ~0.009% - 23-58 rows (was 0.1)
    "SOLUSDT": 0.01,       # ~0.016% - 12-32 rows (was 0.02)
    "DOGEUSDT": 0.00002,   # ~0.024% - 8-21 rows
    "ADAUSDT": 0.00003,    # ~0.019% - 10-26 rows (was 0.00005)
    "TRXUSDT": 0.00005,    # ~0.016% - 13-32 rows
    "LINKUSDT": 0.001,     # ~0.013% - 15-38 rows (was 0.002)
    "AVAXUSDT": 0.001,     # ~0.016% - 12-32 rows (was 0.002)
    "XLMUSDT": 0.00005,    # ~0.028% - 7-18 rows
    "HBARUSDT": 0.00002,   # ~0.026% - 8-20 rows
    "LTCUSDT": 0.01,       # ~0.024% - 8-21 rows (was 0.02, critical fix)
    "DOTUSDT": 0.0002,     # ~0.021% - 9-23 rows (was 0.0005, critical fix)
    "XMRUSDT": 0.05,       # ~0.016% - 12-31 rows
    "NEARUSDT": 0.0005,    # ~0.024% - 8-21 rows
    "UNIUSDT": 0.0005,     # ~0.021% - 10-24 rows
    "FILUSDT": 0.0002,     # ~0.027% - 7-19 rows
    "SUIUSDT": 0.0005,     # Tick size for SUI
    "XAUUSDT": 0.25,       # Tick size for Gold
    "XAGUSDT": 0.002,      # Tick size for Silver
    "CLUSDT": 0.005,       # Tick size for Crude Oil
    "NATGASUSDT": 0.0002,  # Tick size for Natural Gas
}

class FootprintCandle:
    """Tracks a single 15m kline candle's delta and volume profile."""
    def __init__(self, tick_size: float):
        self.tick_size = tick_size
        self.candle_open_ms: int = 0
        self.delta: float = 0.0
        self.volume_profile: Dict[float, float] = collections.defaultdict(float)

    def _bucket(self, price: float) -> float:
        return round(price / self.tick_size) * self.tick_size

    def update(self, candle_open_ms: int, buy_vol: float, sell_vol: float, close_price: float) -> None:
        """Called with the latest kline data. Resets automatically on new candle."""
        if candle_open_ms != self.candle_open_ms:
            # New 15m candle opened — reset everything
            self.candle_open_ms = candle_open_ms
            self.delta = 0.0
            self.volume_profile.clear()
        self.delta = buy_vol - sell_vol
        # Volume profile: accumulate on close price bucket (klines don't provide tick-level data)
        bucket = self._bucket(close_price)
        self.volume_profile[bucket] = buy_vol + sell_vol

    @property
    def poc(self) -> float:
        if not self.volume_profile:
            return 0.0
        return max(self.volume_profile.items(), key=lambda kv: kv[1])[0]


class BinanceTradePriceWebSocketFeed:
    def __init__(self, symbols: List[str], store: SnapshotStore):
        self.symbols = [
            s for s in symbols
            if s not in ("XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT")
        ]
        self.store = store
        self.running = True
        self.last_heartbeat_ns = time.time_ns()
        self.last_emit_ns: Dict[str, int] = {}
        self.tab_id = "binance_ws"
        self.skip_watchdog = True  # WS may be network-blocked (ISP); REST feed covers all price needs
        self.clock_offset_ms: float = 0.0
        
    async def sync_clock_offset(self) -> None:
        try:
            loop = asyncio.get_running_loop()
            def _fetch_time():
                import urllib.request, json
                with urllib.request.urlopen("https://fapi.binance.com/fapi/v1/time", timeout=3) as r:
                    return json.loads(r.read().decode())["serverTime"]
            server_time = await loop.run_in_executor(None, _fetch_time)
            self.clock_offset_ms = (time.time() * 1000) - server_time
        except Exception:
            self.clock_offset_ms = 0.0

    async def run(self) -> None:
        if not self.symbols:
            return
            
        # Exclude commodities not traded on Binance Futures fstream
        crypto_symbols = [s for s in self.symbols if s not in ["XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]]
        streams = "/".join(f"{s.lower()}@aggTrade" for s in crypto_symbols)
        is_testnet = os.environ.get("BINANCE_USE_TESTNET", "false").lower() == "true"
        default_base = "wss://stream.binancefuture.com/stream" if is_testnet else "wss://fstream.binance.com/stream"
        url = os.environ.get("BINANCE_WS_URL", f"{default_base}?streams={streams}")
        print(f"[Binance WS] Starting with URL: {url}")
        await self.sync_clock_offset()
        
        while self.running:
            try:
                # Wrap connect with timeout to prevent watchdog hangs
                async with websockets.connect(
                    url,
                    ping_interval=20,
                    ping_timeout=10,
                    open_timeout=15,  # Prevents hanging during connection
                    max_queue=4096,
                ) as ws:
                    print("[Binance WS] Connected trade price stream.")
                    if self.store and hasattr(self.store, 'pipeline_health'):
                        self.store.pipeline_health["binance_ws_status"] = "CONNECTED"
                        self.store.pipeline_health["binance_ws_url"] = url[:60] + "..."
                    async for raw in ws:
                        if not self.running:
                            break
                            
                        self.last_heartbeat_ns = time.time_ns()
                        if self.store and hasattr(self.store, 'pipeline_health'):
                            self.store.pipeline_health["binance_ws_ticks"] += 1
                        
                        try:
                            msg = json.loads(raw)
                            data = msg.get("data", {})
                            sym = data.get("s")
                            p_str = data.get("p")
                            
                            def finite_float_or_none(v):
                                try:
                                    val = float(v)
                                    import math
                                    if math.isfinite(val): return val
                                    return None
                                except:
                                    return None

                            price = finite_float_or_none(p_str)
                            
                            if sym not in self.symbols or price is None or price <= 0:
                                continue
                                
                            # Track WebSocket message queue & processing lag adjusted for system clock offset
                            event_time_ms = data.get("E")
                            if event_time_ms:
                                adjusted_local_ms = (time.time() * 1000) - self.clock_offset_ms
                                lag_sec = max(0.0, (adjusted_local_ms - event_time_ms) / 1000.0)
                                if lag_sec > 3.0 and self.store and hasattr(self.store, 'pipeline_health'):
                                    self.store.pipeline_health["binance_ws_lag"] = round(lag_sec, 2)
                                
                            now_ns = time.time_ns()
                            last_ns = self.last_emit_ns.get(sym, 0)
                            if now_ns - last_ns < 150_000_000:  # 150 ms
                                continue
                            self.last_emit_ns[sym] = now_ns
                            
                            await self.store.update(sym, source="binance_ws", price=price)
                        except Exception as inner_e:
                            continue
                            
            except Exception as e:
                if self.store and hasattr(self.store, 'pipeline_health'):
                    self.store.pipeline_health["binance_ws_status"] = "RECONNECTING"
                print(f"[Binance WS] Disconnected/error: {e}. Reconnecting in 5s...")
                await asyncio.sleep(5.0)


class BinanceFootprintFeed:
    """Polls Binance Futures klines REST API every 5s to derive 15m candle delta and POC."""
    def __init__(self, symbols: List[str], store: SnapshotStore):
        self.symbols = symbols
        self.store = store
        # Exclude commodities not on Binance Futures
        self.valid_symbols = [s for s in symbols if s not in ["XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]]
        self.candles = {}
        for s in self.valid_symbols:
            tick = TICK_SIZES.get(s)
            if tick is None:
                raise KeyError(f"No TICK_SIZE configured for {s}; refusing to guess POC bucket.")
            self.candles[s] = FootprintCandle(tick)
        self.last_heartbeat_ns = time.time_ns()
        self.running = True
        self.consecutive_failures = 0
        self.was_failing = False

    async def run(self) -> None:
        url = "https://fapi.binance.com/fapi/v1/klines"

        async def _fetch_one(session: aiohttp.ClientSession, idx: int, sym: str) -> None:
            try:
                params = {"symbol": sym, "interval": "15m", "limit": 1}
                async with session.get(url, params=params, timeout=aiohttp.ClientTimeout(total=8)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        if data:
                            item = data[-1]
                            candle_open_ms = int(item[0])
                            tot_vol = float(item[5])
                            buy_vol = float(item[9])
                            sell_vol = tot_vol - buy_vol
                            close_price = float(item[4])
                            candle = self.candles[sym]
                            candle.update(candle_open_ms, buy_vol, sell_vol, close_price)
                            await self.store.update(
                                sym,
                                source="binance",
                                price=close_price,
                                open=float(item[1]),
                                high=float(item[2]),
                                low=float(item[3]),
                                close=close_price,
                                fp_delta=candle.delta,
                                fp_poc=candle.poc
                            )
                            successes[idx] = True
            except Exception:
                pass  # individual symbol errors are covered by the outer consolidated warning

        while self.running:
            try:
                # Use ThreadedResolver to bypass asyncio resolver warnings on Windows
                connector = aiohttp.TCPConnector(
                    resolver=aiohttp.ThreadedResolver(),
                    family=socket.AF_INET,
                    keepalive_timeout=60,
                    enable_cleanup_closed=True,
                )
                # Session lives for 30 minutes before being refreshed
                session_start = time.time()
                async with aiohttp.ClientSession(connector=connector) as session:
                    while self.running and (time.time() - session_start) < 1800:
                        self.last_heartbeat_ns = time.time_ns()
                        successes = [False] * len(self.valid_symbols)
                        await asyncio.gather(*[_fetch_one(session, idx, s) for idx, s in enumerate(self.valid_symbols)])
                        
                        if any(successes):
                            if self.was_failing:
                                print("[Binance Feed] [INFO] Connection restored.")
                                self.was_failing = False
                            if self.store and hasattr(self.store, 'pipeline_health'):
                                self.store.pipeline_health["footprint_status"] = "CONNECTED"
                                self.store.pipeline_health["footprint_ticks"] += 1
                            self.consecutive_failures = 0
                        else:
                            self.consecutive_failures += 1
                            if self.consecutive_failures == 1:
                                print("[Binance Feed] [WARN] Connection issues detected (all queries failed).")
                            elif self.consecutive_failures % 30 == 0:
                                print(f"[Binance Feed] [WARN] Connection is still down (consecutive failures: {self.consecutive_failures})")
                            self.was_failing = True
                            
                        await asyncio.sleep(2.0)
            except Exception as e:
                if self.consecutive_failures <= 3:
                    print(f"[Binance Feed] [WARN] Session error: {e}. Retrying in 10s...")
                await asyncio.sleep(10.0)

# --- COINGLASS JS SHIMS ---
INIT_SCRIPT = ""

# JS run inside the TradingView iframe — extracts OHLCV and all indicator legend
SINGLE_FRAME_EXTRACTION_JS = r'''() => {
    try {
        let minusRe = /[\u2212\u2012\u2013\u2014]/g;
        let getTxt = el => el ? (el.innerText || el.textContent || '').replace(minusRe, '-').trim() : '';

        let data = {
            symbol: '',
            price: 'N/A', open: 'N/A', high: 'N/A', low: 'N/A', close: 'N/A', volume: 'N/A',
            rsi: 'N/A', futures_cvd: 'N/A', spot_cvd: 'N/A', funding_rate: '0.0',
            liquidations_long: '0.0', liquidations_short: '0.0', ls_ratio: 'N/A', open_interest: 'N/A',
            whale_index: 'N/A', taker_buy_count: 'N/A', taker_sell_count: 'N/A',
            coins_bid: 'N/A', coins_ask: 'N/A', dollars_bid: 'N/A', dollars_ask: 'N/A',
            ema_8: 'N/A', ema_21: 'N/A', ema_50: 'N/A', ema_200: 'N/A', ema_800: 'N/A',
            atr: 'N/A', atr_14: 'N/A', atr_100: 'N/A'
        };

        // 1. Symbol & Series OHLC
        let seriesEl = document.querySelector('.pane-legend-title, [class*="legendTitle"], [class*="title"], [data-name="legend-series-item"], [class*="series-"], .pane-legend-line:first-child, [class*="legendMainSourceWrapper"]');
        if (seriesEl) {
            let fullText = getTxt(seriesEl);
            let symMatch = fullText.match(/^([A-Z0-9_.]+)/i);
            if (symMatch) {
                let s = symMatch[1].replace('BINANCE_', '').replace('.P', '').replace('Binance_', '').replace('COINGLASS_', '');
                data.symbol = s;
            }

            let valTitles = Array.from(seriesEl.querySelectorAll('[class*="valueTitle-"], .pane-legend-title')).map(t => getTxt(t));
            let valValues = Array.from(seriesEl.querySelectorAll('[class*="valueValue-"], [class*="legendValue"], [class*="lastValue"], .pane-legend-value')).map(v => getTxt(v));

            for (let k = 0; k < valTitles.length; k++) {
                let t = valTitles[k];
                let v = valValues[k];
                if (v && v !== '∅' && v !== 'N/A') {
                    if (t === 'O') data.open = v;
                    else if (t === 'H') data.high = v;
                    else if (t === 'L') data.low = v;
                    else if (t === 'C') data.close = v;
                    else if (t === 'Vol' || t === 'V') data.volume = v;
                }
            }
            if (data.close === 'N/A') {
                let m = fullText.match(/O\s*([0-9.,]+)\s*H\s*([0-9.,]+)\s*L\s*([0-9.,]+)\s*C\s*([0-9.,]+)/i);
                if (m) {
                    data.open = m[1];
                    data.high = m[2];
                    data.low = m[3];
                    data.close = m[4];
                } else {
                    let priceMatch = fullText.match(/\b([0-9]+(?:\.[0-9]+)?)\b/);
                    if (priceMatch) data.close = priceMatch[1];
                }
            }
            if (data.close !== 'N/A') data.price = data.close;
        }

        // 2. Studies & Indicators
        let legends = Array.from(document.querySelectorAll('.pane-legend-item, [class*="legendItem"], [class*="study"], [data-name="legend-source-item"], [class*="legend-"], [class*="Legend-"], [class*="source-"], [class*="item-"], .legend-TG1_J52N'));

        for (let el of legends) {
            if (el === seriesEl) continue;
            let text = getTxt(el);
            if (!text) continue;
            let upper = text.toUpperCase();

            // Clean text extraction of all distinct numbers in the legend line (removing parentheses)
            let cleanedLegendText = text.replace(/\([^)]*\)/g, ' ');
            let allTextNums = (cleanedLegendText.match(/[-+]?[0-9,]+(?:\.[0-9]+)?[KMBkmb%]?/g) || []).filter(s => s && s !== '-' && s !== '+');

            if ((upper.includes('VOLUME') || upper.includes('VOL')) && !upper.includes('DELTA') && !upper.includes('TAKER') && !upper.includes('BID')) {
                if (allTextNums.length > 0) data.volume = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('EMA') || upper.includes('MOVING AVERAGE') || upper.includes('EXPONENTIAL')) {
                let periodMatch = text.match(/\b(8|21|50|200|800)\b/);
                let p = periodMatch ? periodMatch[1] : '';
                let num = allTextNums.length > 0 ? allTextNums[allTextNums.length - 1] : null;
                if (num) {
                    if (p === '8') data.ema_8 = num;
                    else if (p === '21') data.ema_21 = num;
                    else if (p === '50') data.ema_50 = num;
                    else if (p === '200') data.ema_200 = num;
                    else if (p === '800') data.ema_800 = num;
                }
            } else if (upper.includes('FUTURES CUMULATIVE') || (upper.includes('CVD') && !upper.includes('SPOT')) || upper.includes('FUTURES CVD')) {
                if (allTextNums.length > 0) data.futures_cvd = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('SPOT CUMULATIVE') || (upper.includes('CVD') && upper.includes('SPOT')) || upper.includes('SPOT CVD')) {
                if (allTextNums.length > 0) data.spot_cvd = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('RELATIVE STRENGTH') || upper.includes('RSI')) {
                if (allTextNums.length > 0) data.rsi = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('FUNDING') || upper.includes('FUND') || upper.includes('PREDICTED RATE')) {
                if (allTextNums.length > 0) data.funding_rate = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('LIQUIDATION') || upper.includes('LIQ')) {
                // Aggregated Liquidations: Long is positive (1st), Short is negative (2nd)
                if (allTextNums.length >= 2) {
                    data.liquidations_long = allTextNums[0];
                    data.liquidations_short = allTextNums[1];
                } else if (allTextNums.length === 1) {
                    data.liquidations_long = allTextNums[0];
                }
            } else if (upper.includes('LONG/SHORT') || upper.includes('L/S') || upper.includes('LSR') || upper.includes('RATIO')) {
                if (allTextNums.length > 0) data.ls_ratio = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('OPEN INTEREST') || /\bOI\b/.test(upper) || upper.includes('OPEN_INTEREST')) {
                if (allTextNums.length > 0) data.open_interest = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('WHALE') || upper.includes('WHALE INDEX')) {
                if (allTextNums.length > 0) data.whale_index = allTextNums[allTextNums.length - 1];
            } else if (upper.includes('TAKER') || upper.includes('BUY/SELL')) {
                if (allTextNums.length >= 2) {
                    data.taker_buy_count = allTextNums[0];
                    data.taker_sell_count = allTextNums[1];
                }
            } else if (upper.includes('BID & ASK') || upper.includes('BID AND ASK') || upper.includes('BID/ASK') || upper.includes('DEPTH')) {
                // Aggregated Futures Bid & Ask: Has Bid (positive green) and Ask (negative red)
                let validBidAskNums = allTextNums.filter(n => /[KMBkmb%]/.test(n) || Math.abs(parseFloat(n.replace(/,/g, ''))) > 10.0);
                if (validBidAskNums.length < 2) validBidAskNums = allTextNums.slice(-2);
                
                if (upper.includes('DOLLAR') || upper.includes('$')) {
                    if (validBidAskNums.length >= 2) {
                        data.dollars_bid = validBidAskNums[0];
                        data.dollars_ask = validBidAskNums[1];
                    } else if (validBidAskNums.length === 1) {
                        data.dollars_bid = validBidAskNums[0];
                    }
                } else {
                    if (validBidAskNums.length >= 2) {
                        data.coins_bid = validBidAskNums[0];
                        data.coins_ask = validBidAskNums[1];
                    } else if (validBidAskNums.length === 1) {
                        data.coins_bid = validBidAskNums[0];
                    }
                }
            } else if (upper.includes('AVERAGE TRUE RANGE') || upper.includes('ATR')) {
                if (allTextNums.length > 0) {
                    let num = allTextNums[allTextNums.length - 1];
                    data.atr = num;
                    if (upper.includes('100')) data.atr_100 = num;
                    else data.atr_14 = num;
                }
            }
        }

        let rawLegends = legends.map(el => getTxt(el));
        return { success: true, data: data, rawLegends: rawLegends };
    } catch (err) {
        return { success: false, error: (err && err.message) || String(err) };
    }
}
'''

# --- BROWSER SCRAPER PAGE CLASS ---
class CoinglassTab:
    def __init__(self, context: BrowserContext, symbols: List[str], store: SnapshotStore, tab_id: str):
        self.context = context
        self.symbols = symbols
        self.store = store
        self.tab_id = tab_id
        self.is_seeding = False
        self.page: Optional[Page] = None
        self.last_heartbeat_ns = time.time_ns()
        self.running = True
        self._response_tasks: set[asyncio.Task] = set()
        self._cached_frames = []

    async def get_grid_frames(self) -> List[Any]:
        if not self.page or self.page.is_closed():
            return []
        
        frames = []
        for win_idx in range(1, len(self.symbols) + 1):
            f_found = None
            try:
                container_id = f"tv_chart_container_win{win_idx}"
                selector = f"#{container_id}" if win_idx != 1 else f"#{container_id}, #tv_chart_container_main"
                container = self.page.locator(selector).first
                if await container.count() > 0:
                    iframe = container.locator("iframe").first
                    if await iframe.count() > 0:
                        handle = await iframe.element_handle(timeout=300)
                        if handle:
                            f = await handle.content_frame()
                            if f and not f.is_detached():
                                f_found = f
            except Exception:
                pass
            frames.append(f_found)

        # Fallback: fill any missing frame slots from valid page frames
        valid_frames = [f for f in self.page.frames if f != self.page.main_frame and not f.is_detached()]
        for i in range(len(frames)):
            if frames[i] is None and i < len(valid_frames):
                frames[i] = valid_frames[i]

        return [f for f in frames if f is not None]

    async def ensure_all_cells_15m(self) -> None:
        """Iterate through all 9 grid chart cells and guarantee 15m timeframe is selected."""
        if not self.page or self.page.is_closed():
            return
        print(f"[{self.tab_id}] Verifying and enforcing 15m timeframe across all 9 grid cells...")
        frames = await self.get_grid_frames()
        for idx, frame in enumerate(frames):
            try:
                # Click canvas inside frame to focus this specific chart cell
                canvas = frame.locator("canvas").nth(1)
                if await canvas.count() > 0:
                    await canvas.click(position={"x": 300, "y": 50}, force=True, timeout=3000)
                    await asyncio.sleep(0.3)

                # 1. Try TradingView API inside the iframe
                await frame.evaluate("""() => {
                    if (typeof tradingViewApi !== 'undefined' && tradingViewApi.activeChart) {
                        let ac = tradingViewApi.activeChart();
                        if (typeof ac.setResolution === 'function') {
                            ac.setResolution('15', () => {});
                        }
                    }
                    if (typeof window.tvWidget !== 'undefined' && window.tvWidget.activeChart) {
                        let ac = window.tvWidget.activeChart();
                        if (typeof ac.setResolution === 'function') {
                            ac.setResolution('15', () => {});
                        }
                    }
                }""")

                # 2. Click the 15m toolbar button on the main page while this cell is focused
                btn15 = self.page.locator("button:has-text('15m'), div:has-text('15m'), [data-value='15']").first
                if await btn15.count() > 0 and await btn15.is_visible():
                    await btn15.click(timeout=1500)
                    await asyncio.sleep(0.2)

                # 3. Native TradingView keyboard resolution shortcut (15 + Enter)
                await self.page.keyboard.type("15")
                await asyncio.sleep(0.1)
                await self.page.keyboard.press("Enter")
                await asyncio.sleep(0.3)
                print(f"[{self.tab_id}] Cell {idx+1}/9 locked to 15m timeframe.")
            except Exception as ex:
                print(f"[{self.tab_id}] [WARN] Timeframe lock for cell {idx+1} bypassed: {ex}")



    async def start(self) -> None:
        coinglass_pages = [p for p in self.context.pages if not p.is_closed() and "coinglass" in p.url.lower()]
        all_pages = [p for p in self.context.pages if not p.is_closed() and not p.url.startswith("devtools://")]
        
        target_idx = 0
        if len(coinglass_pages) > target_idx:
            self.page = coinglass_pages[target_idx]
            print(f"[{self.tab_id}] Attached to existing CoinGlass page ({target_idx+1}/{len(coinglass_pages)}): {self.page.url}")
        elif len(all_pages) > target_idx:
            self.page = all_pages[target_idx]
            print(f"[{self.tab_id}] Attached to existing browser page {target_idx+1}: {self.page.url}")
        else:
            print(f"[{self.tab_id}] Creating new page for {self.tab_id}...")
            self.page = await self.context.new_page()

        # Open layout page
        print(f"[{self.tab_id}] Navigating to layout page...")
        await self.page.goto("https://www.coinglass.com/tv/layout/s9", timeout=60000)
        await asyncio.sleep(10.0)

        # Check if redirected to login page
        if "login" in self.page.url.lower():
            print(f"[{self.tab_id}] Redirected to login page. Performing automated login...")
            try:
                email_box = self.page.get_by_role("textbox", name="Email")
                if await email_box.is_visible(timeout=5000):
                    await email_box.click()
                    await email_box.fill("singhkaranbir0248@gmail.com")
                    await email_box.press("Tab")
                    await self.page.get_by_role("textbox", name="Password").click()
                    await self.page.get_by_role("textbox", name="Password").fill("Lu$er2hero")
                    await self.page.get_by_role("button", name="Login").nth(1).click()
                    await asyncio.sleep(5.0)

                    # Return to layout page
                    print(f"[{self.tab_id}] Re-navigating to layout page after login...")
                    await self.page.goto("https://www.coinglass.com/tv/layout/s9", timeout=60000)
                    await asyncio.sleep(10.0)
            except Exception as login_err:
                print(f"[{self.tab_id}] [Setup] Automated login failed/bypassed: {login_err}")

        # Check if we need to load layout L_1 (if it's not already loaded)
        try:
            layout_btn = self.page.get_by_role("button").filter(has_text=re.compile(r"^$")).nth(3)
            if await layout_btn.is_visible(timeout=5000):
                print(f"[{self.tab_id}] Triggering load for custom layout L_1...")
                await layout_btn.click()
                await self.page.get_by_role("menuitem", name="Load Chart Layout").click()
                await self.page.get_by_role("button", name="L_1").click()
                await asyncio.sleep(10.0)
        except Exception as layout_err:
            print(f"[{self.tab_id}] Custom layout L_1 loading bypassed: {layout_err}")

        # Ensure 15m resolution across all 9 grid chart cells
        await self.ensure_all_cells_15m()
        
        # Suppress noisy TradingView internal console spam; only print errors and CoinGlass messages
        def _on_console(msg):
            text = msg.text
            typ = msg.type
            skip_patterns = (
                "Recurring script engine stop",
                "76 custom indicators loaded",
                "Content Security Policy",
                "WebSocket connection to",
                "ERR_NAME_NOT_RESOLVED",
                "502",
                "wss.coinglass.com",
                "net::ERR_",
                "Failed to fetch",
                "unload is not allowed",
            )
            if any(p in text for p in skip_patterns):
                return
            if typ in ("error", "warning") or "coinglass" in text.lower():
                log_live_event(f"{self.tab_id} {typ}: {text[:60]}", "Chrome")

        def _on_page_error(exc):
            msg = str(exc)
            # Filter generic browser resource errors that are not actionable
            if any(p in msg for p in ("unknown compression", "net::", "ERR_", "Failed to fetch", "ResizeObserver", "reading 'symbol'")):
                return
            log_live_event(f"{self.tab_id} page error: {msg[:60]}", "Chrome")

        self.page.on("console", _on_console)
        self.page.on("pageerror", _on_page_error)

        # Primary CDP Network Interception: Intercept CoinGlass structured JSON API responses
        async def _on_response(response):
            try:
                url = response.url.lower()
                if not any(k in url for k in ("coinglass.com/api", "openinterest", "fundingrate", "cvd", "liquidation", "takervolume", "longshortaccount", "fr-chart", "liq-chart")):
                    return
                self.store.pipeline_health[f"{self.tab_id}_last_heartbeat"] = time.time()
                payload = await response.json()
                await self._route_payload({"url": response.url, "body": json.dumps(payload)})
            except Exception:
                pass

        self.page.on("response", lambda res: asyncio.create_task(_on_response(res)))
        
        # Intercept HTTP API responses natively to capture Open Interest and Funding Rates securely
        async def handle_response(response):
            try:
                url = response.url
                if any(k in url for k in ("open-interest", "funding-rate", "liquidation", "long-short", "rsi", "cumulative-volume")):
                    body = await response.text()
                    await self._route_payload({"url": url, "body": body})
            except Exception:
                pass

        def _spawn_response_task(response):
            task = asyncio.create_task(handle_response(response))
            self._response_tasks.add(task)
            task.add_done_callback(self._response_tasks.discard)

        self.page.on("response", _spawn_response_task)
        
        print(f"[{self.tab_id}] Opening layout: {URL}...")
        await self.page.bring_to_front()
        for attempt in range(3):
            try:
                await self.page.goto(URL, wait_until="domcontentloaded", timeout=45000)
                break
            except Exception as e:
                print(f"[{self.tab_id}] [WARN] goto attempt {attempt+1} exception: {e}")
                if attempt == 2:
                    raise
                await asyncio.sleep(2.0)
        print(f"[{self.tab_id}] Waiting for chart layout to mount...")
        try:
            await self.page.wait_for_selector("iframe", state="attached", timeout=25000)
        except Exception:
            pass
        await asyncio.sleep(5.0)

    async def reconnect(self, focus_lock: asyncio.Lock) -> None:
        log_live_event(f"{self.tab_id} reconnecting/restarting tab...", "Recovery")
        self.is_seeding = True
        try:
            self.running = False
            self._cached_frames = []
            if self.page and not self.page.is_closed():
                try:
                    await self.page.close()
                except Exception:
                    pass
            self.running = True
            await self.start()
            await self.inject_and_configure_all(focus_lock)
            log_live_event(f"{self.tab_id} tab restarted and re-configured", "Recovery")
            self.last_heartbeat_ns = time.time_ns()
            self.poll_failures = 0
        except Exception as e:
            log_live_event(f"{self.tab_id} recovery failed: {str(e)[:50]}", "Recovery")
        finally:
            self.is_seeding = False

    async def inject_and_configure_all(self, focus_lock: asyncio.Lock):
        """Symbol & Resolution configuration using Playwright UI interactions.
        
        Performs grid focusing, setting resolution to 15m, and filling/selecting 
        the exact symbol from the CoinGlass search dropdown natively.
        """
        print(f"[{self.tab_id}] Bringing tab to front...")
        await self.page.bring_to_front()
        await asyncio.sleep(1.0)

        # Discover grid iframes
        print(f"[{self.tab_id}] Discovering grid frames...")
        frames = []
        for _ in range(30):
            frames = await self.get_grid_frames()
            if len(frames) >= len(self.symbols):
                break
            await asyncio.sleep(1.0)

        if not frames:
            print(f"[{self.tab_id}] No grid frames found. Skipping layout configuration.")
            return

        print(f"[{self.tab_id}] Found {len(frames)} grid frames. Configuring symbols and timeframes via UI...")

        # Get frame names in grid order from DOM
        print(f"[{self.tab_id}] Mapping grid frames to DOM order...")
        try:
            iframe_names = await self.page.evaluate("""() => {
                let list = [];
                for (let i = 1; i <= 9; i++) {
                    let id = 'tv_chart_container_win' + i;
                    let container = document.getElementById(id);
                    if (i === 1 && !container) {
                        container = document.getElementById('tv_chart_container_main');
                    }
                    if (container) {
                        let f = container.querySelector('iframe');
                        if (f && f.name) list.push(f.name);
                    }
                }
                return list;
            }""")
            print(f"[{self.tab_id}] Ordered frame names: {iframe_names}")
        except Exception as e:
            print(f"[{self.tab_id}] [ERROR] Failed to map DOM frame names: {e}")
            iframe_names = []

        # Iterate and configure each cell via UI
        for i in range(min(len(self.symbols), len(frames))):
            sym = self.symbols[i]
            print(f"[{self.tab_id}] Configuring cell {i+1}/9: {sym}")
            try:
                # Find matching frame from discovered frames
                frame = None
                if i < len(iframe_names):
                    name = iframe_names[i]
                    matching = [f for f in frames if f.name == name]
                    if matching:
                        frame = matching[0]
                if not frame:
                    print(f"[{self.tab_id}] [WARN] Bypassing grid order for cell {i+1}, using fallback frame")
                    frame = frames[i]

                # Focus the specific grid frame by clicking the canvas
                # Click the canvas at coordinate (300, 50) inside the frame content
                await frame.locator("canvas").nth(1).click(position={"x": 300, "y": 50}, force=True, timeout=10000)
                await asyncio.sleep(0.8)

                # Set 15m resolution for this specific chart window
                try:
                    await frame.evaluate("""() => {
                        if (typeof tradingViewApi !== 'undefined' && tradingViewApi.activeChart) {
                            let ac = tradingViewApi.activeChart();
                            if (typeof ac.setResolution === 'function') {
                                ac.setResolution('15', () => {});
                            }
                        }
                    }""")
                    # Also trigger 15m toolbar button on page level if available
                    btn15 = self.page.locator("button:has-text('15m'), div:has-text('15m')").first
                    if await btn15.count() > 0:
                        await btn15.click(timeout=1500)
                except Exception:
                    pass
                await asyncio.sleep(0.5)

                # Change symbol via UI search
                # Click global search button (first button in toolbar)
                await self.page.get_by_role("button").first.click()
                await asyncio.sleep(0.8)
                
                # Fill the symbol input
                await self.page.locator("#tv-ss").fill(sym)
                await asyncio.sleep(1.2)
                
                # Select the search result button matching "Binance sym" (case-insensitive)
                result_btn = self.page.get_by_role("button", name=re.compile(f"Binance {sym}", re.I)).first
                await result_btn.click()
                await asyncio.sleep(3.0) # Wait for symbol loading to settle
                print(f"[{self.tab_id}] Cell {i+1} configured successfully for {sym} (15m)")
            except Exception as e:
                print(f"[{self.tab_id}] [ERROR] Failed to configure cell {i+1} for {sym}: {e}")

        print(f"[{self.tab_id}] Grid symbol configuration complete.")


    async def poll_loop(self) -> None:
        """Background data poller extracting DOM legend values & JS shims"""
        _poll_count = 0
        _poll_start_ns = time.time_ns()
        _last_proactive_reload = time.time()
        PROACTIVE_RELOAD_INTERVAL = 1800  # 30 minutes - reset TradingView canvas throttle

        field_map = {
            "volume": "volume", "open_interest": "oi",
            "funding_rate": "funding", "ls_ratio": "ls_ratio",
            "futures_cvd": "fut_cvd", "spot_cvd": "spot_cvd",
            "liquidations_long": "liq_long", "liquidations_short": "liq_short",
            "coins_bid": "coins_bid", "coins_ask": "coins_ask",
            "dollars_bid": "dollars_bid", "dollars_ask": "dollars_ask",
            "whale_index": "whale_idx",
            "taker_buy_count": "tk_buy_cnt", "taker_sell_count": "tk_sell_cnt", "taker_delta": "tk_delta",
            "ema_8": "ema_8", "ema_21": "ema_21", "ema_50": "ema_50",
            "ema_200": "ema_200", "ema_800": "ema_800",
            "atr_14": "atr_14", "atr_100": "atr_100", "atr": "atr",
        }

        while self.running:
            self.last_heartbeat_ns = time.time_ns()
            if not self.page or self.page.is_closed():
                await asyncio.sleep(1.0)
                continue

            # Proactive page reload every 30 minutes to prevent TradingView canvas throttling
            if time.time() - _last_proactive_reload > PROACTIVE_RELOAD_INTERVAL:
                log_live_event("30-min page reload to prevent canvas throttling...", self.tab_id)
                _last_proactive_reload = time.time()
                try:
                    if hasattr(self, 'focus_lock') and self.focus_lock:
                        await self.reconnect(self.focus_lock)
                    elif self.page and not self.page.is_closed():
                        await self.page.reload(wait_until="load", timeout=30000)
                    self.poll_failures = 0
                    _poll_count = 0
                    _poll_start_ns = time.time_ns()
                except Exception as ex:
                    log_live_event(f"Reload failed: {ex}", self.tab_id)

            try:
                success_count = 0
                frame_errors = 0
                last_frame_err = ""
                frames = await self.get_grid_frames()

                for frame_idx, frame in enumerate(frames):
                    try:
                        res = await asyncio.wait_for(frame.evaluate(SINGLE_FRAME_EXTRACTION_JS), timeout=4.0)
                    except Exception as fe:
                        frame_errors += 1
                        last_frame_err = str(fe)[:80]
                        continue
                    if not res or not res.get("success"):
                        frame_errors += 1
                        last_frame_err = (res or {}).get("error", "no success flag")[:80]
                        continue

                    d = res.get("data", {})
                    sym_raw = (d.get("symbol") or "").strip().upper()
                    for prefix in ("BINANCE_", "BINANCE:", "BINANCE-", "COINGLASS_", "COINGLASS:", "BYBIT:", "OKX:"):
                        sym_raw = sym_raw.replace(prefix, "")
                    sym_clean = sym_raw.split(".")[0].replace("/", "").replace("-", "").strip()

                    sym_actual = None
                    if sym_clean in ALL_SYMBOLS:
                        sym_actual = sym_clean
                    elif (sym_clean + "USDT") in ALL_SYMBOLS:
                        sym_actual = sym_clean + "USDT"
                    else:
                        alias_map = {
                            "XAUUSD": "XAUUSDT", "GOLD": "XAUUSDT", "XAU": "XAUUSDT",
                            "XAGUSD": "XAGUSDT", "SILVER": "XAGUSDT", "XAG": "XAGUSDT",
                            "CLUSD": "CLUSDT", "CRUDE": "CLUSDT", "OIL": "CLUSDT", "CL": "CLUSDT",
                            "NGUSD": "NATGASUSDT", "NATGAS": "NATGASUSDT", "NG": "NATGASUSDT",
                            "BTC": "BTCUSDT", "ETH": "ETHUSDT", "XRP": "XRPUSDT", "SOL": "SOLUSDT",
                            "BNB": "BNBUSDT", "DOGE": "DOGEUSDT", "ADA": "ADAUSDT", "TRX": "TRXUSDT",
                            "LINK": "LINKUSDT", "AVAX": "AVAXUSDT", "SUI": "SUIUSDT", "NEAR": "NEARUSDT",
                            "DOT": "DOTUSDT", "LTC": "LTCUSDT"
                        }
                        sym_actual = alias_map.get(sym_clean)

                    if not sym_actual and frame_idx < len(self.symbols):
                        sym_actual = self.symbols[frame_idx]

                    if not sym_actual:
                        continue

                    update_kwargs = {}
                    for js_key, py_key in field_map.items():
                        raw = d.get(js_key)
                        if raw is None or str(raw).strip().lower() == "n/a":
                            continue
                        fv = finite_float_or_none(raw)
                        if fv is not None:
                            update_kwargs[py_key] = fv

                    raw_rsi = d.get("rsi")
                    if raw_rsi is not None and str(raw_rsi).strip().lower() != "n/a":
                        rsi_val = parse_float(raw_rsi)
                        if rsi_val not in (100.0, 0.0):
                            update_kwargs["rsi"] = rsi_val

                    price_val = parse_float(d.get("close") or d.get("price") or 0.0)
                    if price_val > 0:
                        update_kwargs["price"] = price_val

                    if update_kwargs:
                        await self.store.update(sym_actual, source="coinglass", **update_kwargs)
                        success_count += 1

                has_success = success_count > 0

                if has_success:
                    self.last_heartbeat_ns = time.time_ns()
                    self.poll_failures = 0
                    _poll_count += 1
                    if _poll_count == 1:
                        log_live_event(f"First successful DOM poll! Scraped {success_count}/{len(self.symbols)} symbols.", self.tab_id)
                    if self.store and hasattr(self.store, 'pipeline_health'):
                        now_ns = time.time_ns()
                        elapsed_s = max(1.0, (now_ns - _poll_start_ns) / 1e9)
                        real_fps = round(_poll_count / elapsed_s, 2)
                        self.store.pipeline_health["chrome_polls"] = self.store.pipeline_health.get("chrome_polls", 0) + 1
                        self.store.pipeline_health["chrome_status"] = "CONNECTED"
                        self.store.pipeline_health["chrome_latency_ms"] = 35.0
                        self.store.pipeline_health["scraper_last_parse_ns"] = now_ns
                        self.store.pipeline_health["scraper_fps"] = real_fps
                        self.store.pipeline_health["scraper_frame_ok"] = f"{success_count}/{len(self.symbols)}"
                        self.store.pipeline_health["chrome_poll_failures"] = self.poll_failures

                    # Save periodic full-resolution layout snapshot directly from Playwright
                    if _poll_count % 60 == 1:
                        try:
                            save_dir = os.path.join(base_dir, "live_data", "desktop_screenshots")
                            os.makedirs(save_dir, exist_ok=True)
                            tab_name = "latest_chrome_tab1.png" if self.tab_id == "TAB_1" else "latest_chrome_tab2.png"
                            await self.page.screenshot(path=os.path.join(save_dir, tab_name))
                        except Exception:
                            pass
                else:
                    self.poll_failures += 1
                    if self.store and hasattr(self.store, 'pipeline_health'):
                        self.store.pipeline_health["chrome_poll_failures"] = self.poll_failures
                    if self.poll_failures % 10 == 1:
                        diag = f"frames={len(frames)}, errs={frame_errors}"
                        if last_frame_err:
                            diag += f", last_err={last_frame_err[:50]}"
                        log_live_event(f"{self.tab_id} poll miss #{self.poll_failures}: {diag}", "Scraper")
            except Exception as e:
                self.poll_failures += 1
                if self.store and hasattr(self.store, 'pipeline_health'):
                    self.store.pipeline_health["chrome_poll_failures"] = self.poll_failures
                log_live_event(f"{self.tab_id} poll exception: {str(e)[:60]}", "Scraper")

            frozen = False
            now_ns = time.time_ns()
            if self.store and hasattr(self.store, 'pipeline_health'):
                indicator_ns = self.store.pipeline_health.get("scraper_valid_ns", {})
                frozen_syms = []
                for sym in self.symbols:
                    last_ind = indicator_ns.get(sym, 0)
                    if last_ind > 0 and (now_ns - last_ind) > 120 * 1_000_000_000:
                        frozen_syms.append(sym)
                if len(frozen_syms) >= max(1, len(self.symbols) // 2):
                    frozen = True

            last_heal = getattr(self, '_last_heal_ns', 0)
            if (self.poll_failures > 60 or frozen) and (now_ns - last_heal) > 60 * 1_000_000_000:
                self._last_heal_ns = now_ns
                reason = "Frozen indicators" if frozen else f"Max failures ({self.poll_failures})"
                log_live_event(f"[WATCHDOG] {reason}. Auto-healing tab...", self.tab_id)
                
                # Reset failure count and timestamps immediately to prevent rapid-fire loop spam
                self.poll_failures = 0
                _poll_count = 0
                _poll_start_ns = now_ns
                if self.store and hasattr(self.store, 'pipeline_health'):
                    for sym in self.symbols:
                        self.store.pipeline_health.setdefault("last_change_ns", {})[sym] = now_ns
                        self.store.pipeline_health.setdefault("scraper_valid_ns", {})[sym] = now_ns

                try:
                    if hasattr(self, 'focus_lock') and self.focus_lock:
                        await self.reconnect(self.focus_lock)
                    elif self.page and not self.page.is_closed():
                        await self.page.reload(wait_until="load", timeout=30000)
                except Exception as ex:
                    log_live_event(f"[WATCHDOG] Auto-heal exception: {ex}", self.tab_id)

            await asyncio.sleep(0.5)

    async def _route_payload(self, entry: dict) -> None:
        url = entry.get("url", "")
        body = entry.get("body", "")
        try:
            payload = json.loads(body)
        except Exception:
            return
        
        url_lower = url.lower()
        # Route to appropriate update target
        if any(k in url_lower for k in ("open-interest", "openinterest", "/oi", "open_interest")):
            await self._apply(payload, "oi")
        elif any(k in url_lower for k in ("funding-rate", "fundingrate", "funding", "fr-chart")):
            await self._apply(payload, "funding")
        elif any(k in url_lower for k in ("liquidation", "/liq", "liq-chart")):
            await self._apply_liq(payload)
        elif any(k in url_lower for k in ("long-short", "longshort", "ls_ratio", "ls-rate")):
            await self._apply(payload, "ls_ratio")
        elif any(k in url_lower for k in ("cumulative-volume", "cvd", "volumedelta")):
            if "futures" in url_lower:
                await self._apply(payload, "fut_cvd")
            else:
                await self._apply(payload, "spot_cvd")
        elif "rsi" in url_lower:
            await self._apply(payload, "rsi")

    async def _apply(self, payload: Any, field_name: str) -> None:
        data = payload.get("data", [])
        if isinstance(data, list):
            for row in data:
                sym = row.get("symbol")
                if sym in self.symbols:
                    val = parse_float(row.get("value", 0.0))
                    await self.store.update(sym, source="coinglass", **{field_name: val})
        elif isinstance(data, dict):
            for sym, val in data.items():
                if sym in self.symbols:
                    await self.store.update(sym, source="coinglass", **{field_name: parse_float(val)})

    async def _apply_liq(self, payload: Any) -> None:
        data = payload.get("data", [])
        if isinstance(data, list):
            for row in data:
                sym = row.get("symbol")
                if sym in self.symbols:
                    long_liq = parse_float(row.get("longLiq", 0.0))
                    short_liq = parse_float(row.get("shortLiq", 0.0))
                    await self.store.update(sym, source="coinglass", liq_long=long_liq, liq_short=short_liq)

    async def seed_symbol(self, symbol: str, excel_executor, focus_lock: asyncio.Lock) -> None:
        """Performs visual backward walk to collect 50 candles and export to Excel"""
        self.is_seeding = True
        win_idx = self.symbols.index(symbol) + 1
        container_id = f"tv_chart_container_win{win_idx}"
        selector = f"#{container_id}" if win_idx != 1 else f"#{container_id}, #tv_chart_container_main"
        container = self.page.locator(selector).first
        
        async with focus_lock:
            print(f"[{self.tab_id}] Seeding {symbol} in Window {win_idx}. Acquired focus lock. Bringing tab to front...")
            await self.page.bring_to_front()
            await asyncio.sleep(0.5)
            
            iframe = container.locator("iframe").first
            try:
                await iframe.wait_for(state="attached", timeout=25000)
            except Exception:
                pass
            iframe_handle = await iframe.element_handle(timeout=20000)
            if not iframe_handle:
                print(f"[{self.tab_id}] [ERROR] No iframe handle for seeding {symbol}")
                return
            frame = await iframe_handle.content_frame()
            if not frame:
                print(f"[{self.tab_id}] [ERROR] Content frame missing for seeding {symbol}")
                return
                
            # Resolve the first canvas inside the frame
            canvas = frame.locator("canvas").first
            try:
                await canvas.wait_for(state="visible", timeout=5000)
            except Exception:
                print(f"[{self.tab_id}] [ERROR] Canvas element not visible for {symbol}")
                return
                
            # Click canvas center to focus TradingView inner context
            await canvas.click(force=True, timeout=5000)
            await asyncio.sleep(0.3)
            
            # Explicitly focus the window/document body
            await frame.evaluate("() => { window.focus(); if (document.body) document.body.focus(); }")
            await asyncio.sleep(0.2)
            
            # Press Escape to close any potential dialogs
            await self.page.keyboard.press("Escape")
            await asyncio.sleep(0.2)
            
            # Reset visual using Alt+r
            await self.page.keyboard.press("Alt+r")
            await asyncio.sleep(1.0)
            
            # Wait for canvas to become visible/attached again after chart reset
            try:
                await canvas.wait_for(state="visible", timeout=5000)
            except Exception:
                pass
            
            # Right-click canvas to open context menu (forces browser focus delegation)
            await canvas.click(button="right", force=True, timeout=5000)
            await asyncio.sleep(0.5)
            await self.page.keyboard.press("Escape")
            await asyncio.sleep(0.2)
            
            # Wait up to 10 seconds for indicators to load historical data from network
            print(f"[{self.tab_id}] waiting for indicators to populate historical data for {symbol}...")
            for attempt in range(20):
                res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                if res and res.get("success"):
                    d = res["data"]
                    if (d.get("volume") not in ("N/A", "0", None) and
                        d.get("rsi") not in ("N/A", "100.00", None) and
                        d.get("futures_cvd") != "N/A" and
                        d.get("spot_cvd") != "N/A" and
                        d.get("open_interest") != "N/A"):
                        print(f"[{self.tab_id}] Indicators populated in {attempt * 0.5:.1f}s")
                        break
                await asyncio.sleep(0.5)

            rect = await canvas.bounding_box()
            if not rect:
                print(f"[{self.tab_id}] [ERROR] Cannot get canvas bounding box for {symbol}")
                return

            x_pos = rect["x"] + rect["width"] - 60
            y_pos = rect["y"] + rect["height"] * 0.5

            # Hover and click to focus on the rightmost section of the canvas
            await self.page.mouse.move(x_pos, y_pos)
            await self.page.mouse.click(x_pos, y_pos)
            await asyncio.sleep(0.2)

            # ArrowLeft snaps crosshair to the latest candle
            await self.page.keyboard.press("ArrowLeft")
            await asyncio.sleep(0.3)

            # --- Dynamic Gap Calculation ---
            target_steps = 850
            existing_rows = []
            base_dir = os.path.dirname(os.path.abspath(__file__))
            combined_path = os.path.join(base_dir, "Seeding", "combined_seed_history.xlsx")
            if os.path.exists(combined_path):
                import pandas as pd
                try:
                    df = pd.read_excel(combined_path, sheet_name=symbol)
                    if not df.empty and "open_time" in df.columns:
                        existing_rows = df.to_dict('records')
                        
                        # Handle potential datetime vs int vs str timestamp differences
                        for r in existing_rows:
                            val = r.get("open_time")
                            if hasattr(val, "timestamp"):
                                from datetime import timezone
                                r["open_time"] = int(val.replace(tzinfo=timezone.utc).timestamp())
                            elif isinstance(val, (int, float)):
                                r["open_time"] = int(val)
                            elif isinstance(val, str):
                                try:
                                    val_clean = val.replace(" IST", "").strip()
                                    dt = pd.to_datetime(val_clean)
                                    from datetime import timedelta
                                    dt_utc = dt - timedelta(hours=5, minutes=30)
                                    r["open_time"] = int(dt_utc.timestamp())
                                except Exception:
                                    try:
                                        r["open_time"] = int(float(val))
                                    except Exception:
                                        pass
                        
                        # Filter out invalid open_times for max calc
                        valid_times = [r["open_time"] for r in existing_rows if isinstance(r.get("open_time"), int)]
                        if valid_times:
                            latest_time = max(valid_times)
                            current_time = int((time.time() // 900) * 900)
                            gap_candles = calculate_commodity_gap(symbol, latest_time, current_time)
                            existing_count = len(existing_rows)
                            
                            if existing_count + gap_candles >= 850:
                                target_steps = min(gap_candles + 2, 850)
                            else:
                                target_steps = 850
                                
                            print(f"\n==================================================")
                            print(f"[{self.tab_id}] {symbol} SEEDING DATABASE CHECK:")
                            print(f"[{self.tab_id}] Database has {existing_count} candles.")
                            print(f"[{self.tab_id}] Gap from offline time: {gap_candles} missing candles (calendar adjusted).")
                            print(f"[{self.tab_id}] -> WILL SCRAPE {target_steps} CANDLES NOW TO WARM UP.")
                            print(f"==================================================\n")
                        else:
                            print(f"\n==================================================")
                            print(f"[{self.tab_id}] {symbol} Found Excel sheet but no valid `open_time` ints parsed.")
                            print(f"[{self.tab_id}] -> WILL SCRAPE {target_steps} CANDLES NOW TO WARM UP.")
                            print(f"==================================================\n")
                except Exception as e:
                    print(f"[{self.tab_id}] [WARN] Could not read existing seed for {symbol}: {e}")
            else:
                print(f"\n==================================================")
                print(f"[{self.tab_id}] {symbol} No existing seed history found in Excel.")
                print(f"[{self.tab_id}] -> WILL SCRAPE {target_steps} CANDLES NOW TO WARM UP.")
                print(f"==================================================\n")
            # -------------------------------

            candles = collections.deque(maxlen=1000)
            stalls = 0
            debug_dicts = []
            
            last_key = None
            is_crypto = symbol not in ["XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]
            
            computed_timestamps = get_historical_timestamps(symbol, int((time.time() // 900) * 900), target_steps)

            for step in range(target_steps * 2):
                if len(candles) >= target_steps:
                    print(f"[{self.tab_id}] {symbol} Reached target {target_steps} candles. Stopping walk.")
                    break
                    
                if step % 20 == 0:
                    print(f"[{self.tab_id}] Seeding {symbol}: candle {len(candles)}/{target_steps}...")
                
                # 1. Wait for DOM to update after moving crosshair (avoid duplicate read stalls)
                d = None
                for attempt in range(6):
                    res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                    if res and res.get("success"):
                        temp_d = res["data"]
                        close = parse_float(temp_d.get("close", temp_d.get("price", 0.0)))
                        volume = parse_float(temp_d.get("volume", 0.0))
                        rsi = parse_float(temp_d.get("rsi", 50.0))
                        val_key = (close, volume, rsi)
                        
                        if val_key != last_key:
                            d = temp_d
                            last_key = val_key
                            break
                    await asyncio.sleep(0.04)

                if d is None:
                    stalls += 1
                    if stalls > 4:
                        print(f"[{self.tab_id}] [WARN] Seeding stalled for {symbol} at step {step}. Ending early.")
                        break
                    # Recover visual focus delegation
                    await canvas.focus()
                    await asyncio.sleep(0.05)
                    await self.page.keyboard.press("ArrowLeft")
                    await asyncio.sleep(0.1)
                    continue
                
                stalls = 0
                
                # 2. Wait up to 600ms for lazy-loaded indicators (CVD, OI) to populate if they are currently N/A
                if is_crypto:
                    for load_attempt in range(4):
                        if (d.get("futures_cvd") != "N/A" and 
                            d.get("spot_cvd") != "N/A" and 
                            d.get("open_interest") != "N/A"):
                            break
                        await asyncio.sleep(0.15)
                        res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                        if res and res.get("success"):
                            d = res["data"]
                            
                if symbol == "BTCUSDT":
                    debug_dicts.append({
                        "step": step,
                        "data": d,
                        "rawLegends": res.get("rawLegends", []) if res else []
                    })
                    
                candle_data = {
                    "open_time": computed_timestamps[len(candles)],
                    "open":       parse_float(d.get("open",   0.0)),
                    "high":       parse_float(d.get("high",   0.0)),
                    "low":        parse_float(d.get("low",    0.0)),
                    "close":      parse_float(d.get("close",  d.get("price", 0.0))),
                    "volume":     parse_float(d.get("volume", 0.0)),
                    "rsi":        parse_float(d.get("rsi",    50.0)),
                    "fut_cvd":    parse_float(d.get("futures_cvd",      0.0)),
                    "spot_cvd":   parse_float(d.get("spot_cvd",         0.0)),
                    "funding":    parse_float(d.get("funding_rate",      0.0)),
                    "liq_long":   abs(parse_float(d.get("liquidations_long",  0.0))),
                    "liq_short":  abs(parse_float(d.get("liquidations_short", 0.0))),
                    "ls_ratio":   parse_float(d.get("ls_ratio",           1.0)),
                    "oi":         parse_float(d.get("open_interest",      0.0)),
                    "coins_bid":  parse_float(d.get("coins_bid", 0.0)),
                    "coins_ask":  parse_float(d.get("coins_ask", 0.0)),
                    "dollars_bid": parse_float(d.get("dollars_bid", 0.0)),
                    "dollars_ask": parse_float(d.get("dollars_ask", 0.0)),
                    "whale_idx":  parse_float(d.get("whale_index", 0.0)),
                    "tk_buy_cnt": parse_float(d.get("taker_buy_count", 0.0)),
                    "tk_sell_cnt": parse_float(d.get("taker_sell_count", 0.0)),
                    "tk_delta": parse_float(d.get("taker_delta", 0.0)),
                }
                
                candles.appendleft(candle_data)
                
                # Step left — move crosshair one candle back
                await self.page.keyboard.press("ArrowLeft")
                await asyncio.sleep(0.08)

            # Restore view
            await self.page.keyboard.press("Alt+r")
            await asyncio.sleep(0.5)

            scraped_list = list(candles)
            final_list = scraped_list
            if existing_rows:
                all_data = existing_rows + scraped_list
                # Deduplicate by open_time, keeping newest (scraped over existing due to order)
                dedup = {r["open_time"]: r for r in all_data if isinstance(r.get("open_time"), int)}
                sorted_vals = sorted(dedup.values(), key=lambda x: x["open_time"])
                final_list = sorted_vals

            loop = asyncio.get_running_loop()
            await loop.run_in_executor(excel_executor, _dump_xlsx, symbol, final_list)
            
            if self.store.predictor:
                self.store.predictor.set_history(symbol, final_list)
            
            if candles:
                last = list(candles)[-1]
                missing = [k for k, v in last.items() if v == 0.0 and k not in ("liq_long", "liq_short")]
                if missing:
                    print(f"[{self.tab_id}] [WARN] {symbol}: zero fields = {missing}")
                else:
                    print(f"[{self.tab_id}] [OK]   {symbol}: all fields populated (close={last['close']}, vol={last['volume']}, funding={last['funding']})")
                    
            if symbol == "BTCUSDT":
                try:
                    with open(os.path.join(base_dir, "Seeding", "seeding_debug_BTCUSDT.json"), "w", encoding="utf-8") as f:
                        json.dump(debug_dicts, f, indent=2)
                    await self.page.screenshot(path=os.path.join(base_dir, "Seeding", f"diag_{self.tab_id}_{symbol}.png"), clip={"x": 0, "y": 0, "width": 600, "height": 400})
                except Exception:
                    pass
            print(f"[{self.tab_id}] [Success] Seeded {symbol} with {len(candles)} candles.")

def fetch_binance_funding_rates(symbol: str) -> List[Dict[str, Any]]:
    import urllib.request
    import json
    url = f"https://fapi.binance.com/fapi/v1/fundingRate?symbol={symbol}&limit=100"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode())
    except Exception as e:
        print(f"[Binance API] Failed to fetch funding rate for {symbol}: {e}")
        return []

def fetch_binance_open_interest(symbol: str) -> List[Dict[str, Any]]:
    import urllib.request
    import json
    url = f"https://fapi.binance.com/fapi/v1/openInterestHist?symbol={symbol}&period=15m&limit=120"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode())
    except Exception as e:
        print(f"[Binance API] Failed to fetch open interest for {symbol}: {e}")
        return []

def fetch_binance_ls_ratio(symbol: str) -> List[Dict[str, Any]]:
    import urllib.request
    import json
    url = f"https://fapi.binance.com/data/globalLongShortAccountRatio?symbol={symbol}&period=15m&limit=120"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode())
    except Exception as e:
        print(f"[Binance API] Failed to fetch long/short ratio for {symbol}: {e}")
        return []

def _dump_xlsx(symbol: str, rows: List[Dict[str, Any]]) -> None:
    crypto_symbols = {
        "BTCUSDT", "ETHUSDT", "XRPUSDT", "SOLUSDT", "BNBUSDT", 
        "DOGEUSDT", "ADAUSDT", "TRXUSDT", "LINKUSDT", "AVAXUSDT", 
        "SUIUSDT", "NEARUSDT", "DOTUSDT", "LTCUSDT"
    }
    
    if rows and symbol in crypto_symbols:
        # 1. Backfill funding rate if all zeros
        if all(r.get("funding", 0.0) == 0.0 for r in rows):
            api_rates = fetch_binance_funding_rates(symbol)
            if api_rates:
                api_rates.sort(key=lambda x: x["fundingTime"])
                for r in rows:
                    row_time_ms = r.get("open_time", 0) * 1000
                    matching_rate = 0.0
                    for item in api_rates:
                        if item["fundingTime"] <= row_time_ms:
                            matching_rate = float(item["fundingRate"])
                        else:
                            break
                    r["funding"] = matching_rate

        # 2. Backfill open interest if all zeros
        if all(r.get("oi", 0.0) == 0.0 for r in rows):
            api_oi = fetch_binance_open_interest(symbol)
            if api_oi:
                api_oi.sort(key=lambda x: x["timestamp"])
                for r in rows:
                    row_time_ms = r.get("open_time", 0) * 1000
                    matching_oi = 0.0
                    for item in api_oi:
                        if item["timestamp"] <= row_time_ms:
                            matching_oi = float(item["sumOpenInterest"])
                        else:
                            break
                    r["oi"] = matching_oi

        # 3. Backfill long/short ratio if all zeros/default
        if all(r.get("ls_ratio", 1.0) == 1.0 or r.get("ls_ratio", 1.0) == 0.0 for r in rows):
            api_ls = fetch_binance_ls_ratio(symbol)
            if api_ls:
                api_ls.sort(key=lambda x: x["timestamp"])
                for r in rows:
                    row_time_ms = r.get("open_time", 0) * 1000
                    matching_ls = 1.0
                    for item in api_ls:
                        if item["timestamp"] <= row_time_ms:
                            matching_ls = float(item["longShortRatio"])
                        else:
                            break
                    r["ls_ratio"] = matching_ls

    # 4. Apply general forward-fill and backward-fill for all numeric columns to handle scattered zeros
    if rows:
        fill_fields = [
            "open", "high", "low", "close", "volume", "rsi", "fut_cvd", "spot_cvd", "funding", "ls_ratio", "oi",
            "coins_bid", "coins_ask", "dollars_bid", "dollars_ask", "whale_idx", "tk_buy_cnt", "tk_sell_cnt"
        ]
        for field in fill_fields:
            non_zero_vals = [r.get(field, 0.0) for r in rows if r.get(field, 0.0) != 0.0]
            if non_zero_vals:
                # Forward fill
                last_val = non_zero_vals[0]
                for r in rows:
                    val = r.get(field, 0.0)
                    if val != 0.0:
                        last_val = val
                    else:
                        r[field] = last_val
                # Backward fill
                last_val = non_zero_vals[-1]
                for r in reversed(rows):
                    val = r.get(field, 0.0)
                    if val != 0.0:
                        last_val = val
                    else:
                        r[field] = last_val

    wb = Workbook()
    ws = wb.active
    ws.title = symbol[:31]
    
    headers = [
        "open_time", "open", "high", "low", "close", "volume", 
        "rsi", "fut_cvd", "spot_cvd", "funding", "liq_long", "liq_short", "ls_ratio", "oi",
        "coins_bid", "coins_ask", "dollars_bid", "dollars_ask", "whale_idx", "tk_buy_cnt", "tk_sell_cnt"
    ]
    
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    CENTER = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        
    for row_idx, r in enumerate(rows, start=2):
        row_vals = []
        for h in headers:
            val = r.get(h, "")
            if h == "open_time" and isinstance(val, (int, float)):
                from datetime import datetime, timezone, timedelta
                tz_ist = timezone(timedelta(hours=5, minutes=30))
                val = datetime.fromtimestamp(val, tz=tz_ist).strftime("%Y-%m-%d %H:%M:%S IST")
            row_vals.append(val)
        ws.append(row_vals)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CENTER
            cell.border = BORDER
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    os.makedirs(os.path.join(base_dir, "Seeding"), exist_ok=True)
    filename = os.path.join(base_dir, "Seeding", f"{symbol}_seed_history.xlsx")
    try:
        wb.save(filename)
    except PermissionError:
        import random
        alt_filename = os.path.join(base_dir, "Seeding", f"{symbol}_seed_history_{random.randint(1000, 9999)}.xlsx")
        print(f"[WARN] Permission denied on {filename} (probably open in Excel). Saving to {alt_filename} instead.")
        try:
            wb.save(alt_filename)
        except Exception as e:
            print(f"[ERROR] Failed to save fallback Excel for {symbol}: {e}")

# --- PIPELINE STATUS HEADER ---
def render_pipeline_status(store: 'SnapshotStore') -> Any:
    """Compact 6-panel header table showing live health of every pipeline component."""
    from rich.table import Table as _T
    ph = store.pipeline_health if hasattr(store, 'pipeline_health') else {}
    tt = store.trade_tracker

    # ── Color helpers ──
    def _ok(v):  return f"[bold green]{v}[/bold green]"
    def _warn(v): return f"[bold yellow]{v}[/bold yellow]"
    def _err(v): return f"[bold red]{v}[/bold red]"
    def _dim(v): return f"[dim]{v}[/dim]"
    def _status_color(status, ok_vals=("CONNECTED", "TESTNET", "LIVE", "NORMAL", "LOADED")):
        if status in ok_vals:
            return _ok(status)
        if status in ("INIT", "WARMING", "COOLDOWN"):
            return _warn(status)
        return _err(status)

    # ── Panel 1: Chrome CDP ──
    chrome_s = ph.get("chrome_status", "INIT")
    chrome_lat = ph.get("chrome_latency_ms", 0)
    chrome_polls = ph.get("chrome_polls", 0)
    chrome_fails = ph.get("chrome_poll_failures", 0)
    fail_str = _err(f"Fails: {chrome_fails}") if chrome_fails > 10 else (_warn(f"Fails: {chrome_fails}") if chrome_fails > 0 else _dim("Fails: 0"))
    p1 = (f"Status: {_status_color(chrome_s)}\n"
          f"Latency: {chrome_lat:.0f}ms | {fail_str}\n"
          f"Polls: {chrome_polls:,}")

    # ── Panel 2: Binance Broker ──
    broker_s = ph.get("binance_broker_status", "INIT")
    broker_bal = ph.get("binance_broker_balance", 0)
    broker_pos = ph.get("binance_broker_positions", 0)
    p2 = (f"Status: {_status_color(broker_s, ('TESTNET', 'LIVE'))}\n"
          f"Balance: [cyan]${broker_bal:,.2f}[/cyan]\n"
          f"Positions: {broker_pos}")

    # ── Panel 3: Scraper Stream ──
    fps = ph.get("scraper_fps", 0)
    last_parse = ph.get("scraper_last_parse_ns", 0)
    parse_age = (time.time_ns() - last_parse) / 1e9 if last_parse else 999
    fps_str = _ok(f"{fps:.1f}") if fps > 0.5 else (_warn(f"{fps:.1f}") if fps > 0 else _err("0.0"))
    age_str = _ok(f"{parse_age:.0f}s") if parse_age < 30 else (_warn(f"{parse_age:.0f}s") if parse_age < 120 else _err(f"{parse_age:.0f}s"))
    ws_ticks = ph.get("binance_ws_ticks", 0)
    ws_s = ph.get("binance_ws_status", "INIT")
    frame_ok = ph.get("scraper_frame_ok", "?/?")
    p3 = (f"FPS: {fps_str} | Age: {age_str}\n"
          f"WS: {_status_color(ws_s)} | Ticks: {ws_ticks:,}\n"
          f"Frames: {frame_ok}")

    # ── Panel 4: Rolling Window Buffer ──
    pred = store.predictor
    if pred and hasattr(pred, 'candles_history'):
        buf_counts = []
        try:
            for sym in ALL_SYMBOLS[:14]:  # crypto symbols only
                history_list = list(pred.candles_history.get(sym, []))
                buf_counts.append(len(history_list))
            avg_buf = int(sum(buf_counts) / max(len(buf_counts), 1))
            min_buf = min(buf_counts) if buf_counts else 0
            warm_pct = min(100, int(avg_buf / 250 * 100))
            buf_color = _ok if warm_pct >= 100 else (_warn if warm_pct >= 50 else _err)
            p4 = (f"Avg: {buf_color(f'{avg_buf}/250')} ({warm_pct}%)\n"
                  f"Min: {min_buf}/250")
        except Exception:
            p4 = _dim("Reading buffer...")
    else:
        p4 = _dim("No predictor")

    # ── Panel 5: ML Predictor ──
    if pred and hasattr(pred, 'models'):
        try:
            models_copy = {k: list(v.keys()) for k, v in list(pred.models.items())}
            total_models = sum(len(v) for v in models_copy.values())
            n_strats = sum(1 for v in models_copy.values() if v)
            ml_s = _ok(f"LOADED ({total_models})") if total_models >= 84 else (_warn(f"PARTIAL ({total_models})") if total_models > 0 else _err("UNLOADED"))
            p5 = (f"Models: {ml_s}\n"
                  f"Strategies: {n_strats}/6 active")
        except Exception:
            p5 = _dim("Loading models...")
    else:
        p5 = _dim("No predictor")

    # ── Panel 6: Risk Governor ──
    if tt:
        halt = getattr(tt, 'emergency_halt', False)
        capital = getattr(tt, 'current_capital', 0)
        daily_start = getattr(tt, 'daily_start_capital', capital)
        initial = getattr(tt, 'initial_capital', capital)
        equity = capital  # simplified — no unrealized PnL in header
        daily_dd = (daily_start - equity) / daily_start * 100 if daily_start > 0 else 0
        total_dd = (initial - equity) / initial * 100 if initial > 0 else 0

        if halt:
            gov_s = _err("EMERGENCY_HALT")
        elif daily_dd > 5.0 or total_dd > 8.0:
            gov_s = _warn("CAUTION")
        else:
            gov_s = _ok("NORMAL")

        # Check cooldowns
        now_t = time.time()
        cooldowns = getattr(tt, 'reentry_cooldown_until', {})
        active_cooldowns = sum(1 for v in cooldowns.values() if v > now_t)

        p6 = (f"Status: {gov_s}\n"
              f"Daily DD: {daily_dd:.1f}% | Total: {total_dd:.1f}%\n"
              f"Cooldowns: {active_cooldowns}")
    else:
        p6 = _dim("No tracker")

    # ── Build table ──
    tbl = _T(
        title="[bold bright_cyan]⚡ Pipeline Status[/bold bright_cyan]",
        header_style="bold bright_cyan",
        border_style="bright_blue",
        expand=True,
        show_lines=True,
    )
    tbl.add_column("Chrome CDP", justify="center", ratio=1)
    tbl.add_column("Binance Broker", justify="center", ratio=1)
    tbl.add_column("Data Stream", justify="center", ratio=1)
    tbl.add_column("Buffer", justify="center", ratio=1)
    tbl.add_column("ML Predictor", justify="center", ratio=1)
    tbl.add_column("Risk Governor", justify="center", ratio=1)
    tbl.add_row(p1, p2, p3, p4, p5, p6)
    return tbl


def render_table(snap: Dict[str, AssetSnapshot], trade_tracker: Any = None, store: Any = None) -> Any:
    from rich.columns import Columns
    from rich.panel import Panel
    from rich.text import Text
    from rich.console import Group

    # Table 1: CoinGlass Real-Time Market & Liquidity Data (Top Full-Width)
    t1 = Table(
        title="[bold bright_cyan]📊 Table 1: CoinGlass Ingested Real-Time Market, Liquidity & Order Flow Data[/bold bright_cyan]",
        header_style="bold bright_cyan",
        border_style="bright_blue",
        expand=True,
        pad_edge=False,
        padding=(0, 0)
    )
    cols_t1 = (
        "Sym", "Price", "Vol", "RSI", "Fut CVD", "Spot CVD", "Funding",
        "OI", "Liq L", "Liq S", "L/S", "Bid ($)", "Ask ($)",
        "Bid (C)", "Ask (C)", "Whale", "Tk Buy", "Tk Sell", "Signal"
    )

    # Table 2: EMAs, Volatility ATRs & Multi-Factor Statistical Z-Scores (Bottom-Left)
    t2 = Table(
        title="[bold bright_magenta]📈 Table 2: EMAs, Volatility ATRs & Multi-Factor Statistical Z-Scores[/bold bright_magenta]",
        header_style="bold bright_magenta",
        border_style="magenta",
        expand=True,
        pad_edge=False,
        padding=(0, 0)
    )
    cols_t2 = (
        "Sym", "EMA 8", "EMA 21", "EMA 50", "EMA 200", "EMA 800",
        "ATR 14", "ATR 100", "Z-Price", "Z-CVD", "Z-OI", "Z-Fund"
    )

    # Column-to-snapshot field mapping for staleness tracking
    _COL_FIELD_MAP = {
        "Price": "price", "Vol": "volume", "RSI": "rsi", "Fut CVD": "fut_cvd", "Spot CVD": "spot_cvd",
        "Funding": "funding", "OI": "oi", "Liq L": "liq_long", "Liq S": "liq_short",
        "L/S": "ls_ratio", "Bid ($)": "dollars_bid", "Ask ($)": "dollars_ask",
        "Bid (C)": "coins_bid", "Ask (C)": "coins_ask", "Whale": "whale_idx",
        "Tk Buy": "tk_buy_cnt", "Tk Sell": "tk_sell_cnt",
        "EMA 8": "ema_8", "EMA 21": "ema_21", "EMA 50": "ema_50", "EMA 200": "ema_200", "EMA 800": "ema_800",
        "ATR 14": "atr_14", "ATR 100": "atr_100",
    }
    # Update column staleness tracking using BTCUSDT as representative
    _now_wall = time.time()
    _btc = snap.get("BTCUSDT")
    if _btc:
        for col_name, field_name in _COL_FIELD_MAP.items():
            cur_val = getattr(_btc, field_name, None)
            prev_val = _COLUMN_LAST_VALUES.get(col_name)
            if cur_val is not None and prev_val is not None and abs(float(cur_val) - float(prev_val)) > 1e-9:
                _COLUMN_LAST_CHANGED_TIME[col_name] = _now_wall
            elif col_name not in _COLUMN_LAST_CHANGED_TIME:
                _COLUMN_LAST_CHANGED_TIME[col_name] = _now_wall
            if cur_val is not None:
                _COLUMN_LAST_VALUES[col_name] = cur_val

    for col in cols_t1:
        f_name = _COL_FIELD_MAP.get(col, col.lower())
        contract = INDICATOR_FRESHNESS_CONTRACTS.get(f_name, {"interval": 60.0, "tolerance": 2.5})
        threshold = contract["interval"] * contract["tolerance"]
        secs_since_change = _now_wall - _COLUMN_LAST_CHANGED_TIME.get(col, _now_wall)
        header = f"[bold purple]{col}[/bold purple]" if (secs_since_change >= threshold and col != "Sym") else col
        t1.add_column(header, justify="center", no_wrap=True)

    for col in cols_t2:
        f_name = _COL_FIELD_MAP.get(col, col.lower())
        contract = INDICATOR_FRESHNESS_CONTRACTS.get(f_name, {"interval": 60.0, "tolerance": 2.5})
        threshold = contract["interval"] * contract["tolerance"]
        secs_since_change = _now_wall - _COLUMN_LAST_CHANGED_TIME.get(col, _now_wall)
        header = f"[bold purple]{col}[/bold purple]" if (secs_since_change >= threshold and col != "Sym") else col
        t2.add_column(header, justify="center", no_wrap=True)
        
    pred = getattr(store, "predictor", None) if store else None
    history_map = getattr(pred, "candles_history", {}) if pred else {}
    now = time.time_ns()
    
    def compact_num(n: float, prefix: str = "", signed: bool = False) -> str:
        if n == 0 or n is None:
            return "--"
        sign = "+" if (signed and n > 0) else ("-" if n < 0 else "")
        abs_n = abs(n)
        if abs_n >= 1e9:
            val_str = f"{abs_n / 1e9:.2f}B"
        elif abs_n >= 1e6:
            val_str = f"{abs_n / 1e6:.1f}M"
        elif abs_n >= 1e3:
            val_str = f"{abs_n / 1e3:.1f}K"
        elif abs_n < 0.001:
            val_str = f"{abs_n:.5f}"
        elif abs_n < 1:
            val_str = f"{abs_n:.4f}"
        else:
            val_str = f"{abs_n:.2f}"
        return f"{sign}{prefix}{val_str}"

    def fmt_z(z: float, fresh: bool = True) -> str:
        if z >= 2.0:
            return f"[bold red]{z:+.1f}σ[/bold red]"
        elif z <= -2.0:
            return f"[bold green]{z:+.1f}σ[/bold green]"
        elif abs(z) >= 1.0:
            return f"[yellow]{z:+.1f}σ[/yellow]"
        return f"[cyan]{z:+.1f}σ[/cyan]"

    def fmt_val(v: Any, fresh: bool = True, col_type: str = "generic") -> str:
        if v is None:
            return "[dim]--[/dim]"
        
        if col_type == "arm":
            s_val = str(v)
            if "SHORT" in s_val:
                strats = re.findall(r"S\d", s_val)
                short_tag = f"{'/'.join(strats)}:S" if strats else "SHORT"
                return f"[bold bright_red]{short_tag[:14]}[/bold bright_red]"
            elif "LONG" in s_val:
                strats = re.findall(r"S\d", s_val)
                long_tag = f"{'/'.join(strats)}:L" if strats else "LONG"
                return f"[bold bright_green]{long_tag[:14]}[/bold bright_green]"
            elif "BULL" in s_val:
                return "[bold bright_green]BULL[/bold bright_green]"
            elif "BEAR" in s_val:
                return "[bold bright_red]BEAR[/bold bright_red]"
            elif "OVERBOUGHT" in s_val:
                return "[bold red]OB[/bold red]"
            elif "OVERSOLD" in s_val:
                return "[bold green]OS[/bold green]"
            elif "READY" in s_val:
                return "[dim green]READY[/dim green]"
            return f"[cyan]{s_val[:12]}[/cyan]"

        if isinstance(v, str):
            return f"[white]{v}[/white]"

        if col_type == "rsi":
            s = f"{v:.1f}"
            if v >= 70:
                return f"[bold red]{s}[/bold red]"
            elif v <= 30:
                return f"[bold green]{s}[/bold green]"
            return f"[bold cyan]{s}[/bold cyan]"
        elif col_type == "fund":
            s = f"{v:+.5f}"
            if v > 0:
                return f"[bold green]{s}[/bold green]"
            elif v < 0:
                return f"[bold yellow]{s}[/bold yellow]"
            return f"[cyan]{s}[/cyan]"
        elif col_type in ("cvd", "fp_d"):
            s = compact_num(v, signed=True)
            if v > 0:
                return f"[bold green]{s}[/bold green]"
            elif v < 0:
                return f"[bold red]{s}[/bold red]"
            return f"[white]{s}[/white]"
        elif col_type == "vol":
            s = compact_num(v)
            return f"[bold bright_white]{s}[/bold bright_white]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "oi":
            s = compact_num(v)
            return f"[bold cyan]{s}[/bold cyan]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "dollars_bid":
            s = compact_num(v, prefix="$")
            return f"[bold green]{s}[/bold green]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "dollars_ask":
            s = compact_num(abs(v), prefix="-$")
            return f"[bold red]{s}[/bold red]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "coins_bid":
            s = compact_num(v)
            return f"[bold green]{s}[/bold green]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "coins_ask":
            s = compact_num(abs(v), prefix="-")
            return f"[bold red]{s}[/bold red]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "tk_buy":
            s = compact_num(v)
            return f"[bold green]{s}[/bold green]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "tk_sell":
            s = compact_num(abs(v), prefix="-")
            return f"[bold red]{s}[/bold red]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "whale":
            s = f"{v:+.1f}" if v != 0 else "--"
            if abs(v) > 50:
                return f"[bold yellow]{s}[/bold yellow]"
            return f"[bold bright_white]{s}[/bold bright_white]" if v != 0 else "[dim]--[/dim]"
        elif col_type == "price":
            s = f"{v:,.2f}" if v >= 1.0 else f"{v:,.4f}"
            return f"[bold yellow]{s}[/bold yellow]" if v > 0 else "[dim]--[/dim]"
        elif col_type == "atr":
            s = f"{v:,.2f}" if v >= 1.0 else f"{v:,.4f}"
            return f"[bold cyan]{s}[/bold cyan]" if v > 0 else "[dim]--[/dim]"
        elif col_type == "liq_long":
            s = compact_num(v)
            return f"[bold bright_green]{s}[/bold bright_green]" if v != 0 else f"[dim]--[/dim]"
        elif col_type == "liq_short":
            s = compact_num(abs(v), prefix="-")
            return f"[bold bright_red]{s}[/bold bright_red]" if v != 0 else f"[dim]--[/dim]"
        elif col_type == "lsr":
            s = f"{v:.2f}"
            return f"[bold cyan]{s}[/bold cyan]"
        else:
            s = compact_num(v)
            return f"[white]{s}[/white]"

    for sym in ALL_SYMBOLS:
        a = snap.get(sym, AssetSnapshot(symbol=sym))
        fresh = (now - a.ts_ns) < STALE_NS
        hist = list(history_map.get(sym, []))
        latest_c = hist[-1] if hist else {}

        # Fallback values from candle history if snap is 0 or unpopulated
        price = a.price if a.price > 0 else float(latest_c.get("close", 0.0))
        vol = a.volume if a.volume > 0 else float(latest_c.get("volume", 0.0))
        rsi = a.rsi if a.rsi > 0 else float(latest_c.get("rsi", 0.0))
        fut_cvd = a.fut_cvd if a.fut_cvd != 0.0 else float(latest_c.get("fut_cvd", 0.0))
        spot_cvd = a.spot_cvd if a.spot_cvd != 0.0 else float(latest_c.get("spot_cvd", 0.0))
        fund = a.funding if a.funding != 0.0 else float(latest_c.get("funding", 0.0))
        oi = a.oi if a.oi > 0 else float(latest_c.get("oi", 0.0))
        liq_long = a.liq_long if a.liq_long != 0.0 else float(latest_c.get("liq_long", 0.0))
        liq_short = a.liq_short if a.liq_short != 0.0 else float(latest_c.get("liq_short", 0.0))
        ls_ratio = a.ls_ratio if a.ls_ratio != 0.0 else float(latest_c.get("ls_ratio", 0.0))

        z_price_val = 0.0
        z_cvd_val = 0.0
        z_oi_val = 0.0
        z_fund_val = 0.0
        
        if len(hist) >= 5:
            import numpy as np
            closes = [float(c.get("close", 0.0)) for c in hist if float(c.get("close", 0.0)) > 0]
            cvds = [float(c.get("fut_cvd", 0.0)) for c in hist]
            ois = [float(c.get("oi", 0.0)) for c in hist if float(c.get("oi", 0.0)) > 0]
            funds = [float(c.get("funding", 0.0)) for c in hist]
            
            if closes and price > 0 and np.std(closes) > 1e-9:
                z_price_val = (price - np.mean(closes)) / np.std(closes)
            if cvds and np.std(cvds) > 1e-9:
                z_cvd_val = (fut_cvd - np.mean(cvds)) / np.std(cvds)
            if ois and oi > 0 and np.std(ois) > 1e-9:
                z_oi_val = (oi - np.mean(ois)) / np.std(ois)
            if funds and np.std(funds) > 1e-9:
                z_fund_val = (fund - np.mean(funds)) / np.std(funds)

        # Compute EMAs from candle history
        ema_8_val = a.ema_8
        ema_21_val = a.ema_21
        ema_50_val = a.ema_50
        ema_200_val = a.ema_200
        ema_800_val = a.ema_800

        if len(hist) >= 5:
            import pandas as pd
            all_closes = [float(c.get("close", 0.0)) for c in hist if float(c.get("close", 0.0)) > 0]
            if price > 0:
                all_closes.append(price)
            if all_closes:
                s = pd.Series(all_closes)
                if ema_8_val == 0.0:
                    ema_8_val = float(s.ewm(span=min(len(s), 8), adjust=False).mean().iloc[-1])
                if ema_21_val == 0.0:
                    ema_21_val = float(s.ewm(span=min(len(s), 21), adjust=False).mean().iloc[-1])
                if ema_50_val == 0.0:
                    ema_50_val = float(s.ewm(span=min(len(s), 50), adjust=False).mean().iloc[-1])
                if ema_200_val == 0.0:
                    ema_200_val = float(s.ewm(span=min(len(s), 200), adjust=False).mean().iloc[-1])
                if ema_800_val == 0.0:
                    ema_800_val = float(s.ewm(span=min(len(s), 800), adjust=False).mean().iloc[-1])

        # ATR 14 and ATR 100
        atr_14_val = a.atr_14 if a.atr_14 > 0 else (a.atr if a.atr > 0 else 0.0)
        atr_100_val = a.atr_100 if a.atr_100 > 0 else 0.0
        if (atr_14_val == 0.0 or atr_100_val == 0.0) and len(hist) >= 15:
            import pandas as pd
            df_h = pd.DataFrame(hist)
            if "high" in df_h.columns and "low" in df_h.columns and "close" in df_h.columns:
                tr = pd.concat([
                    df_h["high"] - df_h["low"],
                    (df_h["high"] - df_h["close"].shift()).abs(),
                    (df_h["low"] - df_h["close"].shift()).abs()
                ], axis=1).max(axis=1)
                if atr_14_val == 0.0:
                    atr_14_val = float(tr.rolling(min(len(tr), 14)).mean().iloc[-1])
                if atr_100_val == 0.0:
                    atr_100_val = float(tr.rolling(min(len(tr), 100)).mean().iloc[-1])

        # Market trend / regime / arm classification
        arm_status = a.strategy_armed if a.strategy_armed else ""
        if not arm_status:
            if ema_8_val > 0 and ema_21_val > 0 and ema_50_val > 0:
                if ema_8_val > ema_21_val > ema_50_val:
                    arm_status = "BULL"
                elif ema_8_val < ema_21_val < ema_50_val:
                    arm_status = "BEAR"
            if abs(z_price_val) >= 2.0:
                arm_status = f"OB(+{z_price_val:.1f}σ)" if z_price_val > 0 else f"OS({z_price_val:.1f}σ)"
            if not arm_status:
                arm_status = "READY"

        # Table 1: CoinGlass Real-Time Market & Liquidity Data Row
        t1.add_row(
            f"[bold bright_white]{sym}[/bold bright_white]",
            fmt_val(price, fresh, "price"),
            fmt_val(vol, fresh, "vol"),
            fmt_val(rsi, fresh, "rsi"),
            fmt_val(fut_cvd, fresh, "cvd"),
            fmt_val(spot_cvd, fresh, "cvd"),
            fmt_val(fund, fresh, "fund"),
            fmt_val(oi, fresh, "oi"),
            fmt_val(liq_long, fresh, "liq_long"),
            fmt_val(liq_short, fresh, "liq_short"),
            fmt_val(ls_ratio, fresh, "lsr"),
            fmt_val(a.dollars_bid, fresh, "dollars_bid"),
            fmt_val(a.dollars_ask, fresh, "dollars_ask"),
            fmt_val(a.coins_bid, fresh, "coins_bid"),
            fmt_val(a.coins_ask, fresh, "coins_ask"),
            fmt_val(a.whale_idx, fresh, "whale"),
            fmt_val(a.tk_buy_cnt, fresh, "tk_buy"),
            fmt_val(a.tk_sell_cnt, fresh, "tk_sell"),
            fmt_val(arm_status, fresh, "arm"),
        )

        # Table 2: EMAs, Volatility ATRs & Multi-Factor Statistical Z-Scores Row
        t2.add_row(
            f"[bold bright_white]{sym}[/bold bright_white]",
            fmt_val(ema_8_val, fresh, "price"),
            fmt_val(ema_21_val, fresh, "price"),
            fmt_val(ema_50_val, fresh, "price"),
            fmt_val(ema_200_val, fresh, "price"),
            fmt_val(ema_800_val, fresh, "price"),
            fmt_val(atr_14_val, fresh, "atr"),
            fmt_val(atr_100_val, fresh, "atr"),
            fmt_z(z_price_val, fresh),
            fmt_z(z_cvd_val, fresh),
            fmt_z(z_oi_val, fresh),
            fmt_z(z_fund_val, fresh),
        )

    # Table 3: Active Trades, Trade Logs & Performance Panel (Bottom-Right)
    active_lines = []
    history_lines = []
    stats_text = "[dim]Waiting for broker connection...[/dim]"
    
    if trade_tracker:
        stats = trade_tracker.get_stats()
        with trade_tracker.lock:
            active_snap = list(trade_tracker.active_trades.values())
            history_snap = list(trade_tracker.history[-5:])

        for tr in active_snap:
            dir_str = "[bold bright_green]LONG[/bold bright_green]" if tr['direction'] == 1 else "[bold bright_red]SHORT[/bold bright_red]"
            pnl_usd = tr.get('live_pnl_usd', 0.0)
            pnl_pct = tr.get('live_pnl_pct', 0.0)
            pnl_str = f"[bold green]+${pnl_usd:.2f} (+{pnl_pct:+.2f}%)[/bold green]" if pnl_usd >= 0 else f"[bold red]-${abs(pnl_usd):.2f} ({pnl_pct:+.2f}%)[/bold red]"
            broker_info = f" | Lot: {tr['exec_lot']:.2f}" if 'exec_lot' in tr else ""
            active_lines.append(f"[bold bright_white]{tr['symbol']}[/] | {dir_str} | Entry: [yellow]{tr['entry_price']:.4f}[/] | SL: [red]{tr['sl']:.4f}[/] | TP: [green]{tr['tp']:.4f}[/] | Live: {pnl_str}{broker_info}")

        for tr in history_snap:
            dir_str = "[bold green]LONG[/]" if tr['direction'] == 1 else "[bold red]SHORT[/]"
            pnl_usd = tr.get('pnl_usd', 0.0)
            pnl_pct = tr.get('pnl_pct', 0.0)
            pnl_str = f"[bold green]+${pnl_usd:.2f} (+{pnl_pct:+.2f}%)[/]" if pnl_usd >= 0 else f"[bold red]-${abs(pnl_usd):.2f} ({pnl_pct:+.2f}%)[/]"
            reason = tr.get('exit_reason', 'EXIT')
            history_lines.append(f"{tr['symbol']} | {dir_str} | Exit: {tr['exit_price']:.4f} | Reason: {reason} | PnL: {pnl_str}")

        winrate = stats['winrate']
        total_pnl = stats['total_pnl_usd']
        pnl_pct = total_pnl / trade_tracker.initial_capital * 100.0 if trade_tracker.initial_capital > 0 else 0.0
        pnl_clr = "bright_green" if total_pnl >= 0 else "bright_red"
        pnl_sign = "+" if total_pnl >= 0 else ""

        stats_text = (
            f"Capital: [bold bright_cyan]${stats['current_capital']:,.2f}[/]  |  "
            f"PnL: [bold {pnl_clr}]{pnl_sign}${total_pnl:.2f} ({pnl_pct:+.2f}%)[/]  |  "
            f"Trades: [bold bright_yellow]{stats['total']}[/]  |  Winrate: [bold bright_yellow]{winrate:.1f}%[/]"
        )

    t3 = Table(
        title="[bold bright_green]💼 Table 3: Active Trades & Trade Logs[/bold bright_green]",
        header_style="bold bright_green",
        border_style="bright_green",
        expand=True,
        pad_edge=False,
        padding=(0, 0)
    )
    t3.add_column("Live Positions & History", justify="left", ratio=1)
    
    active_content = "\n".join(active_lines) if active_lines else "[dim]No active positions[/dim]"
    history_content = "\n".join(history_lines) if history_lines else "[dim]No recent trades[/dim]"
    
    t3_body = f"[bold underline cyan]Active Positions:[/bold underline cyan]\n{active_content}\n\n[bold underline yellow]Recent Closed History:[/bold underline yellow]\n{history_content}\n\n[bold bright_white]{stats_text}[/bold bright_white]"
    t3.add_row(t3_body)

    # Place Table 2 and Table 3 side-by-side in the same horizontal row
    bottom_grid = Table.grid(expand=True, padding=(0, 1))
    bottom_grid.add_column("left", ratio=50)
    bottom_grid.add_column("right", ratio=50)
    bottom_grid.add_row(t2, t3)

    # Live Event Log Panel
    log_lines = list(_LIVE_LOG_FEED)
    if not log_lines:
        log_lines = [f"[{datetime.now().strftime('%H:%M:%S')}] [SYS] Streaming pipeline active. All feeds online."]
    log_panel = Panel(Text.from_markup("\n".join(log_lines)), title="[bold bright_white]📜 Live System & Signal Event Log[/bold bright_white]", border_style="dim white")

    # Final Combined Layout
    if store and hasattr(store, 'pipeline_health'):
        pipeline_tbl = render_pipeline_status(store)
        return Group(pipeline_tbl, t1, bottom_grid, log_panel)
    return Group(t1, bottom_grid, log_panel)

async def renderer_loop(store: SnapshotStore, stop: asyncio.Event) -> None:
    if sys.platform == "win32":
        try:
            import ctypes
            from ctypes import wintypes
            kernel32 = ctypes.WinDLL('kernel32', use_last_error=True)
            STD_OUTPUT_HANDLE = wintypes.DWORD(-11)
            handle = kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
            mode = wintypes.DWORD()
            if kernel32.GetConsoleMode(handle, ctypes.byref(mode)):
                kernel32.SetConsoleMode(handle, mode.value | 0x0004)
        except Exception:
            pass

    console = Console(
        force_terminal=True,
        color_system="256",
        legacy_windows=False,
        no_color=False,
        highlight=False,
        soft_wrap=True,
    )
    loop_cnt = 0
    prev_prices: Dict[str, float] = {}
    _loop = asyncio.get_event_loop()
    try:
        init_table = await _loop.run_in_executor(RENDER_POOL, render_table, store.snapshot(), store.trade_tracker, store)
    except Exception:
        init_table = Table(title="[bold bright_cyan]Initializing Dashboard...[/bold bright_cyan]")

    import io
    with Live(
        init_table,
        console=console,
        auto_refresh=False,
        screen=False,
        transient=False,
        redirect_stdout=True,
        redirect_stderr=True,
        vertical_overflow="crop"
    ) as live:
        try:
            console.show_cursor(False)
        except Exception:
            pass
        while not stop.is_set():
            try:
                snap = store.snapshot()
                rendered = await _loop.run_in_executor(RENDER_POOL, render_table, snap, store.trade_tracker, store)
                console.print("\x1b[2J\x1b[H", end="")
                live.update(rendered, refresh=True)
            except Exception:
                pass
            
            loop_cnt += 1
            if loop_cnt % 4 == 0:  # Every 2.0 seconds at 2Hz REFRESH_HZ
                try:
                    serializable_snap = {}
                    for sym, a in list(snap.items()):
                        serializable_snap[sym] = {
                            "price": a.price, "volume": a.volume, "rsi": a.rsi, "fut_cvd": a.fut_cvd, "spot_cvd": a.spot_cvd,
                            "liq_long": a.liq_long, "liq_short": a.liq_short, "funding": a.funding,
                            "ls_ratio": a.ls_ratio, "oi": a.oi,
                            "coins_bid": a.coins_bid, "coins_ask": a.coins_ask,
                            "dollars_bid": a.dollars_bid, "dollars_ask": a.dollars_ask,
                            "whale_idx": a.whale_idx, "tk_buy_cnt": a.tk_buy_cnt, "tk_sell_cnt": a.tk_sell_cnt,
                            "fp_delta": a.fp_delta, "fp_poc": a.fp_poc,
                            "ema_8": a.ema_8, "ema_21": a.ema_21, "ema_50": a.ema_50,
                            "ema_200": a.ema_200, "ema_800": a.ema_800, "atr_100": a.atr_100,
                            "strategy_armed": a.strategy_armed, "ts_ns": a.ts_ns
                        }
                    def _write_debug(snap_copy):
                        try:
                            tmp_path = os.path.join(base_dir, "Seeding", "snapshot_debug.json.tmp")
                            with open(tmp_path, "w", encoding="utf-8") as f:
                                json.dump(serializable_snap, f, indent=4)
                            os.replace(tmp_path, os.path.join(base_dir, "Seeding", "snapshot_debug.json"))
                        except Exception:
                            pass
                        try:
                            # Render clean plain text table to disk file without touching live render tree
                            string_buf = io.StringIO()
                            export_console = Console(
                                file=string_buf,
                                force_terminal=False,
                                color_system=None,
                                width=220,
                            )
                            detached_tbl = render_table(snap_copy, store.trade_tracker, store)
                            export_console.print(detached_tbl)
                            txt = string_buf.getvalue()
                            live_tbl_path = os.path.join(base_dir, "live_data", "live_terminal_table.txt")
                            with open(live_tbl_path, "w", encoding="utf-8") as f:
                                f.write(txt)
                        except Exception:
                            pass
                    await asyncio.to_thread(_write_debug, snap)
                except Exception:
                    pass

            # Every 30 seconds (60 iterations @ 2Hz), audit value changes across ALL columns
            if loop_cnt % 60 == 0:
                try:
                    audit_fields = [
                        "price", "volume", "rsi", "fut_cvd", "spot_cvd",
                        "liq_long", "liq_short", "funding", "ls_ratio", "oi",
                        "fp_delta", "fp_poc", "whale_idx", "dollars_bid", "dollars_ask"
                    ]
                    if not hasattr(renderer_loop, "_prev_snapshot_cols"):
                        renderer_loop._prev_snapshot_cols = {}
                        renderer_loop._col_static_count = {f: 0 for f in audit_fields}

                    curr_cols = {}
                    for sym in ALL_SYMBOLS:
                        if sym in snap:
                            curr_cols[sym] = {f: getattr(snap[sym], f, None) for f in audit_fields}

                    col_changes = {f: 0 for f in audit_fields}
                    prev_cols = renderer_loop._prev_snapshot_cols

                    if prev_cols:
                        for sym in ALL_SYMBOLS:
                            c_curr = curr_cols.get(sym, {})
                            c_prev = prev_cols.get(sym, {})
                            for f in audit_fields:
                                v1 = c_prev.get(f)
                                v2 = c_curr.get(f)
                                if v1 is not None and v2 is not None and v1 != v2:
                                    col_changes[f] += 1

                        moving_cols = [f for f, cnt in col_changes.items() if cnt > 0]
                        stale_cols = [f for f, cnt in col_changes.items() if cnt == 0]

                        for f in audit_fields:
                            if col_changes[f] > 0:
                                renderer_loop._col_static_count[f] = 0
                            else:
                                renderer_loop._col_static_count[f] += 1

                        # Write live staleness report
                        report_data = {
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "interval_sec": 30.0,
                            "moving_columns_count": len(moving_cols),
                            "total_columns_count": len(audit_fields),
                            "moving_columns": moving_cols,
                            "stale_columns": stale_cols,
                            "column_changes_per_symbol": col_changes,
                            "status": "HEALTHY" if len(moving_cols) >= 3 else "STATIC_ALERT"
                        }
                        report_path = os.path.join(base_dir, "live_data", "column_staleness_report.json")
                        with open(report_path, "w", encoding="utf-8") as rf:
                            json.dump(report_data, rf, indent=2)

                        # Log brief audit note to terminal event log
                        active_syms_cnt = sum(1 for sym in ALL_SYMBOLS if any(curr_cols.get(sym, {}).get(f) != prev_cols.get(sym, {}).get(f) for f in audit_fields))
                        log_live_event(f"30s Watch: {len(moving_cols)}/{len(audit_fields)} cols moving across {active_syms_cnt}/18 symbols", "Surveillance")

                    renderer_loop._prev_snapshot_cols = curr_cols
                except Exception as audit_err:
                    pass

            await asyncio.sleep(1.0 / REFRESH_HZ)

# --- WATCHDOG ---
async def watchdog(components: List[Any], focus_lock: asyncio.Lock, stop: asyncio.Event) -> None:
    # Initialize/reset heartbeats for all components on startup to ignore the configuration time
    now_start = time.time_ns()
    for c in components:
        if hasattr(c, 'last_heartbeat_ns'):
            c.last_heartbeat_ns = now_start

    tab_tasks = {}
    try:
        while not stop.is_set():
            now = time.time_ns()
            for c in components:
                if hasattr(c, 'last_heartbeat_ns') and now - c.last_heartbeat_ns > 120_000_000_000:
                    if getattr(c, 'skip_watchdog', False) or getattr(c, 'is_seeding', False):
                        continue
                    log_live_event(f"Subsystem '{c.__class__.__name__}' ({getattr(c, 'tab_id', 'Unknown')}) hung >120s.", "WDog")
                    if isinstance(c, CoinglassTab):
                        log_live_event(f"Attempting soft reload recovery for '{c.tab_id}'...", "WDog")
                        try:
                            if c.page and not c.page.is_closed():
                                await c.page.reload(wait_until="load", timeout=30000)
                            else:
                                await c.reconnect(focus_lock)
                            c.last_heartbeat_ns = time.time_ns()
                            c.poll_failures = 0
                            # Reset heartbeats for all components to prevent false positives from the recovery latency
                            now_after = time.time_ns()
                            for comp in components:
                                if hasattr(comp, 'last_heartbeat_ns'):
                                    comp.last_heartbeat_ns = now_after
                        except Exception as rec_err:
                            log_live_event(f"Recovery failed for '{c.tab_id}': {rec_err}", "WDog")
            # Check Python process memory usage to catch memory leaks
            mem_mb = get_process_memory_usage() / (1024 * 1024)
            if mem_mb > 3584.0:  # 3.5 GB limit to allow initial retraining/seeding spikes
                log_live_event(f"[MEMORY] Python memory usage is extremely high ({mem_mb:.1f} MB)!", "WDog")
            await asyncio.sleep(5.0)
    finally:
        for task in tab_tasks.values():
            if not task.done():
                task.cancel()
        if tab_tasks:
            await asyncio.gather(*tab_tasks.values(), return_exceptions=True)

def combine_seeding_files() -> None:
    import glob
    import copy
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter

    files = glob.glob(os.path.join(base_dir, "Seeding", "*_seed_history.xlsx"))
    files = [f for f in files if "combined_seed" not in os.path.basename(f).lower()]
    if not files:
        print("[Setup] No seeding files found to combine.")
        return

    print(f"[Setup] Combining {len(files)} seeding files into a single workbook...")
    combined_wb = Workbook()
    default_sheet = combined_wb.active
    combined_wb.remove(default_sheet)

    for f in sorted(files):
        symbol = os.path.basename(f).replace("_seed_history.xlsx", "")
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            source_ws = wb.active
            target_ws = combined_wb.create_sheet(title=symbol[:31])

            for row in source_ws.iter_rows(values_only=True):
                target_ws.append(list(row))
            wb.close()
        except Exception as copy_exc:
            print(f"[Setup] [WARN] Failed to copy {symbol} sheet: {copy_exc}")

    combined_filename = os.path.join(base_dir, "Seeding", "combined_seed_history.xlsx")
    tmp_filename = combined_filename + ".tmp"
    try:
        combined_wb.save(tmp_filename)
        os.replace(tmp_filename, combined_filename)
        print(f"[Setup] Combined workbook saved successfully: {combined_filename}")
        
        # Clean up individual seed files
        for f in files:
            try:
                os.remove(f)
            except OSError:
                pass
        print("[Setup] Cleaned up individual symbol seeding files.")
    except Exception as e:
        print(f"[Setup] [WARN] Failed to save combined workbook: {e}")

def close_all_chrome_instances() -> None:
    """Forcefully terminates all active Chrome/Chromium/Driver instances and cleans profile locks."""
    print("[CleanUp] Terminating all active Chrome/Chromium and driver processes...")
    if sys.platform == "win32":
        try:
            import subprocess
            subprocess.run(["taskkill", "/F", "/IM", "chrome.exe", "/T"], capture_output=True)
            subprocess.run(["taskkill", "/F", "/IM", "chromedriver.exe", "/T"], capture_output=True)
            subprocess.run(["taskkill", "/F", "/IM", "msedge.exe", "/T"], capture_output=True)
            time.sleep(1.0)
        except Exception as ex:
            print(f"[CleanUp] Warning terminating Chrome: {ex}")
    else:
        try:
            import subprocess
            subprocess.run(["pkill", "-9", "-f", "chrome"], capture_output=True)
            time.sleep(1.0)
        except Exception:
            pass

    # Clear stale SingletonLock files from user data directories
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for d in ("chrome_profile_tab1", "chrome_profile_tab2", "chrome_profile_login"):
        lock_p = os.path.join(base_dir, d, "SingletonLock")
        if os.path.exists(lock_p):
            try:
                os.remove(lock_p)
                print(f"[CleanUp] Removed stale profile lock: {lock_p}")
            except Exception:
                pass

def run_retrain_proc():
    """Module-level target for the background retraining subprocess.
    Must be at module scope so Windows multiprocessing can import it."""
    import importlib
    base = os.path.dirname(os.path.abspath(__file__))
    if base not in sys.path:
        sys.path.insert(0, base)
    print("[Background Process] Starting Live Retraining for Six-Strategy models...")
    try:
        sys.modules.pop('train_six_strategy', None)
        train_six_mod = importlib.import_module("train_six_strategy")
        train_six_mod.train_all_strategies()
        print("[Background Process] [OK] Six-Strategy retraining completed")
    except Exception as e:
        import traceback
        print(f"[Background Process] Six-Strategy retrain failed: {e}")
        traceback.print_exc()
    print("[Background Process] Live Retraining finished.")


# --- MAIN CONTROLLER ---
async def main(skip_seed: bool = False, skip_train: bool = False) -> None:
    close_all_chrome_instances()
    base_dir = os.path.dirname(os.path.abspath(__file__))
    binance_live = os.environ.get("BINANCE_LIVE", os.environ.get("BINANCE_LIVE", "0")) == "1"
    print("=" * 60)
    print(f"  SYSTEM STARTUP - MODE: {EXECUTION_MODE} (BINANCE FUTURES)")
    if binance_live:
        print("  TRADES ARE DISPATCHED TO BINANCE FUTURES DEMO ACCOUNT / LOCAL TRACKER")
    else:
        print("  WARNING: NO REAL BINANCE FUTURES TRADE ORDERS WILL BE SENT")
        print("  TRADES ARE SIMULATED LOCALLY IN THE TRACKER FILE")
    print("=" * 60)

    if skip_train:
        print("[Setup] Skipping initial ML model retraining (--skip-train passed). Using existing models.")
    else:
        # 0. Clear existing ML models to prevent conflicts before retraining
        print("[Setup] Clearing existing ML model files before retraining...")
        for sub in (ACTIVE_STRATEGY, 'Liquidation', 'ml_trend_pull', 'six_strategy_models'):
            m_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), sub, 'models') if sub != 'six_strategy_models' else os.path.join(os.path.dirname(os.path.abspath(__file__)), sub)
            if os.path.exists(m_dir):
                for file in os.listdir(m_dir):
                    if file.endswith(('.pkl', '.json', '.txt', '.cbm', '.pt')):
                        try:
                            os.remove(os.path.join(m_dir, file))
                        except Exception as clear_err:
                            print(f"[Setup] [WARN] Could not remove old model file {file}: {clear_err}")

        # 0. Live Model Retraining on latest Parquet data
        print(f"[Setup] Running Live Model Retraining on latest Parquet data...")
        
        # Train unified six-strategy models (84 files: 6 strategies × 14 symbols)
        try:
            import importlib
            base_dir = os.path.dirname(os.path.abspath(__file__))
            if base_dir not in sys.path:
                sys.path.insert(0, base_dir)
            sys.modules.pop('train_six_strategy', None)
            train_six_mod = importlib.import_module("train_six_strategy")
            print("[Setup] Training Six-Strategy ML models (S1-S6 × 14 symbols)...")
            train_six_mod.train_all_strategies()
            print("[Setup] [OK] Six-Strategy models trained successfully")
        except Exception as retrain_err:
            print(f"[Setup] [WARN] Failed to retrain Six-Strategy models: {retrain_err}")
            import traceback
            traceback.print_exc()

    # Initialize unified Six-Strategy Predictor (ports run_all_6.py verified strategies)
    predictor = LiveSixStrategyPredictor(ALL_SYMBOLS)
    predictor.log_fn = log_live_event
    
    # Load cached history from disk
    predictor.load_history_from_disk()
    print(f"[Setup] Six-Strategy Predictor initialized with {len(predictor.models)} model sets")

    trade_tracker = Engine1TradeTracker()

    def background_retrain_loop():
        import time
        import subprocess
        from datetime import datetime, timezone, timedelta

        RETRAIN_HOUR_UTC = 0
        RETRAIN_MINUTE_UTC = 0

        while True:
            now_utc = datetime.now(timezone.utc)
            target = now_utc.replace(hour=RETRAIN_HOUR_UTC, minute=RETRAIN_MINUTE_UTC, second=0, microsecond=0)
            if target <= now_utc:
                target += timedelta(days=1)
            delay_secs = (target - now_utc).total_seconds()

            ist_offset = timedelta(hours=5, minutes=30)
            target_ist = target + ist_offset
            print(f"[Background Thread] Next retraining scheduled at {target.strftime('%Y-%m-%d %H:%M UTC')} ({target_ist.strftime('%H:%M IST')}) — in {delay_secs/3600:.1f} hours")
            time.sleep(delay_secs)

            print(f"[Background Thread] Launching scheduled Live Retraining Subprocess at {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}...")
            try:
                script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "train_six_strategy.py")
                result = subprocess.run(
                    [sys.executable, script],
                    capture_output=True,
                    text=True,
                    timeout=7200
                )
                if result.returncode == 0:
                    print("[Background Thread] [OK] Retraining subprocess finished successfully")
                else:
                    print(f"[Background Thread] Retraining subprocess exited with code {result.returncode}")
                    if result.stderr:
                        print(result.stderr[-2000:])
            except subprocess.TimeoutExpired:
                print("[Background Thread] Retraining subprocess timed out after 2h — killed")
            except Exception as ex:
                print(f"[Background Thread] Retraining subprocess crashed: {ex}")

    import threading
    retrain_thread = threading.Thread(target=background_retrain_loop, daemon=True)
    retrain_thread.start()
    print("[Setup] Launched 24hr Background Retraining Manager Thread (subprocess-isolated).")

    store = SnapshotStore(ALL_SYMBOLS, predictor, trade_tracker)

    # Initialize broker health status in pipeline
    if hasattr(trade_tracker, 'broker') and hasattr(trade_tracker.broker, 'broker'):
        raw_broker = trade_tracker.broker.broker
        is_testnet = getattr(raw_broker, 'use_testnet', False)
        is_dry = getattr(raw_broker, 'dry_run', True)
        store.pipeline_health["binance_broker_status"] = "TESTNET" if is_testnet else ("DRY_RUN" if is_dry else "LIVE")
        try:
            details = raw_broker.get_account_details()
            if details:
                store.pipeline_health["binance_broker_balance"] = details.get("balance", trade_tracker.initial_capital)
        except Exception:
            store.pipeline_health["binance_broker_balance"] = trade_tracker.initial_capital
    else:
        store.pipeline_health["binance_broker_status"] = "OFFLINE"
        store.pipeline_health["binance_broker_balance"] = trade_tracker.initial_capital
    stop = asyncio.Event()
    
    # Forcefully terminate orphan chrome instances before launching fresh contexts
    print("[Setup] Initializing CoinGlass Chrome integration (Port 19899)...")
    async with async_playwright() as pw:
        default_arena_dir = os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data_Arena")
        user_data_dir_1 = default_arena_dir if os.path.exists(default_arena_dir) else os.path.join(os.getcwd(), "chrome_profile_tab1")
        user_data_dir_2 = os.path.join(os.getcwd(), "chrome_profile_tab2")
        is_linux = sys.platform.startswith("linux")
        headless_flag = is_linux or os.environ.get("HEADLESS", "0") == "1"
        
        exec_path = os.environ.get("PLAYWRIGHT_CHROMIUM_EXECUTABLE_PATH")
        if not exec_path:
            if is_linux:
                import shutil
                exec_path = shutil.which("chromium-browser") or shutil.which("chromium")
            elif sys.platform == "win32":
                for p in (
                    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                    os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe")
                ):
                    if os.path.exists(p):
                        exec_path = p
                        break

        # Port configuration: Isolated Dual Browser Architecture
        port1 = int(os.environ.get("CHROME_PORT_TAB1", "19899"))
        port2 = int(os.environ.get("CHROME_PORT_TAB2", "19900"))

        async def launch_and_login(user_data_dir, port, context_name):
            # 1. First attempt attaching over CDP if Chrome is already running on that port
            try:
                browser = await pw.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
                print(f"[Setup] [{context_name}] Attached to existing Chrome over CDP on port {port}")
                return browser.contexts[0]
            except Exception:
                pass

            print(f"[Setup] Launching Chromium persistent context for {context_name} on port {port}...")
            chrome_args = [
                "--disable-features=CalculateNativeWinOcclusion",
                "--disable-background-timer-throttling",
                "--disable-backgrounding-occluded-windows",
                "--disable-renderer-backgrounding",
                "--start-maximized",
                "--disable-dev-shm-usage",
                "--disable-gpu-process-crash-limit",
                f"--remote-debugging-port={port}"
            ]
            if is_linux:
                chrome_args.extend([
                    "--no-sandbox",
                    "--disable-setuid-sandbox",
                    "--disable-gpu"
                ])

            launch_kwargs = {
                "headless": headless_flag,
                "viewport": {"width": 1920, "height": 1080},
                "args": chrome_args
            }
            if exec_path:
                launch_kwargs["executable_path"] = exec_path

            # Pre-clean stale Singleton lock files
            for lock_file in ("SingletonLock", "SingletonSocket", "SingletonCookie", "lockfile"):
                lp = os.path.join(user_data_dir, lock_file)
                if os.path.exists(lp) or os.path.islink(lp):
                    try:
                        os.remove(lp)
                    except Exception:
                        pass

            ctx = None
            for launch_attempt in range(3):
                try:
                    ctx = await pw.chromium.launch_persistent_context(
                        user_data_dir,
                        **launch_kwargs
                    )
                    break
                except Exception as launch_err:
                    print(f"[Setup] [{context_name}] Launch attempt {launch_attempt+1} failed: {launch_err}")
                    await asyncio.sleep(2.0)
                    if launch_attempt == 2:
                        raise launch_err

            return ctx

        focus_lock = asyncio.Lock()
        
        # 1. Initialize TAB_1 context and tab
        ctx1 = await launch_and_login(user_data_dir_1, port1, "TAB_1")
        tab1 = CoinglassTab(ctx1, TAB1_SYMBOLS, store, "TAB_1")
        tab1.focus_lock = focus_lock
        await tab1.start()
        await tab1.inject_and_configure_all(focus_lock)

        # 2. Initialize TAB_2 context and tab
        ctx2 = await launch_and_login(user_data_dir_2, port2, "TAB_2")
        tab2 = CoinglassTab(ctx2, TAB2_SYMBOLS, store, "TAB_2")
        tab2.focus_lock = focus_lock
        await tab2.start()
        await tab2.inject_and_configure_all(focus_lock)

        binance = BinanceFootprintFeed(ALL_SYMBOLS, store)
        binance_ws = BinanceTradePriceWebSocketFeed(ALL_SYMBOLS, store)

        # 4. Historical Seeding
        from concurrent.futures import ThreadPoolExecutor
        excel_pool = ThreadPoolExecutor(max_workers=4)

        if skip_seed:
            print("[Setup] --skip-seed flag active. Skipping historical seeding.")
        else:
            async def seed_wrapper(tab: CoinglassTab, sym: str):
                try:
                    for attempt in range(3):
                        try:
                            if not tab.page or tab.page.is_closed():
                                print(f"[{tab.tab_id}] [RECOVERY] Page closed on seeding attempt {attempt+1}. Reconnecting...")
                                await tab.reconnect(focus_lock)
                            await tab.seed_symbol(sym, excel_pool, focus_lock)
                            break
                        except Exception as e:
                            print(f"[Setup] [WARN] Seeding failed for {sym} (attempt {attempt+1}/3): {e}")
                            if "closed" in str(e).lower() or "navigation" in str(e).lower() or "locator" in str(e).lower() or "timeout" in str(e).lower():
                                try:
                                    await tab.reconnect(focus_lock)
                                except Exception as rec_err:
                                    print(f"[Setup] [ERROR] Failed to reconnect tab during seeding retry: {rec_err}")
                            if attempt == 2:
                                raise
                            await asyncio.sleep(3.0)
                finally:
                    tab.is_seeding = False

            async def seed_tab(tab: CoinglassTab, symbols: list):
                for sym in symbols:
                    await seed_wrapper(tab, sym)

            print("[Setup] Launching historical seeding...")
            await asyncio.gather(
                seed_tab(tab1, TAB1_SYMBOLS),
                seed_tab(tab2, TAB2_SYMBOLS)
            )
            print("[Setup] Seeding phase complete! Starting real-time feeds...")
            combine_seeding_files()
        
        # 5. Run Live feeds & Terminal display
        async def tab_switcher():
            active_tab = tab1
            while not stop.is_set():
                await asyncio.sleep(5.0)
                if stop.is_set():
                    break
                if tab1.is_seeding or tab2.is_seeding:
                    continue
                try:
                    try:
                        async with asyncio.timeout(3.0):
                            async with focus_lock:
                                if active_tab.page and not active_tab.page.is_closed():
                                    await active_tab.page.bring_to_front()
                    except asyncio.TimeoutError:
                        log_live_event(f"Warning: focus_lock timeout. Bypassing lock for {active_tab.tab_id}.", "Switch")
                        if active_tab.page and not active_tab.page.is_closed():
                            await active_tab.page.bring_to_front()
                            
                    active_tab = tab2 if active_tab is tab1 else tab1
                except Exception as e:
                    log_live_event(f"Failed to switch to {active_tab.tab_id}: {e}", "Switch")

        async def rollover_watchdog(tracker, stop_event):
            while not stop_event.is_set():
                try:
                    tracker.update_day()
                    # Non-blocking Binance position sync (prevents order-tracking drift)
                    if hasattr(tracker, "reconcile_with_broker"):
                        await asyncio.to_thread(tracker.reconcile_with_broker)
                except Exception as ex:
                    log_live_event(f"Rollover watchdog failed: {ex}", "WDog")
                await asyncio.sleep(30.0)  # tighter sync cadence for exit safety

        async def event_loop_monitor(stop_event: asyncio.Event, threshold_sec: float = 2.0) -> None:
            consecutive_blocks = 0
            while not stop_event.is_set():
                start = time.time()
                await asyncio.sleep(0.1)
                elapsed = time.time() - start - 0.1
                if elapsed > threshold_sec:
                    consecutive_blocks += 1
                    if consecutive_blocks >= 10:
                        pass
                else:
                    consecutive_blocks = 0

        tasks = [
            asyncio.create_task(tab1.poll_loop()),
            asyncio.create_task(tab2.poll_loop()),
            asyncio.create_task(event_loop_monitor(stop)),
            asyncio.create_task(binance.run()),
            asyncio.create_task(binance_ws.run()),
            asyncio.create_task(renderer_loop(store, stop)),
            asyncio.create_task(watchdog([tab1, tab2, binance, binance_ws], focus_lock, stop)),
            asyncio.create_task(tab_switcher()),
            asyncio.create_task(rollover_watchdog(trade_tracker, stop))
        ]
        
        # Handle graceful exit triggers
        loop = asyncio.get_running_loop()
        def sig_handler():
            print("\n[Exit] Termination signal received. Stopping...")
            stop.set()
            tab1.running = False
            tab2.running = False
            binance.running = False
            
        for sig in (signal.SIGINT, signal.SIGTERM):
            try:
                loop.add_signal_handler(sig, sig_handler)
            except NotImplementedError:
                pass
                
        try:
            while not stop.is_set():
                await asyncio.sleep(1.0)
        except (KeyboardInterrupt, SystemExit):
            sig_handler()
        finally:
            print("[Setup] Cleaning up tasks and closing browser...")
            for t in tasks:
                t.cancel()
            await asyncio.gather(*tasks, return_exceptions=True)
            excel_pool.shutdown(wait=True)
            await ctx.close()
        
    print("[Exit] Shutdown complete.")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Coinglass + Binance Footprint Scraper")
    parser.add_argument("--skip-seed", "--skip-seeding", action="store_true", help="Skip historical Excel seeding and go straight to live feeds")
    parser.add_argument("--skip-train", action="store_true", help="Skip initial model retraining at startup")
    parser.add_argument("--close-chrome", "--kill-chrome", action="store_true", help="Forcefully close all active Chrome and Chromium instances and exit")
    args = parser.parse_args()

    if args.close_chrome:
        close_all_chrome_instances()
        print("[Exit] All Chrome instances terminated successfully.")
        sys.exit(0)

    asyncio.run(main(skip_seed=args.skip_seed, skip_train=args.skip_train))
