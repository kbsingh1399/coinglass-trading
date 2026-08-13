"""
Six Strategy Engine — Unified Live Predictor
=============================================
Ports the exact logic from colab_strategies/run_all_6.py into a live streaming predictor.

Strategies:
  S1 - Liquidation:    Trend pullback + abnormal liquidation spike
  S2 - CVD Momentum:   Tight trend pullback on strong CVD moves
  S3 - Trend Follow:   Classic macro trend pullback (EMA 200/800)
  S4 - Mean Reversion: RSI extremes with deep pullback
  S5 - Vol Breakout:   Trend pullback + elevated volatility + CVD
  S6 - OI Coherence:   Trend pullback + OI/CVD directional agreement

All strategies share:
  - Same feature engineering (featurize)
  - Same ML pipeline (LGB + XGB ensemble)
  - Same trade parameters (TP=5R, Trail=0.8ATR, SL=1ATR, max 288 bars)
  - Same walk-forward validation
"""

import os
import sys
import json
import time
import collections
import threading
import numpy as np
import pandas as pd
from typing import Dict, List, Any, Optional
from numba import njit

# ─── Constants (match run_all_6.py exactly) ──────────────────────────
TP_MULT = 5.0
TRAIL_ATR = 0.8
SL_MULT = 1.0
MAX_BARS = 288       # 72 hours of 15m bars
RISK_PCT = 0.004     # 0.4% per trade (matches RSK=20 on $5000)
FEE_PCT = 0.0015     # Round-trip fee

SYMBOLS = [
    'BTCUSDT', 'ETHUSDT', 'BNBUSDT', 'SOLUSDT', 'XRPUSDT', 'DOGEUSDT', 'ADAUSDT',
    'TRXUSDT', 'AVAXUSDT', 'DOTUSDT', 'LINKUSDT', 'LTCUSDT', 'NEARUSDT', 'SUIUSDT'
]

STRATEGY_NAMES = {
    'S1': 'S1_Liquidation',
    'S2': 'S2_CVD_Momentum',
    'S3': 'S3_Trend_Follow',
    'S4': 'S4_Mean_Reversion',
    'S5': 'S5_Vol_Breakout',
    'S6': 'S6_OI_Coherence',
}


# ─── Numba Trade Simulation (exact copy from run_all_6.py) ──────────
@njit(fastmath=True, nogil=True)
def _sim_trade(h, l, c, entry_idx, entry, atr, dr):
    """Simulate a single trade forward from entry_idx."""
    n = len(c)
    sd = atr * SL_MULT
    td = atr * TP_MULT
    trd = atr * TRAIL_ATR
    st = entry - sd if dr == 1 else entry + sd
    cs = st  # current stop
    bp = entry  # best price
    ns = st  # new stop
    mx = min(entry_idx + MAX_BARS + 1, n)
    ep = c[mx - 1]  # exit price
    bh = mx - 1 - entry_idx  # bars held

    for j in range(entry_idx + 1, mx):
        if dr == 1:
            if l[j] <= cs:
                ep = cs; bh = j - entry_idx; break
            if h[j] > bp:
                bp = h[j]
            if (bp - entry) >= td:
                ns = bp - trd
            if ns > cs:
                cs = ns
        else:
            if h[j] >= cs:
                ep = cs; bh = j - entry_idx; break
            if l[j] < bp:
                bp = l[j]
            if (entry - bp) >= td:
                ns = bp + trd
            if ns < cs:
                cs = ns

    units = RISK_PCT / sd if sd > 0 else 0
    gross = units * (ep - entry) if dr == 1 else units * (entry - ep)
    fees = units * entry * FEE_PCT / 2.0 + units * abs(ep) * FEE_PCT / 2.0
    net_pnl = gross - fees
    r_mult = net_pnl / (RISK_PCT) if RISK_PCT > 0 else 0
    win = 1.0 if net_pnl > 0 else 0.0
    return net_pnl, r_mult, win, bh


# ─── Z-Score Helper ──────────────────────────────────────────────────
def _zscore(series, window):
    """Rolling z-score."""
    mean = series.rolling(window, min_periods=1).mean()
    std = series.rolling(window, min_periods=1).std().replace(0, 1e-10)
    return (series - mean) / std


# ─── Feature Engineering (exact copy from run_all_6.py) ──────────────
def featurize(df, btc_ref=None):
    """Compute all features needed by the 6 strategies."""
    if btc_ref is not None:
        cj = [c for c in btc_ref.columns if c not in df.columns]
        if cj:
            df = df.join(btc_ref[cj], how='left')
        if 'btc_CVD' in df.columns:
            df['btc_CVD'] = df['btc_CVD'].ffill().bfill().fillna(0)

    # ATR
    df['atr'] = (df['High'] - df['Low']).rolling(14, min_periods=1).mean()

    # CVD features
    if 'CVD' in df.columns:
        df['cvd_d'] = df['CVD'].diff(5)
        for k in [4, 10, 20]:
            df[f'zc{k}'] = _zscore(df['CVD'], k)
    else:
        df['cvd_d'] = 0.0
        for k in [4, 10, 20]:
            df[f'zc{k}'] = 0.0

    # BTC CVD features
    df['bcvm'] = df['btc_CVD'].diff(2) if 'btc_CVD' in df.columns else 0.0
    for k in [4, 10, 20]:
        df[f'zb{k}'] = _zscore(df['btc_CVD'], k) if 'btc_CVD' in df.columns else 0.0

    # Macro signal: EMA 200/800 crossover
    df['ef'] = df['Close'].ewm(span=200, min_periods=50).mean()
    df['es'] = df['Close'].ewm(span=800, min_periods=100).mean()
    atrs = df['atr'].replace(0, 1e-10)
    df['mc'] = np.where(
        (df['ef'] - df['es']) / atrs > 0.5, 1,
        np.where((df['ef'] - df['es']) / atrs < -0.5, -1, 0)
    )

    # EMA pullbacks
    for s, n in [(8, 'e8'), (21, 'e21'), (50, 'e50')]:
        df[n] = df['Close'].ewm(span=s, min_periods=1).mean()
    df['p8'] = (df['Close'] - df['e8']) / atrs
    df['p21'] = (df['Close'] - df['e21']) / atrs
    df['p50'] = (df['Close'] - df['e50']) / atrs

    # RSI
    d = df['Close'].diff()
    g = d.clip(lower=0).rolling(14, min_periods=1).mean()
    l = (-d.clip(upper=0)).rolling(14, min_periods=1).mean()
    df['rsi'] = 100 - (100 / (1 + g / l.replace(0, 1e-10)))

    # Volatility regime
    df['vr'] = _zscore(df['atr'], 100)

    # Liquidation features (support alternate column aliases: Agg. Liq Short/Long, liq_short/long, liquidations_short/long)
    for s, default_col in [('l', 'Agg. Liq Long'), ('s', 'Agg. Liq Short')]:
        col = None
        for candidate in [default_col, f'liq_{"long" if s=="l" else "short"}', f'liquidations_{"long" if s=="l" else "short"}', f'liq{s}', f'Agg. Liq. {"Long" if s=="l" else "Short"}']:
            if candidate in df.columns:
                col = candidate
                break
        if col is not None:
            df[f'liq{s}'] = pd.to_numeric(df[col], errors='coerce').fillna(0).rolling(5, min_periods=1).sum()
            df[f'liq{s}m'] = df[f'liq{s}'].rolling(100, min_periods=1).mean()
        else:
            df[f'liq{s}'] = 0.0
            df[f'liq{s}m'] = 0.0

    # Open Interest features
    if 'Agg. OI' in df.columns:
        oi = pd.to_numeric(df['Agg. OI'], errors='coerce').ffill()
        df['zoi'] = _zscore(oi, 100)
        df['oid'] = oi.diff(5) / (oi.shift(5) + 1e-10)
        df['oicc'] = np.sign(df['oid'].fillna(0)) * np.sign(df['cvd_d'].fillna(0))
    else:
        df['zoi'] = 0.0
        df['oid'] = 0.0
        df['oicc'] = 0.0

    # LS Ratio
    if 'Long/Short Ratio (Account)' in df.columns:
        df['zls'] = _zscore(pd.to_numeric(df['Long/Short Ratio (Account)'], errors='coerce').ffill(), 100)
    else:
        df['zls'] = 0.0

    # Funding Rate
    if 'Agg. Funding Rate' in df.columns:
        fr = pd.to_numeric(df['Agg. Funding Rate'], errors='coerce').fillna(0)
        # PARITY GUARD: parquet stores decimal fractions (~0.0001–0.001).
        # If live scraper delivers percentage form (|val| >= 0.005), normalize to decimal.
        fr = fr.apply(lambda v: v / 100.0 if abs(v) >= 0.005 else v)
        df['fr'] = fr
        df['zfr'] = _zscore(fr, 20)
    else:
        df['fr'] = 0.0
        df['zfr'] = 0.0

    # Footprint Delta synthesis if missing but Ask/Bid Qty present
    if 'Delta Qty' not in df.columns:
        if 'Ask Qty' in df.columns and 'Bid Qty' in df.columns:
            df['Delta Qty'] = pd.to_numeric(df['Ask Qty'], errors='coerce').fillna(0) - pd.to_numeric(df['Bid Qty'], errors='coerce').fillna(0)
        elif 'Buy Qty' in df.columns and 'Sell Qty' in df.columns:
            df['Delta Qty'] = pd.to_numeric(df['Buy Qty'], errors='coerce').fillna(0) - pd.to_numeric(df['Sell Qty'], errors='coerce').fillna(0)

    # Footprint features
    for c in ['Bid Qty', 'Ask Qty', 'Delta Qty', 'Bid Trades', 'Ask Trades']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
            df[f'z{c.replace(" ", "_").lower()}'] = _zscore(df[c], 10)
        else:
            df[f'z{c.replace(" ", "_").lower()}'] = 0.0

    # Buy/Sell ratio
    if 'Buy Qty' in df.columns and 'Sell Qty' in df.columns:
        buy = pd.to_numeric(df['Buy Qty'], errors='coerce').fillna(0)
        sell = pd.to_numeric(df['Sell Qty'], errors='coerce').fillna(0)
        df['bsr'] = buy / (buy + sell + 1e-10)
    else:
        df['bsr'] = 0.5

    # Volume ratio
    df['vr5'] = df['Volume'] / (df['Volume'].rolling(20, min_periods=1).mean() + 1e-10)

    df = df.fillna(0).replace([np.inf, -np.inf], 0)
    return df


# ─── Signal Generators (exact copy from run_all_6.py) ────────────────
def make_signal_s1(row):
    """S1: Trend pullback + liquidation confirmation"""
    mc, p8 = row.get('mc', 0), row.get('p8', 0)
    ll, llm = row.get('liql', 0), row.get('liqlm', 0)
    ls, lsm = row.get('liqs', 0), row.get('liqsm', 0)
    zc20 = row.get('zc20', 0)

    if mc > 0 and p8 < -0.12 and (ll > llm * 1.2 or zc20 > 0.1):
        return 1
    if mc < 0 and p8 > 0.12 and (ls > lsm * 1.2 or zc20 < -0.1):
        return -1
    return 0

def make_signal_s2(row):
    """S2: CVD Momentum — tighter pullback"""
    mc, p8 = row.get('mc', 0), row.get('p8', 0)
    if mc > 0 and p8 < -0.25:
        return 1
    if mc < 0 and p8 > 0.25:
        return -1
    return 0

def make_signal_s3(row):
    """S3: Pure trend pullback"""
    mc, p8 = row.get('mc', 0), row.get('p8', 0)
    if mc > 0 and p8 < -0.2:
        return 1
    if mc < 0 and p8 > 0.2:
        return -1
    return 0

def make_signal_s4(row):
    """S4: RSI mean reversion"""
    rsi, p8 = row.get('rsi', 50), row.get('p8', 0)
    if rsi < 35 and p8 < -0.5:
        return 1
    if rsi > 65 and p8 > 0.5:
        return -1
    return 0

def make_signal_s5(row):
    """S5: Vol Breakout — trend pullback + vol bonus"""
    mc, p8 = row.get('mc', 0), row.get('p8', 0)
    vr, zc20 = row.get('vr', 0), row.get('zc20', 0)
    rsi = row.get('rsi', 50)

    # Core: trend pullback like S3
    if mc > 0 and p8 < -0.2:
        return 1
    if mc < 0 and p8 > 0.2:
        return -1
    # Bonus: high-vol regime entries
    if mc > 0 and p8 < -0.1 and vr > 1.5 and zc20 > 0.15 and 25 < rsi < 75:
        return 1
    if mc < 0 and p8 > 0.1 and vr > 1.5 and zc20 < -0.15 and 25 < rsi < 75:
        return -1
    return 0

def make_signal_s6(row):
    """S6: OI Coherence — trend pullback + OI/CVD bonus"""
    mc, p8 = row.get('mc', 0), row.get('p8', 0)
    oicc, zc20 = row.get('oicc', 0), row.get('zc20', 0)

    # Core: trend pullback like S3
    if mc > 0 and p8 < -0.2:
        return 1
    if mc < 0 and p8 > 0.2:
        return -1
    # Bonus: OI-CVD coherence
    if mc > 0 and p8 < -0.1 and oicc != 0 and oicc > 0.2 and zc20 > 0.1:
        return 1
    if mc < 0 and p8 > 0.1 and oicc != 0 and oicc < -0.2 and zc20 < -0.1:
        return -1
    return 0

SIGNAL_FUNCS = {
    'S1': make_signal_s1,
    'S2': make_signal_s2,
    'S3': make_signal_s3,
    'S4': make_signal_s4,
    'S5': make_signal_s5,
    'S6': make_signal_s6,
}


# ─── ML Model Training (matches run_all_6.py bmodel) ────────────────
def train_ensemble(X, y):
    """Train LGB + XGB ensemble with feature importance selection."""
    import lightgbm as lgb
    try:
        import xgboost as xgb
        has_xgb = True
    except ImportError:
        has_xgb = False

    if len(X) < 20 or y.sum() < 3 or (len(y) - y.sum()) < 3:
        return None, list(X.columns)

    p = y.sum()
    sw = max(0.1, float((len(y) - p) / p)) if p > 0 else 1.0

    # Feature selection via LGB importance
    sel = lgb.LGBMClassifier(n_estimators=30, max_depth=3, random_state=42,
                              verbose=-1, n_jobs=1, max_bin=31)
    sel.fit(X, y)
    imps = sel.feature_importances_
    cut = np.percentile(imps, 15)
    selected = [c for c, im in zip(X.columns, imps) if im >= cut]
    if len(selected) < 3:
        selected = list(X.columns)

    models = []

    # LightGBM
    m_lgb = lgb.LGBMClassifier(
        max_depth=5, learning_rate=0.02, n_estimators=200,
        scale_pos_weight=sw, random_state=42, n_jobs=1, verbose=-1,
        max_bin=63, min_child_samples=8, subsample=0.8,
        colsample_bytree=0.8, reg_alpha=0.1, reg_lambda=0.1
    )
    m_lgb.fit(X[selected], y)
    models.append(m_lgb)

    # XGBoost
    if has_xgb:
        m_xgb = xgb.XGBClassifier(
            max_depth=4, learning_rate=0.03, n_estimators=200,
            scale_pos_weight=sw, random_state=42, n_jobs=1,
            verbosity=0, subsample=0.8, colsample_bytree=0.8, reg_alpha=0.1
        )
        m_xgb.fit(X[selected], y)
        models.append(m_xgb)

    return models, selected


def predict_ensemble(models, selected_cols, X):
    """Ensemble average prediction."""
    vc = [c for c in selected_cols if c in X.columns]
    if not vc:
        return np.full(len(X), 0.5)
    probs = [m.predict_proba(X[vc].astype(np.float32))[:, 1] for m in models]
    return np.mean(probs, axis=0)


# ─── Unified Live Predictor Class ────────────────────────────────────
class LiveSixStrategyPredictor:
    """
    Runs all 6 strategies from run_all_6.py on live streaming data.
    
    On each 15m candle close:
    1. Compute features via featurize()
    2. Generate signals via make_signal_s1..s6
    3. Filter via ML ensemble (if trained)
    4. Dispatch trades via trade_tracker.trigger_entry()
    """

    def __init__(self, symbols: List[str]):
        self.symbols = symbols
        self.candles_history: Dict[str, collections.deque] = {}
        self.current_candle: Dict[str, dict] = {}
        self._last_predict_bar: Dict[str, int] = {}
        self._cached_signals: Dict[str, Dict[str, str]] = {s: {} for s in symbols}
        self._lock = threading.RLock()

        # ML models per strategy per symbol
        self.models: Dict[str, Dict[str, Any]] = {k: {} for k in SIGNAL_FUNCS}
        self.selected_cols: Dict[str, Dict[str, list]] = {k: {} for k in SIGNAL_FUNCS}
        self.thresholds: Dict[str, Dict[str, float]] = {k: {s: 0.55 for s in symbols} for k in SIGNAL_FUNCS}

        # BTC reference for cross-asset features
        self.btc_ref = None

        self.load_models()

    def load_models(self):
        """Load pre-trained models from disk."""
        base_dir = os.path.dirname(os.path.abspath(__file__))
        models_dir = os.path.join(base_dir, 'six_strategy_models')
        if not os.path.exists(models_dir):
            print(f"[SixStrategy] No pre-trained models at {models_dir} — will train on first data")
            return

        import pickle
        for strat_key in SIGNAL_FUNCS:
            for sym in self.symbols:
                path = os.path.join(models_dir, f'{strat_key}_{sym}.pkl')
                if os.path.exists(path):
                    try:
                        with open(path, 'rb') as f:
                            data = pickle.load(f)
                        self.models[strat_key][sym] = data['models']
                        self.selected_cols[strat_key][sym] = data['selected_cols']
                        self.thresholds[strat_key][sym] = data.get('threshold', 0.55)
                    except Exception as e:
                        print(f"[SixStrategy] Error loading {strat_key}_{sym}: {e}")

        total = sum(len(v) for v in self.models.values())
        print(f"[SixStrategy] Loaded {total} models across {len(SIGNAL_FUNCS)} strategies")

    def set_history(self, symbol: str, candles):
        """Set historical candle data for a symbol."""
        now_open = int(time.time() // 900) * 900
        cleaned = []
        for c in candles:
            try:
                ot = int(c.get('open_time', 0))
            except Exception:
                continue
            if ot > 0 and ot < now_open:
                row = dict(c)
                row['open_time'] = ot
                cleaned.append(row)

        cleaned.sort(key=lambda r: r['open_time'])
        cleaned = cleaned[-1200:]
        self.candles_history[symbol] = collections.deque(cleaned, maxlen=1200)
        if cleaned:
            self._last_predict_bar[symbol] = cleaned[-1]['open_time']

    def load_history_from_disk(self):
        """Load historical candles from combined_seed_history.xlsx."""
        base_dir = os.path.dirname(os.path.abspath(__file__))
        combined_path = os.path.join(base_dir, "Seeding", "combined_seed_history.xlsx")
        if not os.path.exists(combined_path):
            print(f"[SixStrategy] No combined seeding file at {combined_path}")
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(combined_path, read_only=True)
            loaded = 0
            for sheetname in wb.sheetnames:
                sym = sheetname
                if sym not in self.symbols:
                    continue
                ws = wb[sym]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                headers = rows[0]
                data_rows = rows[1:][-1200:]
                candle_list = []
                EXCEL_EPOCH_OFFSET = 25569  # days between 1900-01-01 and 1970-01-01
                for row in data_rows:
                    d = dict(zip(headers, row))
                    val = d.get("open_time")
                    if val is None:
                        continue
                    if hasattr(val, 'timestamp'):
                        from datetime import timezone
                        try:
                            d["open_time"] = int(val.replace(tzinfo=timezone.utc).timestamp())
                        except Exception:
                            d["open_time"] = int(val.timestamp())
                    elif isinstance(val, (int, float)):
                        v = float(val)
                        # Excel serial date: values < 100000 are day-counts from 1900-01-01
                        if v < 100_000:
                            d["open_time"] = int((v - EXCEL_EPOCH_OFFSET) * 86400)
                        else:
                            d["open_time"] = int(v)
                    else:
                        continue
                    candle_list.append(d)
                self.set_history(sym, candle_list)
                loaded += 1
            print(f"[SixStrategy] Loaded history for {loaded} symbols from disk cache.")
        except Exception as e:
            print(f"[SixStrategy] Error loading disk history: {e}")

    def on_tick_update(self, symbol: str, snap, trade_tracker=None):
        """Called on every tick. Only runs prediction on candle close."""
        with self._lock:
            return self._on_tick_locked(symbol, snap, trade_tracker)

    def _on_tick_locked(self, symbol, snap, trade_tracker):
        if snap.price <= 0:
            return snap

        now = time.time()
        open_time = int(now // 900) * 900

        if symbol not in self.candles_history:
            self.candles_history[symbol] = collections.deque(maxlen=1200)

        history = self.candles_history[symbol]

        # Candle rollover
        if symbol not in self.current_candle or self.current_candle[symbol].get('open_time') != open_time:
            prev = self.current_candle.get(symbol)
            if prev and int(prev.get('open_time', 0)) < open_time:
                prev_ot = int(prev['open_time'])
                if not history or int(history[-1].get('open_time', 0)) != prev_ot:
                    history.append(dict(prev))
            self.current_candle[symbol] = {
                'open_time': open_time, 'open': snap.price, 'high': snap.price,
                'low': snap.price, 'close': snap.price, 'volume': snap.volume,
                'fut_cvd': snap.fut_cvd, 'spot_cvd': snap.spot_cvd,
                'funding': snap.funding, 'liq_long': snap.liq_long,
                'liq_short': snap.liq_short, 'ls_ratio': snap.ls_ratio,
                'oi': snap.oi, 'coins_bid': snap.coins_bid,
                'coins_ask': snap.coins_ask, 'dollars_bid': snap.dollars_bid,
                'dollars_ask': snap.dollars_ask, 'whale_idx': snap.whale_idx,
                'tk_buy_cnt': snap.tk_buy_cnt, 'tk_sell_cnt': snap.tk_sell_cnt,
                'fp_poc': snap.fp_poc,
            }
        else:
            c = self.current_candle[symbol]
            c['close'] = snap.price
            if snap.price > c['high']: c['high'] = snap.price
            if snap.price < c['low'] or c['low'] == 0: c['low'] = snap.price
            c['volume'] = snap.volume
            c['fut_cvd'] = snap.fut_cvd
            c['liq_long'] = snap.liq_long
            c['liq_short'] = snap.liq_short
            c['oi'] = snap.oi

        # Only predict on candle close
        last_bar = history[-1].get('open_time', 0) if history else 0
        if last_bar == self._last_predict_bar.get(symbol, 0):
            return snap  # Already predicted this bar

        if len(history) < 250:
            import dataclasses
            return dataclasses.replace(snap, strategy_armed=f"WARM({len(history)}/250)")

        self._last_predict_bar[symbol] = last_bar

        # Build DataFrame for feature engineering
        try:
            df = self._build_df(symbol)
            if df is None or len(df) < 250:
                import dataclasses
                return dataclasses.replace(snap, strategy_armed=f"WARM({len(history)}/250)")

            # Get BTC reference
            btc_ref = None
            if symbol != 'BTCUSDT' and 'BTCUSDT' in self.candles_history:
                btc_df = self._build_df('BTCUSDT')
                if btc_df is not None:
                    btc_ref = btc_df[['Close', 'CVD']].copy() if 'CVD' in btc_df.columns else btc_df[['Close']].copy()
                    btc_ref.columns = [f'btc_{c}' for c in btc_ref.columns]

            # Featurize
            df = featurize(df.copy(), btc_ref)
            last_row = df.iloc[-1].to_dict()
            atr_val = float(last_row.get('atr', 0))
            if atr_val <= 0 or np.isnan(atr_val):
                return snap
            # Floor ATR at 0.05% of price so SL never falls below MIN_STOP_PCT thresholds
            atr_val = max(atr_val, snap.price * 0.0005)

            # Run all 6 strategies
            armed_parts = []

            # GUARD: Skip symbols that have no trained models for ANY strategy.
            # Trading a symbol without backtest-validated models is unvalidated speculation.
            modeled_strategies = {sk for sk in SIGNAL_FUNCS if symbol in self.models.get(sk, {})}
            if not modeled_strategies:
                import dataclasses
                return dataclasses.replace(snap, strategy_armed="NO_MODEL")

            for strat_key, signal_func in SIGNAL_FUNCS.items():
                direction = signal_func(last_row)
                if direction == 0:
                    continue

                strat_name = STRATEGY_NAMES[strat_key]

                # Check for active trade in this strategy
                if trade_tracker:
                    with trade_tracker.lock:
                        has_active = any(
                            t['symbol'] == symbol and t['strategy'] == strat_name
                            for t in trade_tracker.active_trades.values()
                        )
                    if has_active:
                        continue

                # ML filter (if model available)
                if symbol in self.models.get(strat_key, {}):
                    try:
                        fcs = self.selected_cols[strat_key][symbol]
                        X = pd.DataFrame([{c: last_row.get(c, 0) for c in fcs}]).astype(np.float32)
                        prob = predict_ensemble(
                            self.models[strat_key][symbol], fcs, X
                        )[0]
                        thresh = self.thresholds[strat_key].get(symbol, 0.55)
                        if prob < thresh:
                            continue
                    except Exception as e:
                        print(f"[SixStrategy] ML failed closed {strat_key} {symbol}: {e}")
                        continue  # If ML fails, DO NOT let signal through

                # Compute SL/TP
                sl = snap.price - SL_MULT * atr_val if direction == 1 else snap.price + SL_MULT * atr_val
                tp = snap.price + TP_MULT * atr_val if direction == 1 else snap.price - TP_MULT * atr_val

                # Dispatch trade
                if trade_tracker:
                    trade_tracker.trigger_entry(
                        symbol, strat_name, direction, snap.price,
                        sl, tp, atr_val, macro=int(last_row.get('mc', 0)),
                        vol_regime=float(last_row.get('vr', 0)),
                        risk_mult=1.0, trail_act=TP_MULT, regime_val=0
                    )

                dir_str = 'LONG' if direction == 1 else 'SHORT'
                armed_parts.append(f"{strat_key}:{dir_str}")

            # Cache armed signals for display
            self._cached_signals[symbol] = {'armed_str': ' | '.join(armed_parts) if armed_parts else ''}

        except Exception as e:
            import traceback
            print(f"[SixStrategy] {symbol} error: {e}\n{traceback.format_exc()}")

        # Replay cached signal
        cached = self._cached_signals.get(symbol, {})
        armed_str = cached.get('armed_str', '')

        # Show active trades
        if trade_tracker:
            with trade_tracker.lock:
                trades = [t for t in trade_tracker.active_trades.values() if t['symbol'] == symbol]
            if trades:
                parts = []
                for t in trades:
                    d = 'LONG' if t['direction'] == 1 else 'SHORT'
                    pnl = t.get('live_pnl_pct', 0)
                    sk = t.get('strategy', '?')[:2]
                    parts.append(f"{sk}:{d}({pnl:+.1f}%)")
                armed_str = ' '.join(parts)
        elif not armed_str:
            armed_str = "READY"

        import dataclasses
        snap = dataclasses.replace(snap, strategy_armed=armed_str)
        return snap

    def _build_df(self, symbol):
        """Build a DataFrame from candle history."""
        history = list(self.candles_history.get(symbol, []))
        if not history:
            return None

        df = pd.DataFrame(history)
        # Map to expected column names
        col_map = {
            'open': 'Open', 'high': 'High', 'low': 'Low', 'close': 'Close',
            'volume': 'Volume', 'fut_cvd': 'CVD', 'oi': 'Agg. OI',
            'ls_ratio': 'Long/Short Ratio (Account)', 'funding': 'Agg. Funding Rate',
            'liq_long': 'Agg. Liq Long', 'liq_short': 'Agg. Liq Short',
            'coins_bid': 'Bid Qty', 'coins_ask': 'Ask Qty',
            'tk_buy_cnt': 'Bid Trades', 'tk_sell_cnt': 'Ask Trades',
        }
        for old, new in col_map.items():
            if old in df.columns:
                df[new] = pd.to_numeric(df[old], errors='coerce').fillna(0)

        # Timestamp index
        if 'open_time' in df.columns:
            df['ts'] = pd.to_datetime(df['open_time'], unit='s')
            df = df.set_index('ts').sort_index()

        required = ['Open', 'High', 'Low', 'Close', 'Volume']
        for r in required:
            if r not in df.columns:
                return None

        return df
