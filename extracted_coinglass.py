class CoinglassTab:
    def __init__(self, context: BrowserContext, symbols: List[str], store: SnapshotStore, tab_id: str):
        self.context = context
        self.symbols = symbols
        self.store = store
        self.tab_id = tab_id
        self.is_seeding = False
        self.page: Optional[Page] = None
        self.last_heartbeat_ns = time.time_ns()
        self.running = True
        self._response_tasks: set[asyncio.Task] = set()
        self.poll_failures = 0

    async def start(self) -> None:
        self.page = await self.context.new_page()
        
        # Suppress noisy TradingView internal console spam; only print errors and CoinGlass messages
        def _on_console(msg):
            text = msg.text
            typ = msg.type
            skip_patterns = (
                "Recurring script engine stop",
                "76 custom indicators loaded",
                "Content Security Policy",
                "WebSocket connection to",
                "ERR_NAME_NOT_RESOLVED",
                "502",
                "wss.coinglass.com",
                "net::ERR_",
                "Failed to fetch",
            )
            if any(p in text for p in skip_patterns):
                return
            if typ in ("error", "warning") or "coinglass" in text.lower():
                print(f"[{self.tab_id} CONSOLE] {typ} {text}")

        def _on_page_error(exc):
            msg = str(exc)
            # Filter generic browser resource errors that are not actionable
            if any(p in msg for p in ("unknown compression", "net::", "ERR_", "Failed to fetch", "ResizeObserver", "reading 'symbol'")):
                return
            print(f"[{self.tab_id} PAGE ERROR] {msg}")

        self.page.on("console", _on_console)
        self.page.on("pageerror", _on_page_error)
        
        # Intercept HTTP API responses natively to capture Open Interest and Funding Rates securely
        # without introducing compression encoding errors on the page.
        async def handle_response(response):
            try:
                url = response.url
                if any(k in url for k in ("open-interest", "funding-rate", "liquidation", "long-short", "rsi", "cumulative-volume")):
                    body = await response.text()
                    await self._route_payload({"url": url, "body": body})
            except Exception:
                pass

        def _spawn_response_task(response):
            task = asyncio.create_task(handle_response(response))
            self._response_tasks.add(task)
            task.add_done_callback(self._response_tasks.discard)

        self.page.on("response", _spawn_response_task)
        
        print(f"[{self.tab_id}] Opening layout: {URL}...")
        await self.page.goto(URL, wait_until="load", timeout=45000)
        print(f"[{self.tab_id}] Waiting 15 seconds for layout to load...")
        await asyncio.sleep(15)  # Wait 15 seconds for S9 chart renders

    async def reconnect(self, focus_lock: asyncio.Lock) -> None:
        print(f"[{self.tab_id}] [RECOVERY] Attempting to reconnect/restart the tab...")
        self.is_seeding = True
        try:
            self.running = False
            if self.page:
                try:
                    await self.page.close()
                except Exception:
                    pass
            self.running = True
            await self.start()
            await self.inject_and_configure_all(focus_lock)
            print(f"[{self.tab_id}] [RECOVERY] Tab successfully restarted and re-configured.")
            self.last_heartbeat_ns = time.time_ns()
        except Exception as e:
            print(f"[{self.tab_id}] [RECOVERY ERROR] Failed to restart tab: {e}")
        finally:
            self.is_seeding = False

    async def inject_and_configure_all(self, focus_lock: asyncio.Lock):
        """Programmatic JS-based S9 indicator & symbol configuration"""
        print(f"[{self.tab_id}] Bringing tab to front...")
        await self.page.bring_to_front()
        await asyncio.sleep(0.5)
        
        # Wait for layout containers to render fully
        try:
            print(f"[{self.tab_id}] Waiting for layout containers to render...")
            await self.page.wait_for_selector("#tv_chart_container_win1, #tv_chart_container_main", state="attached", timeout=30000)
            await self.page.wait_for_selector("#tv_chart_container_win9", state="attached", timeout=30000)
            await asyncio.sleep(2.0)
        except Exception as e:
            print(f"[{self.tab_id}] [WARN] Timeout waiting for layout containers: {e}")

        print(f"[{self.tab_id}] Configuring symbols and indicators on grid layout via JS API...")
        for win_idx, symbol in enumerate(self.symbols, start=1):
            print(f"[{self.tab_id}] [Config] Configuring window {win_idx}/9 for {symbol}")
            container_id = f"tv_chart_container_win{win_idx}"
            selector = f"#{container_id}" if win_idx != 1 else f"#{container_id}, #tv_chart_container_main"
            container = self.page.locator(selector).first
            
            if await container.count() > 0:
                try:
                    iframe = container.locator("iframe").first
                    await iframe.wait_for(state="attached", timeout=15000)
                    iframe_handle = await iframe.element_handle(timeout=15000)
                    if iframe_handle:
                        frame = await iframe_handle.content_frame()
                        if frame:
                            # Evaluate programmatic setup script inside the iframe
                            res = await frame.evaluate(f'''async () => {{
                                try {{
                                    // 1. Wait for TradingView API and CoinGlass custom studies metadata cache to load
                                    let apiReady = false;
                                    for (let i = 0; i < 40; i++) {{
                                        if (typeof tradingViewApi !== 'undefined' && 
                                            tradingViewApi.activeChart && 
                                            tradingViewApi._chartApiInstance && 
                                            tradingViewApi._chartApiInstance._studyEngine && 
                                            tradingViewApi._chartApiInstance._studyEngine._metainfoCache) {{
                                            
                                            let cache = tradingViewApi._chartApiInstance._studyEngine._metainfoCache;
                                            let keys = Object.keys(cache);
                                            if (keys.some(k => cache[k].description && cache[k].description.includes('CoinGlass'))) {{
                                                apiReady = true;
                                                break;
                                            }}
                                        }}
                                        await new Promise(r => setTimeout(r, 500));
                                    }}
                                    
                                    if (!apiReady) {{
                                        return {{ success: false, error: 'CoinGlass indicator metadata cache not ready' }};
                                    }}
                                    
                                    // 2. Set Symbol
                                    tradingViewApi.changeSymbol("Binance_{symbol}");
                                    
                                    // 3. Set Timeframe to 15m
                                    if (typeof chartWidgetCollection !== 'undefined') {{
                                        chartWidgetCollection.setResolution('15');
                                    }}
                                    
                                    // 4. Scan and inject missing indicators
                                    let ac = tradingViewApi.activeChart();
                                    if (ac) {{
                                         let existing = [];
                                         try {{
                                             existing = ac.getAllStudies() || [];
                                         }} catch (err) {{}}
                                         
                                         let hasVolume = existing.some(s => s.name && s.name.includes('Volume') && !s.name.includes('Delta'));
                                         let hasFutCVD = existing.some(s => s.name && s.name.includes('Futures Cumulative Volume Delta'));
                                         let hasSpotCVD = existing.some(s => s.name && s.name.includes('Spot Cumulative Volume Delta'));
                                         let hasRSI = existing.some(s => s.name && s.name.includes('Relative Strength Index'));
                                         let hasFunding = existing.some(s => s.name && s.name.includes('Funding Rates'));
                                         let hasLiq = existing.some(s => s.name && s.name.includes('Aggregated Liquidations'));
                                         let hasLS = existing.some(s => s.name && s.name.includes('Long/Short Ratio'));
                                         let hasOI = existing.some(s => s.name && s.name.includes('Open Interest'));
                                         let hasWhale = existing.some(s => s.name && s.name.includes('Whale Index'));
                                         let hasTaker = existing.some(s => s.name && s.name.includes('Taker Buy/Sell Count'));
                                         
                                         let bidAsks = existing.filter(s => s.name && s.name.includes('Bid & Ask'));
                                         
                                         let injectedAny = false;
                                         if (!hasVolume) {{ ac.createStudy('Volume', false, false); injectedAny = true; }}
                                         if (!hasFutCVD) {{ ac.createStudy('<CoinGlass> Aggregated Futures Cumulative Volume Delta (CVD)', false, false); injectedAny = true; }}
                                         if (!hasSpotCVD) {{ ac.createStudy('<CoinGlass> Aggregated Spot Cumulative Volume Delta (CVD)', false, false); injectedAny = true; }}
                                         if (!hasRSI) {{ ac.createStudy('Relative Strength Index', false, false); injectedAny = true; }}
                                         if (!hasFunding) {{ ac.createStudy('<CoinGlass> Funding Rates(Open Interest Weighted,Candles)', false, false); injectedAny = true; }}
                                         if (!hasLiq) {{ ac.createStudy('<CoinGlass> Aggregated Liquidations ', false, false); injectedAny = true; }}
                                         if (!hasLS) {{ ac.createStudy('<CoinGlass> Long/Short Ratio (Accounts)', false, false); injectedAny = true; }}
                                         if (!hasOI) {{ ac.createStudy('<CoinGlass> Aggregated Open Interest(STABLECOIN-margined,Candles)', false, false); injectedAny = true; }}
                                         if (!hasWhale) {{ ac.createStudy('<CoinGlass> Whale Index', false, false); injectedAny = true; }}
                                         if (!hasTaker) {{ ac.createStudy('<CoinGlass> Taker Buy/Sell Count', false, false); injectedAny = true; }}
                                         
                                         if (bidAsks.length < 2) {{
                                             for (let b of bidAsks) {{
                                                 try {{ ac.removeStudy(b.id); }} catch(e) {{}}
                                             }}
                                             ac.createStudy('<CoinGlass> Aggregated Futures Bid & Ask ', false, false, {{ "Depth": 1, "symbol": "Main chart symbol", "Measure": "Coins" }});
                                             ac.createStudy('<CoinGlass> Aggregated Futures Bid & Ask ', false, false, {{ "Depth": 1, "symbol": "Main chart symbol", "Measure": "Dollars" }});
                                             injectedAny = true;
                                         }}
                                         
                                         // Force save layout if anything was injected
                                         if (injectedAny && tradingViewApi._saveChartService && typeof tradingViewApi._saveChartService.saveChart === 'function') {{
                                             try {{
                                                 tradingViewApi._saveChartService.saveChart();
                                             }} catch(se) {{}}
                                         }}
                                         
                                         let dump = ac.getAllStudies().map(s => ({{id: s.id, name: s.name}}));
                                         return {{ success: true, dump: dump }};
                                    }}
                                    return {{ success: false, error: 'Active chart not found' }};
                                }} catch (e) {{
                                    return {{ success: false, error: e.message }};
                                }}
                            }}''')
                            if res and "dump" in res:
                                try:
                                    with open(os.path.join(base_dir, "Seeding", f"studies_{self.tab_id}_{symbol}.json"), "w") as f:
                                        json.dump(res["dump"], f, indent=2)
                                except Exception: pass
                                
                            if not res or not res.get("success"):
                                print(f"[{self.tab_id}] [WARN] Programmatic setup failed for {symbol}: {res.get('error') if res else 'Unknown'}")
                            else:
                                print(f"[{self.tab_id}] [Config] Symbol & Indicators verified/configured for {symbol}")
                except Exception as e:
                    print(f"[{self.tab_id}] [WARN] Error configuring window {win_idx} for {symbol}: {e}")
            await asyncio.sleep(0.1)

        # Wait for studies to load data from network
        print(f"[{self.tab_id}] Waiting 15 seconds for TradingView studies to load historical data...")
        await asyncio.sleep(15.0)

        try:
            await self.page.screenshot(path=os.path.join(base_dir, "Seeding", f"{self.tab_id}_layout.png"))
        except Exception as e:
            print(f"[{self.tab_id}] [WARN] Screenshot failed: {e}")
        print(f"[{self.tab_id}] Setup & Indicator injection complete.")

    async def poll_loop(self) -> None:
        """Background data poller extracting DOM legend values & JS shims"""
        async def _fetch_frame(win_idx: int) -> bool:
            sym = self.symbols[win_idx - 1]
            container_id = f"tv_chart_container_win{win_idx}"
            selector = f"#{container_id}" if win_idx != 1 else f"#{container_id}, #tv_chart_container_main"
            container = self.page.locator(selector).first
            
            if await container.count() > 0:
                iframe = container.locator("iframe").first
                if await iframe.count() > 0:
                    iframe_handle = await iframe.element_handle(timeout=10000)
                    frame = await iframe_handle.content_frame() if iframe_handle else None
                    if frame:
                        try:
                            res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                        except Exception as eval_exc:
                            print(f"[{self.tab_id}] [POLL ERROR] {sym} frame eval: {eval_exc}")
                            return False
                        if res and res.get("success"):
                            d = res["data"]
                            sym_actual = (d.get("symbol") or "").strip().upper()
                            # Check actual extracted symbol matches what we expect
                            if sym_actual and sym_actual in self.symbols:
                                target_sym = sym_actual
                            else:
                                target_sym = sym
                                
                            price_val = parse_float(d.get("close") or d.get("price") or 0.0)
                            rsi_val = parse_float(d.get("rsi", 0.0))
                            if rsi_val in (100.0, 0.0):
                                rsi_val = self.store._data.get(target_sym, AssetSnapshot(symbol=target_sym)).rsi
                            await self.store.update(
                                target_sym,
                                source="coinglass",
                                price=price_val,
                                volume=parse_float(d.get("volume", 0.0)),
                                rsi=rsi_val,
                                fut_cvd=parse_float(d.get("futures_cvd", 0.0)),
                                spot_cvd=parse_float(d.get("spot_cvd", 0.0)),
                                funding=parse_float(d.get("funding_rate", 0.0)),
                                liq_long=parse_float(d.get("liquidations_long", 0.0)),
                                liq_short=parse_float(d.get("liquidations_short", 0.0)),
                                ls_ratio=parse_float(d.get("ls_ratio", 0.0)),
                                oi=parse_float(d.get("open_interest", 0.0)),
                                coins_bid=parse_float(d.get("coins_bid", 0.0)),
                                coins_ask=parse_float(d.get("coins_ask", 0.0)),
                                dollars_bid=parse_float(d.get("dollars_bid", 0.0)),
                                dollars_ask=parse_float(d.get("dollars_ask", 0.0)),
                                whale_idx=parse_float(d.get("whale_index", 0.0)),
                                tk_buy_cnt=parse_float(d.get("taker_buy_count", 0.0)),
                                tk_sell_cnt=parse_float(d.get("taker_sell_count", 0.0))
                            )
                            return True
            return False

        while self.running:
            try:
                results = await asyncio.gather(*[_fetch_frame(i) for i in range(1, 10)], return_exceptions=True)
                has_success = False
                for r in results:
                    if isinstance(r, Exception):
                        print(f"[{self.tab_id}] [POLL ERROR] Subtask failed: {r}")
                        self.poll_failures += 1
                    elif r is True:
                        has_success = True

                if has_success:
                    self.last_heartbeat_ns = time.time_ns()
                    self.poll_failures = 0
                else:
                    self.poll_failures += 1
            except Exception as e:
                print(f"[{self.tab_id}] [POLL ERROR] Outer: {e}")
                self.poll_failures += 1
            
            if self.poll_failures > 5:
                print(f"[{self.tab_id}] [WATCHDOG] Max failures exceeded ({self.poll_failures}). Auto-healing by reloading page...")
                try:
                    await self.page.reload(wait_until="load", timeout=30000)
                    self.poll_failures = 0
                except Exception as ex:
                    print(f"[{self.tab_id}] [WATCHDOG] Failed to reload page: {ex}")
                    self.poll_failures = 0

            await asyncio.sleep(0.5)

    async def _route_payload(self, entry: dict) -> None:
        url = entry.get("url", "")
        body = entry.get("body", "")
        try:
            payload = json.loads(body)
        except Exception:
            return
        
        # Route to appropriate update target
        if "open-interest" in url:
            await self._apply(payload, "oi")
        elif "funding-rate" in url:
            await self._apply(payload, "funding")
        elif "liquidation" in url:
            await self._apply_liq(payload)
        elif "long-short" in url:
            await self._apply(payload, "ls_ratio")
        elif "cumulative-volume" in url:
            if "futures" in url:
                await self._apply(payload, "fut_cvd")
            else:
                await self._apply(payload, "spot_cvd")
        elif "rsi" in url:
            await self._apply(payload, "rsi")

    async def _apply(self, payload: Any, field_name: str) -> None:
        data = payload.get("data", [])
        if isinstance(data, list):
            for row in data:
                sym = row.get("symbol")
                if sym in self.symbols:
                    val = parse_float(row.get("value", 0.0))
                    await self.store.update(sym, source="coinglass", **{field_name: val})
        elif isinstance(data, dict):
            for sym, val in data.items():
                if sym in self.symbols:
                    await self.store.update(sym, source="coinglass", **{field_name: parse_float(val)})

    async def _apply_liq(self, payload: Any) -> None:
        data = payload.get("data", [])
        if isinstance(data, list):
            for row in data:
                sym = row.get("symbol")
                if sym in self.symbols:
                    long_liq = parse_float(row.get("longLiq", 0.0))
                    short_liq = parse_float(row.get("shortLiq", 0.0))
                    await self.store.update(sym, source="coinglass", liq_long=long_liq, liq_short=short_liq)

    async def seed_symbol(self, symbol: str, excel_executor, focus_lock: asyncio.Lock) -> None:
        """Performs visual backward walk to collect 50 candles and export to Excel"""
        self.is_seeding = True
        win_idx = self.symbols.index(symbol) + 1
        container_id = f"tv_chart_container_win{win_idx}"
        selector = f"#{container_id}" if win_idx != 1 else f"#{container_id}, #tv_chart_container_main"
        container = self.page.locator(selector).first
        
        async with focus_lock:
            print(f"[{self.tab_id}] Seeding {symbol} in Window {win_idx}. Acquired focus lock. Bringing tab to front...")
            await self.page.bring_to_front()
            await asyncio.sleep(0.5)
            
            iframe = container.locator("iframe").first
            try:
                await iframe.wait_for(state="attached", timeout=15000)
            except Exception:
                pass
            iframe_handle = await iframe.element_handle(timeout=10000)
            if not iframe_handle:
                print(f"[{self.tab_id}] [ERROR] No iframe handle for seeding {symbol}")
                return
            frame = await iframe_handle.content_frame()
            if not frame:
                print(f"[{self.tab_id}] [ERROR] Content frame missing for seeding {symbol}")
                return
                
            # Resolve the first canvas inside the frame
            canvas = frame.locator("canvas").first
            try:
                await canvas.wait_for(state="visible", timeout=5000)
            except Exception:
                print(f"[{self.tab_id}] [ERROR] Canvas element not visible for {symbol}")
                return
                
            # Click canvas center to focus TradingView inner context
            await canvas.click(force=True, timeout=5000)
            await asyncio.sleep(0.3)
            
            # Explicitly focus the window/document body
            await frame.evaluate("() => { window.focus(); if (document.body) document.body.focus(); }")
            await asyncio.sleep(0.2)
            
            # Press Escape to close any potential dialogs
            await self.page.keyboard.press("Escape")
            await asyncio.sleep(0.2)
            
            # Reset visual using Alt+r
            await self.page.keyboard.press("Alt+r")
            await asyncio.sleep(1.0)
            
            # Wait for canvas to become visible/attached again after chart reset
            try:
                await canvas.wait_for(state="visible", timeout=5000)
            except Exception:
                pass
            
            # Right-click canvas to open context menu (forces browser focus delegation)
            await canvas.click(button="right", force=True, timeout=5000)
            await asyncio.sleep(0.5)
            await self.page.keyboard.press("Escape")
            await asyncio.sleep(0.2)
            
            # Wait up to 10 seconds for indicators to load historical data from network
            print(f"[{self.tab_id}] waiting for indicators to populate historical data for {symbol}...")
            for attempt in range(20):
                res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                if res and res.get("success"):
                    d = res["data"]
                    if (d.get("volume") not in ("N/A", "0", None) and
                        d.get("rsi") not in ("N/A", "100.00", None) and
                        d.get("futures_cvd") != "N/A" and
                        d.get("spot_cvd") != "N/A" and
                        d.get("open_interest") != "N/A"):
                        print(f"[{self.tab_id}] Indicators populated in {attempt * 0.5:.1f}s")
                        break
                await asyncio.sleep(0.5)

            rect = await canvas.bounding_box()
            if not rect:
                print(f"[{self.tab_id}] [ERROR] Cannot get canvas bounding box for {symbol}")
                return

            x_pos = rect["x"] + rect["width"] - 60
            y_pos = rect["y"] + rect["height"] * 0.5

            # Hover and click to focus on the rightmost section of the canvas
            await self.page.mouse.move(x_pos, y_pos)
            await self.page.mouse.click(x_pos, y_pos)
            await asyncio.sleep(0.2)

            # ArrowLeft snaps crosshair to the latest candle
            await self.page.keyboard.press("ArrowLeft")
            await asyncio.sleep(0.3)

            # --- Dynamic Gap Calculation ---
            target_steps = 850
            existing_rows = []
            base_dir = os.path.dirname(os.path.abspath(__file__))
            combined_path = os.path.join(base_dir, "Seeding", "combined_seed_history.xlsx")
            if os.path.exists(combined_path):
                import pandas as pd
                try:
                    df = pd.read_excel(combined_path, sheet_name=symbol)
                    if not df.empty and "open_time" in df.columns:
                        existing_rows = df.to_dict('records')
                        
                        # Handle potential datetime vs int vs str timestamp differences
                        for r in existing_rows:
                            val = r.get("open_time")
                            if hasattr(val, "timestamp"):
                                from datetime import timezone
                                r["open_time"] = int(val.replace(tzinfo=timezone.utc).timestamp())
                            elif isinstance(val, (int, float)):
                                r["open_time"] = int(val)
                            elif isinstance(val, str):
                                try:
                                    val_clean = val.replace(" IST", "").strip()
                                    dt = pd.to_datetime(val_clean)
                                    from datetime import timedelta
                                    dt_utc = dt - timedelta(hours=5, minutes=30)
                                    r["open_time"] = int(dt_utc.timestamp())
                                except Exception:
                                    try:
                                        r["open_time"] = int(float(val))
                                    except Exception:
                                        pass
                        
                        # Filter out invalid open_times for max calc
                        valid_times = [r["open_time"] for r in existing_rows if isinstance(r.get("open_time"), int)]
                        if valid_times:
                            latest_time = max(valid_times)
                            current_time = int((time.time() // 900) * 900)
                            gap_candles = calculate_commodity_gap(symbol, latest_time, current_time)
                            existing_count = len(existing_rows)
                            
                            if existing_count + gap_candles >= 850:
                                target_steps = min(gap_candles + 2, 850)
                            else:
                                target_steps = 850
                                
                            print(f"\n==================================================")
                            print(f"[{self.tab_id}] {symbol} SEEDING DATABASE CHECK:")
                            print(f"[{self.tab_id}] Database has {existing_count} candles.")
                            print(f"[{self.tab_id}] Gap from offline time: {gap_candles} missing candles (calendar adjusted).")
                            print(f"[{self.tab_id}] -> WILL SCRAPE {target_steps} CANDLES NOW TO WARM UP.")
                            print(f"==================================================\n")
                        else:
                            print(f"\n==================================================")
                            print(f"[{self.tab_id}] {symbol} Found Excel sheet but no valid `open_time` ints parsed.")
                            print(f"[{self.tab_id}] -> WILL SCRAPE {target_steps} CANDLES NOW TO WARM UP.")
                            print(f"==================================================\n")
                except Exception as e:
                    print(f"[{self.tab_id}] [WARN] Could not read existing seed for {symbol}: {e}")
            else:
                print(f"\n==================================================")
                print(f"[{self.tab_id}] {symbol} No existing seed history found in Excel.")
                print(f"[{self.tab_id}] -> WILL SCRAPE {target_steps} CANDLES NOW TO WARM UP.")
                print(f"==================================================\n")
            # -------------------------------

            candles = collections.deque(maxlen=1000)
            stalls = 0
            debug_dicts = []
            
            last_key = None
            is_crypto = symbol not in ["XAUUSDT", "XAGUSDT", "CLUSDT", "NATGASUSDT"]
            
            computed_timestamps = get_historical_timestamps(symbol, int((time.time() // 900) * 900), target_steps)

            for step in range(target_steps * 2):
                if len(candles) >= target_steps:
                    print(f"[{self.tab_id}] {symbol} Reached target {target_steps} candles. Stopping walk.")
                    break
                    
                if step % 20 == 0:
                    print(f"[{self.tab_id}] Seeding {symbol}: candle {len(candles)}/{target_steps}...")
                
                # 1. Wait for DOM to update after moving crosshair (avoid duplicate read stalls)
                d = None
                for attempt in range(6):
                    res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                    if res and res.get("success"):
                        temp_d = res["data"]
                        close = parse_float(temp_d.get("close", temp_d.get("price", 0.0)))
                        volume = parse_float(temp_d.get("volume", 0.0))
                        rsi = parse_float(temp_d.get("rsi", 50.0))
                        val_key = (close, volume, rsi)
                        
                        if val_key != last_key:
                            d = temp_d
                            last_key = val_key
                            break
                    await asyncio.sleep(0.04)

                if d is None:
                    stalls += 1
                    if stalls > 4:
                        print(f"[{self.tab_id}] [WARN] Seeding stalled for {symbol} at step {step}. Ending early.")
                        break
                    # Recover visual focus delegation
                    await canvas.focus()
                    await asyncio.sleep(0.05)
                    await self.page.keyboard.press("ArrowLeft")
                    await asyncio.sleep(0.1)
                    continue
                
                stalls = 0
                
                # 2. Wait up to 600ms for lazy-loaded indicators (CVD, OI) to populate if they are currently N/A
                if is_crypto:
                    for load_attempt in range(4):
                        if (d.get("futures_cvd") != "N/A" and 
                            d.get("spot_cvd") != "N/A" and 
                            d.get("open_interest") != "N/A"):
                            break
                        await asyncio.sleep(0.15)
                        res = await frame.evaluate(SINGLE_FRAME_EXTRACTION_JS)
                        if res and res.get("success"):
                            d = res["data"]
                            
                if symbol == "BTCUSDT":
                    debug_dicts.append({
                        "step": step,
                        "data": d,
                        "rawLegends": res.get("rawLegends", []) if res else []
                    })
                    
                candle_data = {
                    "open_time": computed_timestamps[len(candles)],
                    "open":       parse_float(d.get("open",   0.0)),
                    "high":       parse_float(d.get("high",   0.0)),
                    "low":        parse_float(d.get("low",    0.0)),
                    "close":      parse_float(d.get("close",  d.get("price", 0.0))),
                    "volume":     parse_float(d.get("volume", 0.0)),
                    "rsi":        parse_float(d.get("rsi",    50.0)),
                    "fut_cvd":    parse_float(d.get("futures_cvd",      0.0)),
                    "spot_cvd":   parse_float(d.get("spot_cvd",         0.0)),
                    "funding":    parse_float(d.get("funding_rate",      0.0)),
                    "liq_long":   abs(parse_float(d.get("liquidations_long",  0.0))),
                    "liq_short":  abs(parse_float(d.get("liquidations_short", 0.0))),
                    "ls_ratio":   parse_float(d.get("ls_ratio",           1.0)),
                    "oi":         parse_float(d.get("open_interest",      0.0)),
                    "coins_bid":  parse_float(d.get("coins_bid", 0.0)),
                    "coins_ask":  parse_float(d.get("coins_ask", 0.0)),
                    "dollars_bid": parse_float(d.get("dollars_bid", 0.0)),
                    "dollars_ask": parse_float(d.get("dollars_ask", 0.0)),
                    "whale_idx":  parse_float(d.get("whale_index", 0.0)),
                    "tk_buy_cnt": parse_float(d.get("taker_buy_count", 0.0)),
                    "tk_sell_cnt": parse_float(d.get("taker_sell_count", 0.0)),
                }
                
                candles.appendleft(candle_data)
                
                # Step left — move crosshair one candle back
                await self.page.keyboard.press("ArrowLeft")
                await asyncio.sleep(0.08)

            # Restore view
            await self.page.keyboard.press("Alt+r")
            await asyncio.sleep(0.5)

            scraped_list = list(candles)
            final_list = scraped_list
            if existing_rows:
                all_data = existing_rows + scraped_list
                # Deduplicate by open_time, keeping newest (scraped over existing due to order)
                dedup = {r["open_time"]: r for r in all_data if isinstance(r.get("open_time"), int)}
                sorted_vals = sorted(dedup.values(), key=lambda x: x["open_time"])
                final_list = sorted_vals

            loop = asyncio.get_running_loop()
            await loop.run_in_executor(excel_executor, _dump_xlsx, symbol, final_list)
            
            if self.store.predictor:
                self.store.predictor.set_history(symbol, final_list)
            
            if candles:
                last = list(candles)[-1]
                missing = [k for k, v in last.items() if v == 0.0 and k not in ("liq_long", "liq_short")]
                if missing:
                    print(f"[{self.tab_id}] [WARN] {symbol}: zero fields = {missing}")
                else:
                    print(f"[{self.tab_id}] [OK]   {symbol}: all fields populated (close={last['close']}, vol={last['volume']}, funding={last['funding']})")
                    
            if symbol == "BTCUSDT":
                try:
                    with open(os.path.join(base_dir, "Seeding", "seeding_debug_BTCUSDT.json"), "w", encoding="utf-8") as f:
                        json.dump(debug_dicts, f, indent=2)
                    await self.page.screenshot(path=os.path.join(base_dir, "Seeding", f"diag_{self.tab_id}_{symbol}.png"), clip={"x": 0, "y": 0, "width": 600, "height": 400})
                except Exception:
                    pass
            print(f"[{self.tab_id}] [Success] Seeded {symbol} with {len(candles)} candles.")

def fetch_binance_funding_rates(symbol: str) -> List[Dict[str, Any]]:
    import urllib.request
    import json
    url = f"https://fapi.binance.com/fapi/v1/fundingRate?symbol={symbol}&limit=100"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode())
    except Exception as e:
        print(f"[Binance API] Failed to fetch funding rate for {symbol}: {e}")
        return []

def fetch_binance_open_interest(symbol: str) -> List[Dict[str, Any]]:
    import urllib.request
    import json
    url = f"https://fapi.binance.com/fapi/v1/openInterestHist?symbol={symbol}&period=15m&limit=120"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode())
    except Exception as e:
        print(f"[Binance API] Failed to fetch open interest for {symbol}: {e}")
        return []

def fetch_binance_ls_ratio(symbol: str) -> List[Dict[str, Any]]:
    import urllib.request
    import json
    url = f"https://fapi.binance.com/data/globalLongShortAccountRatio?symbol={symbol}&period=15m&limit=120"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            return json.loads(response.read().decode())
    except Exception as e:
        print(f"[Binance API] Failed to fetch long/short ratio for {symbol}: {e}")
        return []

def _dump_xlsx(symbol: str, rows: List[Dict[str, Any]]) -> None:
    crypto_symbols = {
        "BTCUSDT", "ETHUSDT", "XRPUSDT", "SOLUSDT", "BNBUSDT", 
        "DOGEUSDT", "ADAUSDT", "TRXUSDT", "LINKUSDT", "AVAXUSDT", 
        "SUIUSDT", "NEARUSDT", "DOTUSDT", "LTCUSDT"
    }
    
    if rows and symbol in crypto_symbols:
        # 1. Backfill funding rate if all zeros
        if all(r.get("funding", 0.0) == 0.0 for r in rows):
            api_rates = fetch_binance_funding_rates(symbol)
            if api_rates:
                api_rates.sort(key=lambda x: x["fundingTime"])
                for r in rows:
                    row_time_ms = r.get("open_time", 0) * 1000
                    matching_rate = 0.0
                    for item in api_rates:
                        if item["fundingTime"] <= row_time_ms:
                            matching_rate = float(item["fundingRate"])
                        else:
                            break
                    r["funding"] = matching_rate

        # 2. Backfill open interest if all zeros
        if all(r.get("oi", 0.0) == 0.0 for r in rows):
            api_oi = fetch_binance_open_interest(symbol)
            if api_oi:
                api_oi.sort(key=lambda x: x["timestamp"])
                for r in rows:
                    row_time_ms = r.get("open_time", 0) * 1000
                    matching_oi = 0.0
                    for item in api_oi:
                        if item["timestamp"] <= row_time_ms:
                            matching_oi = float(item["sumOpenInterest"])
                        else:
                            break
                    r["oi"] = matching_oi

        # 3. Backfill long/short ratio if all zeros/default
        if all(r.get("ls_ratio", 1.0) == 1.0 or r.get("ls_ratio", 1.0) == 0.0 for r in rows):
            api_ls = fetch_binance_ls_ratio(symbol)
            if api_ls:
                api_ls.sort(key=lambda x: x["timestamp"])
                for r in rows:
                    row_time_ms = r.get("open_time", 0) * 1000
                    matching_ls = 1.0
                    for item in api_ls:
                        if item["timestamp"] <= row_time_ms:
                            matching_ls = float(item["longShortRatio"])
                        else:
                            break
                    r["ls_ratio"] = matching_ls

    # 4. Apply general forward-fill and backward-fill for all numeric columns to handle scattered zeros
    if rows:
        fill_fields = [
            "open", "high", "low", "close", "volume", "rsi", "fut_cvd", "spot_cvd", "funding", "ls_ratio", "oi",
            "coins_bid", "coins_ask", "dollars_bid", "dollars_ask", "whale_idx", "tk_buy_cnt", "tk_sell_cnt"
        ]
        for field in fill_fields:
            non_zero_vals = [r.get(field, 0.0) for r in rows if r.get(field, 0.0) != 0.0]
            if non_zero_vals:
                # Forward fill
                last_val = non_zero_vals[0]
                for r in rows:
                    val = r.get(field, 0.0)
                    if val != 0.0:
                        last_val = val
                    else:
                        r[field] = last_val
                # Backward fill
                last_val = non_zero_vals[-1]
                for r in reversed(rows):
                    val = r.get(field, 0.0)
                    if val != 0.0:
                        last_val = val
                    else:
                        r[field] = last_val

    wb = Workbook()
    ws = wb.active
    ws.title = symbol[:31]
    
    headers = [
        "open_time", "open", "high", "low", "close", "volume", 
        "rsi", "fut_cvd", "spot_cvd", "funding", "liq_long", "liq_short", "ls_ratio", "oi",
        "coins_bid", "coins_ask", "dollars_bid", "dollars_ask", "whale_idx", "tk_buy_cnt", "tk_sell_cnt"
    ]
    
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    CENTER = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        
    for row_idx, r in enumerate(rows, start=2):
        row_vals = []
        for h in headers:
            val = r.get(h, "")
            if h == "open_time" and isinstance(val, (int, float)):
                from datetime import datetime, timezone, timedelta
                tz_ist = timezone(timedelta(hours=5, minutes=30))
                val = datetime.fromtimestamp(val, tz=tz_ist).strftime("%Y-%m-%d %H:%M:%S IST")
            row_vals.append(val)
        ws.append(row_vals)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CENTER
            cell.border = BORDER
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    os.makedirs(os.path.join(base_dir, "Seeding"), exist_ok=True)
    filename = os.path.join(base_dir, "Seeding", f"{symbol}_seed_history.xlsx")
    try:
        wb.save(filename)
    except PermissionError:
        import random
        alt_filename = os.path.join(base_dir, "Seeding", f"{symbol}_seed_history_{random.randint(1000, 9999)}.xlsx")
        print(f"[WARN] Permission denied on {filename} (probably open in Excel). Saving to {alt_filename} instead.")
        try:
            wb.save(alt_filename)
        except Exception as e:
            print(f"[ERROR] Failed to save fallback Excel for {symbol}: {e}")

# --- DASHBOARD RENDERER ---
def render_table(snap: Dict[str, AssetSnapshot], trade_tracker: Any = None) -> Any:
    t = Table(title="Coinglass + Binance Footprint Scraper Terminal", expand=True)
    cols = (
        "Symbol", "Price", "RSI", "FutCVD", "SpotCVD", "LiqL", "LiqS", "Fund", "LSR", "OI", 
        "CoinsB", "CoinsA", "USDB", "USDA", "Whale", "BuyC", "SellC", "FP_D", "FP_P", "ARM"
    )
    for col in cols:
        t.add_column(col, justify="center", no_wrap=True)
        
    now = time.time_ns()
    
    def fmt(v: float, fresh: bool, col_type: str = "generic") -> str:
        if v is None:
            return "[dim]--[/dim]"
        if col_type == "rsi":
            s = f"{v:.2f}"
        elif col_type == "fund":
            s = f"{v:+.6f}"
        elif col_type in ("cvd", "fp_d"):
            s = f"{v:+,.2f}"
            if abs(v) > 1e6:
                s = f"{v:+,.0f}"
        else:
            s = f"{v:,.2f}"
            if abs(v) > 1e6 and col_type not in ("price", "rsi", "fund", "lsr"):
                s = f"{v:,.0f}"
        return s if fresh else f"[red]{s}[/red]"

    for sym in ALL_SYMBOLS:
        a = snap.get(sym, AssetSnapshot(symbol=sym))
        fresh = (now - a.ts_ns) < STALE_NS
        
        t.add_row(
            sym,
            fmt(a.price, fresh, "price"),
            fmt(a.rsi, fresh, "rsi"),
            fmt(a.fut_cvd, fresh, "cvd"),
            fmt(a.spot_cvd, fresh, "cvd"),
            fmt(a.liq_long, fresh, "liq"),
            fmt(a.liq_short, fresh, "liq"),
            fmt(a.funding, fresh, "fund"),
            fmt(a.ls_ratio, fresh),
            fmt(a.oi, fresh),
            fmt(a.coins_bid, fresh),
            fmt(a.coins_ask, fresh),
            fmt(a.dollars_bid, fresh),
            fmt(a.dollars_ask, fresh),
            fmt(a.whale_idx, fresh),
            fmt(a.tk_buy_cnt, fresh),
            fmt(a.tk_sell_cnt, fresh),
            fmt(a.fp_delta, fresh),
            fmt(a.fp_poc, fresh),
            f"[green]{a.strategy_armed}[/green]" if a.strategy_armed else "[dim]--[/dim]"
        )

    if trade_tracker is None:
        return t

    from rich.console import Group
    from rich.panel import Panel
    from rich.text import Text

    stats = trade_tracker.get_stats()
    
    # Active trades section
    active_lines = []
    with trade_tracker.lock:
        active_snap = list(trade_tracker.active_trades.values())
        history_snap = list(trade_tracker.history[-3:])

    for tr in active_snap:
        dir_str = "[bold green]LONG[/]" if tr['direction'] == 1 else "[bold red]SHORT[/]"
        pnl_usd = tr.get('live_pnl_usd', 0.0)
        pnl_pct = tr.get('live_pnl_pct', 0.0)
        pnl_str = f"[bold green]+${pnl_usd:.2f} (+{pnl_pct:+.2f}%)[/]" if pnl_usd >= 0 else f"[bold red]-${abs(pnl_usd):.2f} ({pnl_pct:+.2f}%)[/]"
        mt5_info = f" | MT5 Entry: {tr['mt5_entry']:.4f} (Lot: {tr['mt5_lot']:.2f})" if 'mt5_entry' in tr else ""
        active_lines.append(f"{tr['symbol']} | {dir_str} | Entry: {tr['entry_price']:.4f} | SL: {tr['sl']:.4f} | TP: {tr['tp']:.4f} | Live PnL: {pnl_str}{mt5_info}")

    active_text = "\n".join(active_lines) if active_lines else "[dim]No active trades[/dim]"

    # History trades section (last 3)
    history_lines = []
    for tr in history_snap:
        dir_str = "[bold green]LONG[/]" if tr['direction'] == 1 else "[bold red]SHORT[/]"
        pnl_usd = tr.get('pnl_usd', 0.0)
        pnl_pct = tr.get('pnl_pct', 0.0)
        pnl_str = f"[bold green]+${pnl_usd:.2f} (+{pnl_pct:+.2f}%)[/]" if pnl_usd >= 0 else f"[bold red]-${abs(pnl_usd):.2f} ({pnl_pct:+.2f}%)[/]"
        reason = tr.get('exit_reason', 'EXIT')
        history_lines.append(f"{tr['symbol']} | {dir_str} | Exit: {tr['exit_price']:.4f} | Reason: {reason} | Final: {pnl_str}")

    history_text = "\n".join(history_lines) if history_lines else "[dim]No trade history[/dim]"

    # Stats string
    winrate = stats['winrate']
    total_pnl = stats['total_pnl_usd']
    pnl_pct = total_pnl / trade_tracker.initial_capital * 100.0 if trade_tracker.initial_capital > 0 else 0.0
    pnl_clr = "green" if total_pnl >= 0 else "red"
    pnl_sign = "+" if total_pnl >= 0 else ""

    stats_text = (
        f"Initial Capital: [bold]${trade_tracker.initial_capital:,.2f}[/]  |  Current Capital: [bold]${stats['current_capital']:.2f}[/]  |  "
        f"Total PnL: [bold {pnl_clr}]{pnl_sign}${total_pnl:.2f} ({pnl_pct:+.2f}%)[/]  |  "
        f"Trades: [bold]{stats['total']}[/]  |  Winrate: [bold]{winrate:.1f}%[/]"
    )

    trade_table = Table(show_header=True, header_style="bold bright_magenta", border_style="magenta", expand=True)
    trade_table.add_column("Active Trades", justify="left", ratio=1)
    trade_table.add_column(stats_text, justify="left", ratio=1)
    trade_table.add_row(active_text, history_text)

    return Group(t, trade_table)

async def renderer_loop(store: SnapshotStore, stop: asyncio.Event) -> None:
    console = Console()
    loop_cnt = 0
    with Live(render_table(store.snapshot(), store.trade_tracker), console=console, refresh_per_second=REFRESH_HZ, screen=True) as live:
        while not stop.is_set():
            snap = store.snapshot()
            live.update(render_table(snap, store.trade_tracker))
            
            loop_cnt += 1
            if loop_cnt % 20 == 0:  # Every 10 seconds at 2Hz REFRESH_HZ
                try:
                    serializable_snap = {}
                    for sym, a in snap.items():
                        serializable_snap[sym] = {
                            "price": a.price, "volume": a.volume, "rsi": a.rsi, "fut_cvd": a.fut_cvd, "spot_cvd": a.spot_cvd,
                            "liq_long": a.liq_long, "liq_short": a.liq_short, "funding": a.funding,
                            "ls_ratio": a.ls_ratio, "oi": a.oi,
                            "coins_bid": a.coins_bid, "coins_ask": a.coins_ask,
                            "dollars_bid": a.dollars_bid, "dollars_ask": a.dollars_ask,
                            "whale_idx": a.whale_idx, "tk_buy_cnt": a.tk_buy_cnt, "tk_sell_cnt": a.tk_sell_cnt,
                            "fp_delta": a.fp_delta, "fp_poc": a.fp_poc,
                            "strategy_armed": a.strategy_armed, "ts_ns": a.ts_ns
                        }
                    def _write_debug():
                        try:
                            tmp_path = os.path.join(base_dir, "Seeding", "snapshot_debug.json.tmp")
                            with open(tmp_path, "w", encoding="utf-8") as f:
                                json.dump(serializable_snap, f, indent=4)
                            os.replace(tmp_path, os.path.join(base_dir, "Seeding", "snapshot_debug.json"))
                        except Exception:
                            pass
                    await asyncio.to_thread(_write_debug)
                except Exception:
                    pass
            await asyncio.sleep(1.0 / REFRESH_HZ)

# --- WATCHDOG ---
async def watchdog(components: List[Any], focus_lock: asyncio.Lock, stop: asyncio.Event) -> None:
    # Initialize/reset heartbeats for all components on startup to ignore the configuration time
    now_start = time.time_ns()
    for c in components:
        if hasattr(c, 'last_heartbeat_ns'):
            c.last_heartbeat_ns = now_start

    tab_tasks = {}
    for c in components:
        if isinstance(c, CoinglassTab):
            tab_tasks[c] = asyncio.create_task(c.poll_loop())
            
    try:
        while not stop.is_set():
            now = time.time_ns()
            for c in components:
                if hasattr(c, 'last_heartbeat_ns') and now - c.last_heartbeat_ns > 90_000_000_000:
                    if getattr(c, 'skip_watchdog', False):
                        continue
                    print(f"[Watchdog] [WARN] Subsystem '{c.__class__.__name__}' ({getattr(c, 'tab_id', 'Unknown')}) hung. Heartbeat silent >90s.")
                    if isinstance(c, CoinglassTab):
                        print(f"[Watchdog] [RECOVERY] Attempting recovery for '{c.tab_id}'...")
                        if c in tab_tasks and not tab_tasks[c].done():
                            tab_tasks[c].cancel()
                            try:
                                await tab_tasks[c]
                            except asyncio.CancelledError:
                                pass
                        try:
                            await c.reconnect(focus_lock)
                            tab_tasks[c] = asyncio.create_task(c.poll_loop())
                            # Reset heartbeats for all components to prevent false positives from the blocking recovery
                            now_after = time.time_ns()
                            for comp in components:
                                if hasattr(comp, 'last_heartbeat_ns'):
                                    comp.last_heartbeat_ns = now_after
                        except Exception as rec_err:
                            print(f"[Watchdog] [ERROR] Recovery failed for '{c.tab_id}': {rec_err}")
            # Check Python process memory usage to catch memory leaks
            mem_mb = get_process_memory_usage() / (1024 * 1024)
            if mem_mb > 3584.0:  # 3.5 GB limit to allow initial retraining/seeding spikes
                print(f"\n[Watchdog] [ALERT] [MEMORY] Python memory usage is extremely high ({mem_mb:.1f} MB)!")
            await asyncio.sleep(5.0)
    finally:
        for task in tab_tasks.values():
            if not task.done():
                task.cancel()
        if tab_tasks:
            await asyncio.gather(*tab_tasks.values(), return_exceptions=True)

def combine_seeding_files() -> None:
    import glob
    import copy
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter

    files = glob.glob(os.path.join(base_dir, "Seeding", "*_seed_history.xlsx"))
    files = [f for f in files if "combined_seed" not in os.path.basename(f).lower()]
    if not files:
        print("[Setup] No seeding files found to combine.")
        return

    print(f"[Setup] Combining {len(files)} seeding files into a single workbook...")
    combined_wb = Workbook()
    default_sheet = combined_wb.active
    combined_wb.remove(default_sheet)

    for f in sorted(files):
        symbol = os.path.basename(f).replace("_seed_history.xlsx", "")
        try:
            wb = load_workbook(f)
            source_ws = wb.active
            target_ws = combined_wb.create_sheet(title=symbol[:31])

            for row in source_ws.iter_rows():
                for cell in row:
                    new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy.copy(cell.font)
                        new_cell.fill = copy.copy(cell.fill)
                        new_cell.border = copy.copy(cell.border)
                        new_cell.alignment = copy.copy(cell.alignment)
                        new_cell.number_format = cell.number_format

            for col in source_ws.columns:
                col_letter = get_column_letter(col[0].column)
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
            wb.close()
        except Exception as copy_exc:
            print(f"[Setup] [WARN] Failed to copy {symbol} sheet: {copy_exc}")

    combined_filename = os.path.join(base_dir, "Seeding", "combined_seed_history.xlsx")
    tmp_filename = combined_filename + ".tmp"
    try:
        combined_wb.save(tmp_filename)
        os.replace(tmp_filename, combined_filename)
        print(f"[Setup] Combined workbook saved successfully: {combined_filename}")
        
        # Clean up individual seed files
        for f in files:
            try:
                os.remove(f)
            except OSError:
                pass
        print("[Setup] Cleaned up individual symbol seeding files.")
    except Exception as e:
        print(f"[Setup] [WARN] Failed to save combined workbook: {e}")

# --- MAIN CONTROLLER ---
async def main(skip_seed: bool = False) -> None:
    print("=" * 60)
    print(f"  SYSTEM STARTUP - MODE: {EXECUTION_MODE}")
    print("  WARNING: NO REAL METATRADER 5 TRADE ORDERS WILL BE SENT")
    print("  TRADES ARE SIMULATED LOCALLY IN THE TRACKER FILE")
    print("=" * 60)

    # 0. Clear existing ML models to prevent conflicts before retraining
    print("[Setup] Clearing existing ML model files before retraining...")
    for sub in (ACTIVE_STRATEGY, 'Liquidation', 'ml_trend_pull'):
        m_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), sub, 'models')
        if os.path.exists(m_dir):
            for file in os.listdir(m_dir):
                if file.endswith(('.pkl', '.txt', '.json', '.cbm')):
                    try:
                        os.remove(os.path.join(m_dir, file))
                    except Exception as clear_err:
                        print(f"[Setup] [WARN] Could not remove old model file {file}: {clear_err}")

    # 0. Live Model Retraining on latest Parquet data
    print(f"[Setup] Running Live Model Retraining on latest Parquet data for {ACTIVE_STRATEGY}...")
    try:
        as_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ACTIVE_STRATEGY)
        if as_path not in sys.path:
            sys.path.insert(0, as_path)
        import importlib
        model_trainer_mod = importlib.import_module("model_trainer")
        # Ensure we reload the correct strategy module if it was previously loaded
        importlib.reload(model_trainer_mod)
        model_trainer_mod.train_models()
    except Exception as retrain_err:
        print(f"[Setup] [WARN] Failed to retrain {ACTIVE_STRATEGY} models: {retrain_err}")

    try:
        liq_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Liquidation')
        if liq_path not in sys.path:
            sys.path.append(liq_path)
        from train import train_all_symbols
        print("[Setup] Retraining ML Liquidation models on latest data...")
        train_all_symbols()
    except Exception as retrain_err:
        print(f"[Setup] [WARN] Failed to retrain ML Liquidation models: {retrain_err}")

    # Retrain ML_Trend_Pull models
    try:
        tp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ml_trend_pull')
        if tp_path not in sys.path:
            sys.path.insert(0, tp_path)
        import importlib
        sys.modules.pop('model_trainer', None)
        tp_trainer = importlib.import_module('model_trainer')
        importlib.reload(tp_trainer)
        print("[Setup] Retraining ML_Trend_Pull models on latest data...")
        tp_trainer.train_models()
    except Exception as retrain_err:
        print(f"[Setup] [WARN] Failed to retrain ML_Trend_Pull models: {retrain_err}")

    # Initialize LiveStrategyPredictor & load cached history
    predictor = LiveStrategyPredictor(ALL_SYMBOLS)
    predictor.load_history_from_disk()
    
    liquidation_predictor = LiveLiquidationPredictor(ALL_SYMBOLS)
    trend_pull_predictor = LiveTrendPullPredictor(ALL_SYMBOLS)

    # Warm up history from AlphaSqueezer's seeded disk data safely (Deepcopy)
    import copy
    for sym in ALL_SYMBOLS:
        if sym in predictor.candles_history:
            liquidation_predictor.candles_history[sym] = collections.deque(
                [copy.deepcopy(c) for c in predictor.candles_history[sym]], maxlen=1200
            )
            trend_pull_predictor.candles_history[sym] = collections.deque(
                [copy.deepcopy(c) for c in predictor.candles_history[sym]], maxlen=1200
            )
    print(f"[Setup] Warmed up ML Liquidation history deque with {len(liquidation_predictor.candles_history.get(ALL_SYMBOLS[0], []))} rows.")
    print(f"[Setup] Warmed up ML_Trend_Pull history deque with {len(trend_pull_predictor.candles_history.get(ALL_SYMBOLS[0], []))} rows.")

    trade_tracker = Engine1TradeTracker()
    liquidation_predictor.recent_capitals = [trade_tracker.current_capital]
    trade_tracker.on_close_callbacks.append(
        lambda strategy, capital: liquidation_predictor.record_closed_capital(capital)
        if strategy == "ML_Liquidation_Runner" else None
    )
    def run_retrain_proc():
        import sys
        import os
        import importlib
        base_dir = os.path.dirname(os.path.abspath(__file__))
        as_path = os.path.join(base_dir, ACTIVE_STRATEGY)
        liq_path = os.path.join(base_dir, 'Liquidation')
        if as_path not in sys.path:
            sys.path.insert(0, as_path)
        if liq_path not in sys.path:
            sys.path.append(liq_path)
            
        print(f"[Background Process] Starting Live Retraining for {ACTIVE_STRATEGY}...")
        try:
            model_trainer_mod = importlib.import_module("model_trainer")
            importlib.reload(model_trainer_mod)
            model_trainer_mod.train_models()
        except Exception as e:
            print(f"[Background Process] {ACTIVE_STRATEGY} retrain failed: {e}")
        try:
            from train import train_all_symbols
            train_all_symbols()
        except Exception as e:
            print(f"[Background Process] Liquidation retrain failed: {e}")
        try:
            tp_path = os.path.join(base_dir, 'ml_trend_pull')
            if tp_path not in sys.path:
                sys.path.insert(0, tp_path)
            sys.modules.pop('model_trainer', None)
            tp_trainer = importlib.import_module('model_trainer')
            importlib.reload(tp_trainer)
            tp_trainer.train_models()
        except Exception as e:
            print(f"[Background Process] ML_Trend_Pull retrain failed: {e}")
        print("[Background Process] Live Retraining finished.")

    def background_retrain_loop():
        import time
        import multiprocessing
        while True:
            # Sleep for 24 hours (86400 seconds)
            time.sleep(86400)
            print("[Background Thread] Launching 24hr Live Retraining Subprocess...")
            try:
                p = multiprocessing.Process(target=run_retrain_proc)
                p.start()
                p.join()
            except Exception as ex:
                print(f"[Background Thread] Subprocess retraining manager crashed: {ex}")

    import threading
    retrain_thread = threading.Thread(target=background_retrain_loop, daemon=True)
    retrain_thread.start()
    print("[Setup] Launched 24hr Background Retraining Manager Thread (Process-isolated).")

    store = SnapshotStore(ALL_SYMBOLS, predictor, liquidation_predictor, trade_tracker, trend_pull_predictor)
    stop = asyncio.Event()
    
    print("[Setup] Launching Chromium instance with persistent profile...")
    async with async_playwright() as pw:
        user_data_dir = os.path.join(os.getcwd(), "chrome_profile")
        ctx = await pw.chromium.launch_persistent_context(
            user_data_dir,
            headless=False,
            viewport={"width": 1920, "height": 1080},
            args=[
                "--disable-features=CalculateNativeWinOcclusion",
                "--disable-background-timer-throttling",
                "--start-maximized",
                "--remote-debugging-port=9223"
            ]
        )
        
        # 1. Performing Session Login first
        print("[Setup] Navigating to Coinglass Login...")
        login_page = await ctx.new_page()
        
        for attempt in range(3):
            try:
                await login_page.goto("https://www.coinglass.com/login", wait_until="load", timeout=45000)
                break
            except Exception as exc:
                print(f"[Setup] [WARN] Login navigation attempt {attempt+1} failed: {exc}")
                if attempt == 2:
                    raise exc
                await asyncio.sleep(5.0)
        await asyncio.sleep(5)
        
        os.makedirs(os.path.join(base_dir, "Seeding"), exist_ok=True)
        await login_page.screenshot(path=os.path.join(base_dir, "Seeding", "login_init.png"))
        
        email_input = login_page.locator("input[placeholder='Email']").first
        if await email_input.count() > 0:
            email = os.environ.get("COINGLASS_EMAIL")
            password = os.environ.get("COINGLASS_PASSWORD")
            if not email or not password:
                raise ValueError("Missing COINGLASS_EMAIL or COINGLASS_PASSWORD environment variables.")
            
            await email_input.click()
            await email_input.fill(email)
            await asyncio.sleep(0.3)

            pass_input = login_page.locator("input[placeholder='Password']").first
            await pass_input.click()
            await pass_input.fill(password)
            await asyncio.sleep(0.3)

            await login_page.screenshot(path=os.path.join(base_dir, "Seeding", "login_filled.png"))
            print("[Setup] Submitting login form...")

            # Try explicit button click first, fallback to JS click, last resort Enter key
            try:
                btn = login_page.locator("button:has-text('Login')").first
                if await btn.count() > 0:
                    await btn.wait_for(state="visible", timeout=5000)
                    await btn.click()
                else:
                    raise Exception("button not found via locator")
            except Exception:
                try:
                    await login_page.evaluate('''() => {
                        const b = Array.from(document.querySelectorAll('button'))
                            .find(el => el.textContent.trim() === 'Login');
                        if (b) b.click();
                    }''')
                except Exception:
                    # Most reliable: press Enter on password field
                    await pass_input.press("Enter")

            print("[Setup] Waiting for post-login redirect...")
            try:
                await login_page.wait_for_url(lambda url: "/login" not in url, timeout=20000)
                print("[Setup] Login successful — redirected away from /login.")
            except Exception:
                print("[Setup] [WARN] No redirect detected — may already be logged in or login failed.")
            await login_page.screenshot(path=os.path.join(base_dir, "Seeding", "login_after_submit.png"))
            print("[Setup] Waiting 5 seconds to ensure session cookies are fully persisted...")
            await asyncio.sleep(5.0)
        else:
            print("[Setup] Form inputs not detected, assuming session already active.")
        
        # 2. Open Scraping Tabs (while login page is open to hold session cache)
        tab1 = CoinglassTab(ctx, TAB1_SYMBOLS, store, "TAB_1")
        tab2 = CoinglassTab(ctx, TAB2_SYMBOLS, store, "TAB_2")
        binance = BinanceFootprintFeed(ALL_SYMBOLS, store)
        binance_ws = BinanceTradePriceWebSocketFeed(ALL_SYMBOLS, store)
        
        await asyncio.gather(tab1.start(), tab2.start())
        
        # Close login page now that layout tabs have initialized
        try:
            await login_page.close()
        except Exception:
            pass
        
        # 3. Configure grid symbols & indicators
        focus_lock = asyncio.Lock()
        await asyncio.gather(
            tab1.inject_and_configure_all(focus_lock),
            tab2.inject_and_configure_all(focus_lock)
        )

        # 4. Historical Seeding
        from concurrent.futures import ThreadPoolExecutor
        excel_pool = ThreadPoolExecutor(max_workers=4)

        if skip_seed:
            print("[Setup] --skip-seed flag active. Skipping historical seeding.")
        else:
            async def seed_wrapper(tab: CoinglassTab, sym: str):
                try:
                    for attempt in range(3):
                        try:
                            if not tab.page or tab.page.is_closed():
                                print(f"[{tab.tab_id}] [RECOVERY] Page closed on seeding attempt {attempt+1}. Reconnecting...")
                                await tab.reconnect(focus_lock)
                            await tab.seed_symbol(sym, excel_pool, focus_lock)
                            break
                        except Exception as e:
                            print(f"[Setup] [WARN] Seeding failed for {sym} (attempt {attempt+1}/3): {e}")
                            if "closed" in str(e).lower() or "navigation" in str(e).lower() or "locator" in str(e).lower() or "timeout" in str(e).lower():
                                try:
                                    await tab.reconnect(focus_lock)
                                except Exception as rec_err:
                                    print(f"[Setup] [ERROR] Failed to reconnect tab during seeding retry: {rec_err}")
                            if attempt == 2:
                                raise
                            await asyncio.sleep(3.0)
                finally:
                    tab.is_seeding = False

            seeding_tasks = [seed_wrapper(tab1, sym) for sym in TAB1_SYMBOLS] + \
                            [seed_wrapper(tab2, sym) for sym in TAB2_SYMBOLS]

            print("[Setup] Launching historical seeding...")
            await asyncio.gather(*seeding_tasks)
            print("[Setup] Seeding phase complete! Starting real-time feeds...")
            combine_seeding_files()
        
        # 5. Run Live feeds & Terminal display
        async def tab_switcher():
            active_tab = tab1
            while not stop.is_set():
                await asyncio.sleep(60.0)
                if stop.is_set():
                    break
                if tab1.is_seeding or tab2.is_seeding:
                    continue
                try:
                    try:
                        async with asyncio.timeout(3.0):
                            async with focus_lock:
                                if active_tab.page and not active_tab.page.is_closed():
                                    await active_tab.page.bring_to_front()
                    except asyncio.TimeoutError:
                        print(f"[Switcher] Warning: focus_lock timeout. Bypassing lock to force {active_tab.name} to front.")
                        if active_tab.page and not active_tab.page.is_closed():
                            await active_tab.page.bring_to_front()
                            
                    active_tab = tab2 if active_tab is tab1 else tab1
                except Exception as e:
                    print(f"[Switcher] Failed to switch to {active_tab.name}: {e}")

        async def rollover_watchdog(tracker, stop_event):
            while not stop_event.is_set():
                try:
                    tracker.update_day()
                    # Non-blocking MT5 position sync (prevents order-tracking drift)
                    if hasattr(tracker, "reconcile_with_mt5"):
                        await asyncio.to_thread(tracker.reconcile_with_mt5)
                except Exception as ex:
                    print(f"[Watchdog] [ERROR] Rollover watchdog failed: {ex}")
                await asyncio.sleep(30.0)  # tighter sync cadence for exit safety

        async def event_loop_monitor(stop_event: asyncio.Event, threshold_sec: float = 0.5) -> None:
            consecutive_blocks = 0
            while not stop_event.is_set():
                start = time.time()
                await asyncio.sleep(0.1)
                elapsed = time.time() - start - 0.1
                if elapsed > threshold_sec:
                    consecutive_blocks += 1
                    print(f"\n[ALERT] [LATENCY] Event loop blocked for {elapsed:.2f}s! Potential CPU-bound task in event loop. Consecutive count: {consecutive_blocks}")
                    if consecutive_blocks >= 5:
                        print("\n[Watchdog] [ALERT] [LATENCY_CRITICAL] Event loop blocked consecutively 5 times. Process is hung.")
                else:
                    consecutive_blocks = 0

        tasks = [
            asyncio.create_task(event_loop_monitor(stop)),
            asyncio.create_task(binance.run()),
            asyncio.create_task(binance_ws.run()),
            asyncio.create_task(renderer_loop(store, stop)),
            asyncio.create_task(watchdog([tab1, tab2, binance, binance_ws], focus_lock, stop)),
            asyncio.create_task(tab_switcher()),
            asyncio.create_task(rollover_watchdog(trade_tracker, stop))
        ]
        
        # Handle graceful exit triggers
        loop = asyncio.get_running_loop()
        def sig_handler():
            print("\n[Exit] Termination signal received. Stopping...")
            stop.set()
            tab1.running = False
            tab2.running = False
            binance.running = False
            
        for sig in (signal.SIGINT, signal.SIGTERM):
            try:
                loop.add_signal_handler(sig, sig_handler)
            except NotImplementedError:
                pass
                
        try:
            while not stop.is_set():
                await asyncio.sleep(1.0)
        except (KeyboardInterrupt, SystemExit):
            sig_handler()
        finally:
            print("[Setup] Cleaning up tasks and closing browser...")
            for t in tasks:
                t.cancel()
            await asyncio.gather(*tasks, return_exceptions=True)
            excel_pool.shutdown(wait=True)
            await ctx.close()
        
    print("[Exit] Shutdown complete.")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Coinglass + Binance Footprint Scraper")
    parser.add_argument("--skip-seed", action="store_true", help="Skip historical Excel seeding and go straight to live feeds")
    args = parser.parse_args()
    asyncio.run(main(skip_seed=args.skip_seed))
